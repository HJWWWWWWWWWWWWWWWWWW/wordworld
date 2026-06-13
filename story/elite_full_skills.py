"""
Elite full skill upgrade: 6-26 skills per elite based on level tier.
Same tier system as bosses, scaled down appropriately.
"""
from pathlib import Path
import openpyxl

WORKBOOK_PATH = Path("story/text_game_event_schema_v4.xlsx")
wb = openpyxl.load_workbook(WORKBOOK_PATH)

ws_skills = ws_enemies = None
for name in wb.sheetnames:
    if name.startswith("Skills_"): ws_skills = wb[name]
    if name.startswith("Enemies_"): ws_enemies = wb[name]

enemy_headers = [cell.value for cell in next(ws_enemies.iter_rows(min_row=1, max_row=1))]
skills_col = enemy_headers.index("Skills")

# Existing skill sets (already defined)
existing_skills = set()
for row in ws_skills.iter_rows(min_row=2):
    if row[0].value: existing_skills.add(str(row[0].value))

# ══════════════════════════════════════════════════════════
# SKILL TIERS (same as boss system, scaled per level)
# ══════════════════════════════════════════════════════════

TIER1_Lv10 = ["skill_body_fortify", "skill_sword_qi", "skill_palm_strike", "skill_stamina_save"]
TIER2_Lv20 = ["skill_qi_shield", "skill_qi_burst", "skill_defensive_stance", "skill_spirit_sense"]
TIER3_Lv30 = ["skill_shadow_step", "skill_iron_body", "skill_hundred_fists", "skill_battle_cry", "skill_counter_strike"]
TIER4_Lv45 = ["skill_qi_explosion", "skill_air_slash", "skill_soaring_sky", "skill_heavy_strike", "skill_meditate"]
TIER5_Lv60 = ["skill_mountain_splitter", "skill_self_regen", "skill_lightning_flash", "skill_soul_blast", "skill_battle_meditation"]
TIER6_Lv80 = ["skill_ancient_roar", "skill_mind_break", "skill_shadow_dance", "skill_time_space_lock"]

# Element pools
FIRE = ["skill_fire_spit", "skill_fire_serpent", "skill_fire_storm", "skill_inferno"]
ICE = ["skill_ice_breath", "skill_ice_dragon", "skill_ice_prison", "skill_absolute_zero"]
WIND = ["skill_wind_slash", "skill_wind_push", "skill_wind_explosion", "skill_gale_strike"]
EARTH = ["skill_earth_shake", "skill_earth_wall", "skill_earthquake", "skill_sand_king"]
THUNDER = ["skill_lightning_bolt", "skill_thunder_clap", "skill_thunder_rain", "skill_thunder_wrath"]
POISON = ["skill_venom_inject", "skill_toxic_mist", "skill_toxic_explosion", "skill_corrosive_acid"]
SOUL = ["skill_soul_absorption", "skill_soul_shock", "skill_soul_pain", "skill_soul_rend", "skill_life_drain"]
DRAGON = ["skill_dragon_breath_attack", "skill_dragon_roar", "skill_dragon_scale_armor", "skill_dragon_wing_slash"]
ANCIENT = ["skill_bloodline_pressure", "skill_ancient_seal_art", "skill_emperor_will"]

def assign_elite_skills(level, faction_pool, element_pools, unique_pool):
    """Assign 6-26 skills based on level tier, faction, and element."""
    lv = int(level)
    pool = list(unique_pool)

    # Faction skills
    for fs in faction_pool:
        pool.extend(fs[:2])  # First 2 faction skills

    # Element skills by level
    for ep in element_pools:
        if lv >= 15: pool.append(ep[0])
        if lv >= 30: pool.extend(ep[1:2])
        if lv >= 50: pool.extend(ep[2:3])
        if lv >= 75: pool.extend(ep[3:4])

    # Tiered generic skills
    if lv >= 10: pool.extend(TIER1_Lv10)
    if lv >= 20: pool.extend(TIER2_Lv20)
    if lv >= 30: pool.extend(TIER3_Lv30[:3])
    if lv >= 35: pool.extend(TIER3_Lv30[3:])
    if lv >= 45: pool.extend(TIER4_Lv45[:3])
    if lv >= 50: pool.extend(TIER4_Lv45[3:])
    if lv >= 60: pool.extend(TIER5_Lv60[:3])
    if lv >= 70: pool.extend(TIER5_Lv60[3:])
    if lv >= 80: pool.extend(TIER6_Lv80[:2])
    if lv >= 88: pool.extend(TIER6_Lv80[2:])

    # Deduplicate
    seen = set()
    result = []
    for s in pool:
        if s not in seen and s in existing_skills:
            seen.add(s)
            result.append(s)

    return result

# ══════════════════════════════════════════════════════════
# ELITE SKILL MAP - organized by faction/region
# ══════════════════════════════════════════════════════════

# Helper: elite (eid, level, faction_pool, element_pools, unique_pool)
ELITE_DATA = [
    # ── 沧澜帝国 Lv 10-28 ──
    ("elite_wolf_head_leader", 10, [], [], ["skill_beast_charge"]),
    ("elite_xiao_guard_captain", 12, [], [], ["skill_beast_charge"]),
    ("elite_wutan_guard_chief", 12, [], [], ["skill_defensive_stance"]),
    ("elite_wutan_arena_champ", 14, [], [], ["skill_arena_rage"]),
    ("elite_alchemist_guard", 15, [], [], ["skill_pill_boost"]),
    ("elite_alchemist_examiner", 16, [], [FIRE], ["skill_pill_boost", "skill_alchemy"]),
    ("elite_desert_merc_captain", 16, [], [EARTH], ["skill_beast_charge"]),
    ("elite_salt_gang_leader", 18, [], [], ["skill_beast_charge"]),
    ("elite_miteer_chief_guard", 20, [], [], []),
    ("elite_miteer_trade_master", 22, [], [], ["skill_pill_boost"]),
    ("elite_black_rock_gang_boss", 20, [], [], ["skill_heavy_strike"]),
    ("elite_snake_temple_guard", 18, [], [POISON], ["skill_snake_seal"]),
    ("elite_snake_temple_priest", 24, [], [POISON], ["skill_snake_seal", "skill_snake_eye"]),
    ("elite_desert_caravan_master", 20, [], [EARTH], []),
    ("elite_sand_pirate_chief", 22, [], [EARTH], ["skill_sand_king"]),
    ("elite_desert_ancient_guardian", 24, [], [EARTH, ANCIENT], ["skill_ancient_seal_art"]),
    ("elite_imperial_guard", 22, [], [], ["skill_defensive_stance"]),
    ("elite_imperial_court_mage", 26, [], [FIRE], ["skill_pill_boost"]),
    ("elite_nalan_elder", 24, [], [WIND], ["skill_cloud_wind_sword", "skill_wind_wall"]),
    ("elite_jia_ma_general", 26, [], [], ["skill_battle_cry", "skill_bloodline_pressure"]),
    ("elite_jia_ma_spy_master", 28, [], [], ["skill_shadow_strike"]),
    ("elite_jia_ma_arena_champion", 28, [], [], ["skill_arena_rage"]),

    # ── 青岚宗 Lv 18-28 ──
    ("elite_yunlan_elder", 18, [], [WIND], ["skill_cloud_wind_sword", "skill_wind_wall"]),
    ("elite_yunlan_enforcer", 20, [], [WIND], ["skill_cloud_wind_sword"]),
    ("elite_yunlan_guard_captain", 25, [], [WIND], ["skill_cloud_wind_sword", "skill_wind_wall"]),
    ("elite_yunlan_deacon", 28, [], [WIND], ["skill_cloud_wind_sword", "skill_wind_wall"]),
    ("elite_yunlan_ritual_master", 22, [], [WIND], ["skill_cloud_wind_sword", "skill_wind_slash"]),
    ("elite_yunlan_sword_master", 26, [], [WIND], ["skill_cloud_wind_sword", "skill_sword_qi"]),

    # ── 暗角域 Lv 22-36 ──
    ("elite_blood_sect_elder", 22, [], [], ["skill_blood_art", "skill_blood_transform"]),
    ("elite_black_alliance_captain", 24, [], [], ["skill_battle_cry"]),
    ("elite_eight_gates_enforcer", 26, [], [], ["skill_eight_gates_palm"]),
    ("elite_eight_gates_auctioneer", 28, [], [], ["skill_eight_gates_palm", "skill_underground_authority"]),
    ("elite_feng_city_guard", 28, [], [], ["skill_defensive_stance"]),
    ("elite_feng_city_spy", 30, [], [], ["skill_shadow_strike"]),
    ("elite_black_emperor_guard", 30, [], [], ["skill_defensive_stance"]),
    ("elite_black_emperor_advisor", 32, [], [], ["skill_pill_boost"]),
    ("elite_demon_valley_elder", 34, [], [FIRE, POISON], ["skill_fire_poison_dual"]),
    ("elite_black_domain_hunter", 32, [], [], ["skill_shadow_strike"]),
    ("elite_black_blood_merc", 36, [], [], ["skill_life_drain"]),
    ("elite_black_seal_auctioneer", 26, [], [], ["skill_underground_authority"]),
    ("elite_blood_refiner", 28, [], [], ["skill_blood_transform", "skill_life_drain"]),
    ("elite_plain_mercenary_chief", 24, [], [], ["skill_beast_charge"]),

    # ── 迦南学院 Lv 20-32 ──
    ("elite_canaan_exam_officer", 20, [], [], []),
    ("elite_strong_rank_10", 22, [], [], []),
    ("elite_strong_rank_5", 26, [], [], ["skill_battle_cry"]),
    ("elite_inner_enforcer", 25, [], [], ["skill_defensive_stance"]),
    ("elite_tower_trial_guard", 28, [], [FIRE], ["skill_skyfire_guard"]),
    ("elite_academy_instructor", 30, [], [], ["skill_pill_boost", "skill_battle_cry"]),
    ("elite_library_guardian", 30, [], [ANCIENT], ["skill_skyfire_guard", "skill_ancient_seal_art"]),
    ("elite_mission_hall_officer", 27, [], [], []),
    ("elite_inner_duel_referee", 32, [], [], ["skill_battle_cry"]),

    # ── 出云帝国 Lv 30-35 ──
    ("elite_poison_sect_elder", 30, [], [POISON], ["skill_toxic_mist"]),
    ("elite_poison_craft_master", 34, [], [POISON], ["skill_toxic_mist", "skill_pill_boost"]),
    ("elite_scorpion_gate_elder", 32, [], [POISON], ["skill_ten_thousand_scorpion_master"]),
    ("elite_scorpion_tamer_master", 35, [], [POISON], ["skill_ten_thousand_scorpion_master"]),
    ("elite_golden_goose_elder", 33, [], [WIND], ["skill_golden_goose_formation"]),
    ("elite_golden_goose_rider", 32, [], [WIND], ["skill_golden_goose_formation", "skill_gale_strike"]),
    ("elite_mulan_valley_elder", 33, [], [], ["skill_mulan_herb_art"]),
    ("elite_mulan_valley_herbalist", 34, [], [], ["skill_mulan_herb_art", "skill_pill_boost"]),
    ("elite_chuyun_imperial_guard", 35, [], [], ["skill_defensive_stance"]),

    # ── 黑渊殿 Lv 35-82 ──
    ("elite_soul_hall_protector", 35, [], [SOUL], ["skill_soul_handprint", "skill_soul_chain"]),
    ("elite_soul_hall_elder_protector", 42, [], [SOUL], ["skill_soul_handprint", "skill_soul_chain", "skill_soul_pain"]),
    ("elite_soul_hall_soul_reaper", 48, [], [SOUL], ["skill_soul_rend"]),
    ("elite_soul_hall_person_hall_guard", 55, [], [SOUL], ["skill_soul_handprint", "skill_soul_chain"]),
    ("elite_soul_hall_earth_hall_guard", 60, [], [SOUL], ["skill_soul_rend", "skill_soul_chain"]),
    ("elite_soul_hall_heaven_hall_guard", 65, [], [SOUL], ["skill_soul_art", "skill_soul_rend", "skill_soul_chain"]),
    ("elite_soul_hall_messenger", 50, [], [SOUL], ["skill_soul_handprint", "skill_shadow_strike"]),
    ("elite_death_corpse_guardian", 52, [], [SOUL], ["skill_life_drain"]),
    ("elite_soul_hall_two_star", 75, [], [SOUL], ["skill_soul_art", "skill_soul_rend", "skill_soul_pain"]),
    ("elite_soul_hall_three_star", 78, [], [SOUL], ["skill_soul_art", "skill_soul_rend", "skill_soul_domain"]),
    ("elite_soul_hall_five_star", 82, [], [SOUL], ["skill_soul_art", "skill_soul_rend", "skill_ten_thousand_souls"]),
    ("elite_hun_spy", 56, [], [SOUL], ["skill_soul_handprint", "skill_shadow_strike"]),

    # ── 中州 Lv 38-55 ──
    ("elite_wind_lightning_deacon", 38, [], [WIND, THUNDER], ["skill_gale_strike"]),
    ("elite_thunder_pavilion_north", 44, [], [WIND, THUNDER], ["skill_thunder_arc_dance", "skill_gale_strike"]),
    ("elite_thunder_pavilion_east", 43, [], [WIND, THUNDER], ["skill_thunder_arc_dance", "skill_gale_strike"]),
    ("elite_huangquan_deacon", 40, [], [SOUL], ["skill_huangquan_palm"]),
    ("elite_wanjian_deacon", 42, [], [], ["skill_sword"]),
    ("elite_burning_valley_deacon", 44, [], [FIRE], ["skill_fallen_heart_pulse"]),
    ("elite_fire_valley_smith", 46, [], [FIRE], ["skill_pill_boost"]),
    ("elite_ice_river_deacon", 45, [], [ICE], ["skill_ice_dragon"]),
    ("elite_ice_valley_freezer", 48, [], [ICE], ["skill_absolute_zero"]),
    ("elite_sky_demon_deacon", 46, [], [SOUL], ["skill_soul_art"]),
    ("elite_flower_sect_elder", 48, [], [], ["skill_flower", "skill_pill_boost"]),
    ("elite_dan_tower_deacon", 50, [], [FIRE], ["skill_alchemy", "skill_pill_boost"]),
    ("elite_dan_region_guard_captain", 52, [], [], ["skill_defensive_stance", "skill_pill_boost"]),
    ("elite_dan_region_alchemist", 55, [], [FIRE], ["skill_alchemy", "skill_pill_boost"]),
    ("elite_dan_tower_examiner", 54, [], [FIRE], ["skill_alchemy", "skill_pill_boost"]),
    ("elite_central_domain_merc", 50, [], [], ["skill_beast_charge", "skill_battle_cry"]),
    ("elite_tianhuang_guard", 54, [], [], ["skill_defensive_stance", "skill_battle_cry", "skill_bloodline_pressure"]),
    ("elite_black_fire_sect_leader", 48, [], [FIRE, POISON], ["skill_black_fire_ancestor_flame"]),
    ("elite_space_trade_guard", 52, [], [], ["skill_bloodline_pressure"]),
    ("elite_ancient_ruin_explorer", 56, [], [ANCIENT], ["skill_ancient_seal_art"]),
    ("elite_beast_region_tamer", 54, [], [], ["skill_beast_roar"]),

    # ── 兽域/龙岛 Lv 52-78 ──
    ("elite_transformed_beast", 52, [], [ANCIENT], ["skill_beast_roar", "skill_bloodline_pressure"]),
    ("elite_nether_python_elder", 56, [], [POISON, SOUL], ["skill_bloodline_pressure"]),
    ("elite_sky_demon_phoenix", 60, [], [WIND], ["skill_sky_phoenix_elder_art", "skill_bloodline_pressure"]),
    ("elite_beast_region_chieftain", 58, [], [], ["skill_beast_roar", "skill_bloodline_pressure"]),
    ("elite_ancient_beast_guardian", 62, [], [DRAGON, ANCIENT], ["skill_beast_roar", "skill_bloodline_pressure"]),
    ("elite_dragon_warrior", 58, [], [DRAGON], ["skill_beast_charge"]),
    ("elite_dragon_guard_captain", 62, [], [DRAGON], ["skill_dragon_roar", "skill_defensive_stance"]),
    ("elite_west_dragon_commander", 65, [], [DRAGON], ["skill_dragon_roar", "skill_battle_cry"]),
    ("elite_south_dragon_commander", 65, [], [DRAGON], ["skill_dragon_roar", "skill_battle_cry"]),
    ("elite_north_dragon_commander", 68, [], [DRAGON], ["skill_dragon_claw_void", "skill_battle_cry"]),
    ("elite_east_dragon_commander", 62, [], [DRAGON], ["skill_dragon_roar", "skill_defensive_stance"]),
    ("elite_ancient_dragon_spirit", 72, [], [DRAGON, ANCIENT], ["skill_ancient_dragon_spirit_roar", "skill_bloodline_pressure"]),
    ("elite_nine_serene_guard", 60, [], [POISON, SOUL], ["skill_bloodline_pressure"]),
    ("elite_void_ship_captain", 60, [], [WIND], ["skill_space_trade_master_art"]),

    # ── 远古种族 Lv 60-78 ──
    ("elite_gu_clan_warrior", 62, [], [ANCIENT], ["skill_bloodline_pressure"]),
    ("elite_gu_clan_elder", 72, [], [ANCIENT], ["skill_ancient_emperor_seal", "skill_bloodline_pressure"]),
    ("elite_gu_clan_ritual_elder", 74, [], [ANCIENT], ["skill_ancient_emperor_seal", "skill_ancient_seal_art", "skill_bloodline_pressure"]),
    ("elite_yan_clan_warrior", 64, [], [FIRE, ANCIENT], ["skill_bloodline_pressure"]),
    ("elite_yan_clan_flame_master", 68, [], [FIRE, ANCIENT], ["skill_bloodline_pressure"]),
    ("elite_lei_clan_warrior", 64, [], [THUNDER, ANCIENT], ["skill_bloodline_pressure"]),
    ("elite_lei_clan_thunder_master", 68, [], [THUNDER, ANCIENT], ["skill_bloodline_pressure"]),
    ("elite_shi_clan_warrior", 64, [], [EARTH, ANCIENT], ["skill_bloodline_pressure"]),
    ("elite_ling_clan_scout", 62, [], [ANCIENT], ["skill_shadow_strike", "skill_bloodline_pressure"]),
    ("elite_ancient_alliance_officer", 75, [], [ANCIENT], ["skill_battle_cry", "skill_bloodline_pressure"]),
    ("elite_hun_clan_warrior", 68, [], [SOUL, ANCIENT], ["skill_bloodline_pressure"]),
    ("elite_hun_clan_elder", 78, [], [SOUL, ANCIENT], ["skill_soul_art", "skill_soul_rend", "skill_bloodline_pressure"]),
    ("elite_yao_clan_guard", 60, [], [], ["skill_alchemy", "skill_pill_boost"]),
    ("elite_yao_clan_elder", 70, [], [], ["skill_alchemy", "skill_pill_boost", "skill_battle_meditation"]),

    # ── 莽荒/妖火/洞府/战场 Lv 62-92 ──
    ("elite_manghuang_hunter", 64, [], [], ["skill_beast_charge", "skill_shadow_strike"]),
    ("elite_demon_python_guardian", 68, [], [POISON], ["skill_bloodline_pressure"]),
    ("elite_ancient_platform_guard", 70, [], [ANCIENT], ["skill_ancient_seal_art", "skill_bloodline_pressure"]),
    ("elite_bodhi_illusion_guard", 72, [], [SOUL, ANCIENT], ["skill_ancient_seal_art", "skill_bloodline_pressure"]),
    ("elite_bodhi_guardian_spirit", 68, [], [ANCIENT], ["skill_ancient_seal_art", "skill_healing", "skill_bloodline_pressure"]),
    ("elite_demon_flame_warden", 66, [], [FIRE], ["skill_bloodline_pressure"]),
    ("elite_saint_remains_guard", 70, [], [FIRE], ["skill_purifying_saint", "skill_bloodline_pressure"]),
    ("elite_emperor_cave_sentry", 78, [], [DRAGON, ANCIENT], ["skill_emperor_cave_dragon_guard", "skill_bloodline_pressure"]),
    ("elite_emperor_formation_guard", 82, [], [ANCIENT], ["skill_emperor_pill_guard", "skill_ancient_seal_art", "skill_bloodline_pressure"]),
    ("elite_emperor_treasure_guard", 85, [], [ANCIENT], ["skill_emperor_pill_guard", "skill_bloodline_pressure", "skill_ancient_seal_art"]),
    ("elite_allied_forces_general", 88, [], [], ["skill_alliance_commander_art", "skill_battle_cry"]),
    ("elite_alliance_champion", 92, [], [ANCIENT], ["skill_alliance_commander_art", "skill_bloodline_pressure"]),
    ("elite_hun_army_commander", 90, [], [SOUL], ["skill_soul_art", "skill_soul_rend", "skill_bloodline_pressure"]),
    ("elite_small_dan_tower_guard", 62, [], [FIRE], ["skill_small_dan_elder_art", "skill_pill_boost"]),
    ("elite_wan_yao_mountain_guard", 54, [], [], ["skill_wan_yao_mountain_lord_art", "skill_pill_boost"]),
]

# ══════════════════════════════════════════════════════════
# APPLY
# ══════════════════════════════════════════════════════════
known_elite_ids = {eid for eid, *_ in ELITE_DATA}
updated = 0
for row in ws_enemies.iter_rows(min_row=2):
    eid = str(row[0].value)
    if eid not in known_elite_ids: continue
    etype = str(row[2].value) if row[2].value else ""
    if etype != "elite": continue

    # Find the data
    data = next((d for d in ELITE_DATA if d[0] == eid), None)
    if not data: continue

    _, level, faction, element, uniques = data
    skills = assign_elite_skills(level, faction, element, uniques)

    ws_enemies.cell(row=row[0].row, column=skills_col+1).value = ",".join(skills)
    updated += 1
    name = str(row[1].value) if row[1].value else "?"
    print(f"  {name} Lv{level}: {len(skills)}技")

wb.save(WORKBOOK_PATH)
print(f"\nUpdated {updated} elites with full skill sets.")
print("Done!")
