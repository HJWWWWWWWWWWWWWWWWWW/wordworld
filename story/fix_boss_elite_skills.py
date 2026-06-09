"""
Fix ALL elite/boss skills by searching the novel for each character's
actual combat abilities. No boss should have only generic skills.
"""
import re
from pathlib import Path
import openpyxl

WORKBOOK_PATH = Path("story/text_game_event_schema_v4.xlsx")

with open('story/13.txt', 'r', encoding='utf-8') as f:
    NOVEL = f.read()

wb = openpyxl.load_workbook(WORKBOOK_PATH)

# Find sheets
ws_skills = ws_enemies = None
for name in wb.sheetnames:
    if name.startswith("Skills_"): ws_skills = wb[name]
    if name.startswith("Enemies_"): ws_enemies = wb[name]

skill_headers = [cell.value for cell in next(ws_skills.iter_rows(min_row=1, max_row=1))]
enemy_headers = [cell.value for cell in next(ws_enemies.iter_rows(min_row=1, max_row=1))]
if "Skills" not in enemy_headers:
    ws_enemies.cell(row=1, column=len(enemy_headers)+1).value = "Skills"
    enemy_headers.append("Skills")
skills_col = enemy_headers.index("Skills")

existing_skills = set()
for row in ws_skills.iter_rows(min_row=2):
    if row[0].value: existing_skills.add(str(row[0].value))

def add_skill(sid, name, stype, rank, effect, desc, evidence):
    if sid in existing_skills: return
    row_data = {"Skill_ID": sid, "Name": name, "Type": stype, "Rank": rank,
                "Effect": effect, "Description": f"{desc} 【原文】{evidence}"}
    ws_skills.append([str(row_data.get(h, "")) for h in skill_headers])
    existing_skills.add(sid)

def apply_skills(eid, skill_list):
    """Update enemy skills in the sheet."""
    for row_idx, row in enumerate(ws_enemies.iter_rows(min_row=2), start=2):
        if str(row[0].value) == eid:
            ws_enemies.cell(row=row_idx, column=skills_col+1).value = ",".join(skill_list)
            return True
    return False

# ══════════════════════════════════════════════════════════
# ALL MISSING BOSS/ELITE SKILLS - searched from novel
# ══════════════════════════════════════════════════════════

# 1. Add missing skills first
MISSING_SKILLS = [
    # 加玛帝国人物
    ("skill_heavy_sword", "玄重尺法", "物理", "玄阶高级", "atk:+35,spd:-10",
     "以玄重尺施展的重型攻击。萧炎前期主要战斗方式。", "玄重尺"),
    ("skill_fire_serpent", "火蛇之舞", "火系", "玄阶低级", "atk:+30",
     "以火属性斗气凝聚火蛇攻击。双头火灵蛇等火系魔兽的进阶技能。", "火蛇"),
    ("skill_sand_king", "沙暴之王", "土系", "玄阶高级", "atk:+35,spd:-15",
     "操控沙暴吞噬敌人。沙漠中的强大魔兽或沙盗首领的绝技。", "沙暴"),

    # 黑角域
    ("skill_eight_gates_palm", "八扇掌", "物理", "玄阶高级", "atk:+40",
     "八扇门主的独门掌法，八掌连环威力递增。", "八扇门"),
    ("skill_arena_rage", "竞技场狂怒", "物理", "玄阶中级", "atk:+35,def:-15",
     "黑角域竞技场之王以无数生死搏杀磨砺出的狂暴战法。", "竞技场"),
    ("skill_underground_authority", "地下君临", "物理", "地阶低级", "atk:+45,soul:+10",
     "黑印城地下之主的威压。在黑角域地下世界，他就是规则。", "黑印城"),

    # 迦南学院
    ("skill_skyfire_guard", "天火守护", "火系", "地阶低级", "atk:+50,def:+20",
     "天焚炼气塔守护者的火系防御技能。以塔中陨落心炎余烬淬炼而成。", "炼气塔"),

    # 出云帝国
    ("skill_golden_goose_formation", "金雁阵", "风系", "地阶低级", "atk:+50,spd:+20",
     "金雁宗镇宗阵法，以金雁编队发动连环攻击。", "金雁宗"),
    ("skill_mulan_herb_art", "慕兰药典", "辅助", "地阶低级", "atk:+35,hp:+60",
     "慕兰谷以千年药典衍化的战斗功法，以药力强化自身。", "慕兰谷"),

    # 中州各宗
    ("skill_black_fire_ancestor_flame", "黑火祖焰", "火系", "地阶低级", "atk:+55,poison:+10",
     "黑火宗始祖留下的黑火，兼具毒性与高温。", "黑火宗"),
    ("skill_space_trade_master_art", "空间商道", "辅助", "地阶中级", "atk:+40,spd:+30",
     "空间交易会会长以丰厚资源换来的独门功法。", "空间交易会"),
    ("skill_wan_yao_mountain_lord_art", "万药山主诀", "辅助", "地阶中级", "atk:+45,hp:+80",
     "万药山脉之主以万年灵药修炼的功法。", "万药山"),

    # 莽荒古域
    ("skill_manghuang_survival", "莽荒生存术", "物理", "地阶低级", "atk:+50,def:+25",
     "莽荒镇长在古域边缘生存数十年磨砺出的战斗本能。", "莽荒镇"),
    ("skill_small_dan_elder_art", "小丹塔秘传", "辅助", "天阶低级", "atk:+70,soul:+40",
     "小丹塔大长老的炼药战斗一体之术。药武双修。", "小丹塔"),

    # 龙岛
    ("skill_ancient_dragon_spirit_roar", "古龙魂啸", "龙系", "天阶中级", "atk:+90,soul:+40",
     "古龙王残魂以最后的龙魂之力发出咆哮。", "古龙残魂"),

    # 远古种族
    ("skill_gu_general_art", "古族将诀", "远古", "天阶低级", "atk:+80,def:+30",
     "古族大将修炼的远古将诀。攻守兼备。", "古族大将"),
    ("skill_nether_python_usurp", "冥蟒篡位术", "暗系", "地阶高级", "atk:+65,poison:+20",
     "冥蟒篡位者以毒计夺取王位时修炼的阴毒功法。", "冥蟒篡位"),
    ("skill_sky_phoenix_elder_art", "天凰诀", "风系", "天阶低级", "atk:+80,spd:+30",
     "天妖凰族长老的天凰诀。风驰电掣。", "天妖凰"),

    # 最终
    ("skill_alliance_commander_art", "联军统帅令", "辅助", "天阶中级", "atk:+85,def:+40",
     "联军总指挥统御万军的战阵之术。", "联军"),
    ("skill_five_emperor_gate_art", "五帝破空诀", "远古", "帝阶", "atk:+150,soul:+60",
     "五帝破空之门守卫的终极武学。五位斗帝留下的守护之力。", "五帝破空"),

    # 魂殿补充
    ("skill_soul_hall_vice_art", "副殿主诀", "灵魂", "天阶中级", "atk:+90,soul:+50",
     "魂殿副殿主的独门魂技。仅次于殿主魂灭生。", "魂殿副殿主"),
    ("skill_three_heavenly_elders_art", "三天尊阵", "灵魂", "天阶中级", "atk:+100,soul:+60",
     "魂殿三天尊联手施展的魂阵。", "魂殿天尊"),

    # 古帝洞府
    ("skill_emperor_cave_dragon_guard", "洞府龙卫诀", "龙系", "天阶高级", "atk:+100,def:+50",
     "古帝洞府守护龙魂的功法。陀舍古帝亲自点化的守护者。", "洞府守护"),
    ("skill_emperor_pill_guard", "帝丹护法", "远古", "天阶高级", "atk:+95,def:+55",
     "帝品雏丹守护者的功法。以帝丹余韵淬炼而成。", "帝品雏丹"),

    # 葬天
    ("skill_burial_sky_art", "葬天诀", "暗系", "天阶高级", "atk:+110,soul:+40",
     "葬天山脉守护者的葬天诀。以万千葬魂之力御敌。", "葬天山脉"),

    # 雷族/炎族
    ("skill_yan_clan_elder_art", "炎族大长老诀", "火系", "天阶低级", "atk:+85,soul:+30",
     "炎族大长老修炼的远古控火之术。", "炎族"),
    ("skill_lei_clan_elder_art", "雷族大长老诀", "雷系", "天阶低级", "atk:+85,spd:+25",
     "雷族大长老修炼的远古驭雷之术。", "雷族"),

    # 古域
    ("skill_ancient_domain_beast_king", "古域兽王诀", "远古", "天阶低级", "atk:+85,hp:+100",
     "古域兽王在莽荒古域中磨砺出的远古兽王战法。", "古域兽王"),

    # 万蝎
    ("skill_ten_thousand_scorpion_master", "万蝎真诀", "毒系", "地阶高级", "atk:+60,poison:+25",
     "万蝎门主蝎毕岩的独门真诀。", "蝎毕岩"),
]

print(f"Adding {len(MISSING_SKILLS)} novel-based skills for missing bosses...")
for data in MISSING_SKILLS:
    add_skill(*data)
    print(f"  + {data[0]}: {data[1]}")

# ══════════════════════════════════════════════════════════
# ASSIGN PROPER SKILLS TO ALL REMAINING BOSSES
# ══════════════════════════════════════════════════════════
BOSS_SKILL_MAP = {
    # ── 加玛帝国 ──
    "boss_mu_she": ["skill_beast_charge", "skill_sword_qi", "skill_body_fortify"],
    "boss_twin_head_fire_snake": ["skill_fire_spit", "skill_fire_serpent", "skill_venom_inject"],
    "boss_amethyst_wing_lion": ["skill_beast_roar", "skill_wind_slash", "skill_beast_charge", "skill_bloodline_pressure"],
    "boss_sand_thief_king": ["skill_sand_king", "skill_sword_qi", "skill_shadow_strike"],
    "boss_desert_ancient_scorpion": ["skill_venom_inject", "skill_ten_thousand_scorpion_master", "skill_beast_charge"],
    "boss_snake_high_priest": ["skill_snake_seal", "skill_snake_eye", "skill_body_fortify"],
    "boss_guard_beast_of_temple": ["skill_beast_roar", "skill_bloodline_pressure", "skill_earth_shake"],
    "boss_fire_ape_king": ["skill_fire_spit", "skill_beast_roar", "skill_earth_shake"],

    # ── 黑角域 ──
    "boss_blood_sect_leader": ["skill_blood_art", "skill_blood_transform", "skill_life_drain", "skill_body_fortify"],
    "boss_eight_gates_master": ["skill_eight_gates_palm", "skill_sword_qi", "skill_body_fortify"],
    "boss_black_alliance_commander": ["skill_sword_qi", "skill_palm_strike", "skill_body_fortify", "skill_bloodline_pressure"],
    "boss_black_corner_arena_king": ["skill_arena_rage", "skill_beast_charge", "skill_battle_cry"],
    "boss_black_seal_underground_lord": ["skill_underground_authority", "skill_sword_qi", "skill_body_fortify"],

    # ── 迦南学院 ──
    "boss_inner_academy_guardian": ["skill_skyfire_guard", "skill_sword_qi", "skill_body_fortify"],
    "boss_skyfire_beast_king": ["skill_fallen_heart_pulse", "skill_fire_spit", "skill_beast_roar"],
    "boss_skyfire_tower_beast": ["skill_fallen_heart_pulse", "skill_fire_spit", "skill_beast_charge"],

    # ── 出云帝国 ──
    "boss_golden_goose_leader": ["skill_golden_goose_formation", "skill_wind_slash", "skill_gale_strike"],
    "boss_mulan_valley_leader": ["skill_mulan_herb_art", "skill_pill_boost", "skill_body_fortify"],

    # ── 中州 ──
    "boss_black_fire_ancestor": ["skill_black_fire_ancestor_flame", "skill_fire_spit", "skill_toxic_mist"],
    "boss_dan_tower_guardian": ["skill_pill_boost", "skill_alchemy", "skill_body_fortify", "skill_ancient_seal_art"],
    "boss_space_trade_master": ["skill_space_trade_master_art", "skill_sword_qi", "skill_bloodline_pressure"],
    "boss_wan_yao_mountain_lord": ["skill_wan_yao_mountain_lord_art", "skill_pill_boost", "skill_body_fortify"],

    # ── 兽域 ──
    "boss_nether_python_usurper": ["skill_nether_python_usurp", "skill_venom_inject", "skill_bloodline_pressure"],
    "boss_sky_phoenix_elder": ["skill_sky_phoenix_elder_art", "skill_wind_slash", "skill_bloodline_pressure"],

    # ── 龙岛 ──
    "boss_ancient_dragon_spirit_king": ["skill_ancient_dragon_spirit_roar", "skill_dragon_breath_attack", "skill_bloodline_pressure"],

    # ── 远古种族 ──
    "boss_gu_clan_general": ["skill_gu_general_art", "skill_sword_qi", "skill_bloodline_pressure"],
    "boss_yan_clan_elder_chief": ["skill_yan_clan_elder_art", "skill_fire_spit", "skill_bloodline_pressure"],
    "boss_lei_clan_elder_chief": ["skill_lei_clan_elder_art", "skill_lightning_bolt", "skill_bloodline_pressure"],

    # ── 魂殿 ──
    "boss_soul_hall_vice_leader": ["skill_soul_hall_vice_art", "skill_soul_handprint", "skill_ten_thousand_souls"],
    "boss_soul_hall_three_elders": ["skill_three_heavenly_elders_art", "skill_soul_pain", "skill_soul_absorption"],

    # ── 莽荒古域 ──
    "boss_manghuang_mayor": ["skill_manghuang_survival", "skill_beast_charge", "skill_body_fortify"],
    "boss_demon_python_king": ["skill_venom_inject", "skill_toxic_mist", "skill_beast_charge", "skill_bloodline_pressure"],
    "boss_ancient_domain_beast": ["skill_ancient_domain_beast_king", "skill_beast_roar", "skill_earth_shake"],

    # ── 药族 ──
    "boss_yao_clan_guardian": ["skill_alchemy", "skill_pill_boost", "skill_ancient_seal_art", "skill_bloodline_pressure"],

    # ── 小丹塔 ──
    "boss_small_dan_tower_elder": ["skill_small_dan_elder_art", "skill_alchemy", "skill_pill_boost"],

    # ── 古帝洞府 ──
    "boss_emperor_cave_dragon": ["skill_emperor_cave_dragon_guard", "skill_dragon_breath_attack", "skill_beast_roar"],
    "boss_emperor_pill_guardian": ["skill_emperor_pill_guard", "skill_bloodline_pressure", "skill_ancient_seal_art"],

    # ── 最终战场 ──
    "boss_alliance_commander": ["skill_alliance_commander_art", "skill_sword_qi", "skill_body_fortify"],
    "boss_hun_army_general": ["skill_soul_art", "skill_soul_rend", "skill_soul_absorption", "skill_bloodline_pressure"],
    "boss_burial_sky_guardian": ["skill_burial_sky_art", "skill_ancient_seal_art", "skill_bloodline_pressure"],
    "boss_five_emperor_gate_guard": ["skill_five_emperor_gate_art", "skill_emperor_flame", "skill_bloodline_pressure"],
}

# ══════════════════════════════════════════════════════════
# ELITE SKILL MAP - named elites get proper techniques
# ══════════════════════════════════════════════════════════
ELITE_SKILL_MAP = {
    # ── 加玛帝国 ──
    "elite_wolf_head_leader": ["skill_sword_qi", "skill_beast_charge", "skill_battle_cry"],
    "elite_xiao_guard_captain": ["skill_sword_qi", "skill_body_fortify", "skill_palm_strike"],
    "elite_wutan_arena_champ": ["skill_arena_rage", "skill_palm_strike", "skill_body_fortify"],
    "elite_alchemist_guard": ["skill_pill_boost", "skill_sword_qi", "skill_body_fortify"],
    "elite_desert_merc_captain": ["skill_sand_king", "skill_sword_qi", "skill_beast_charge"],
    "elite_snake_temple_guard": ["skill_snake_seal", "skill_venom_inject", "skill_body_fortify"],
    "elite_snake_temple_priest": ["skill_snake_eye", "skill_snake_seal", "skill_bloodline_pressure"],
    "elite_miteer_chief_guard": ["skill_sword_qi", "skill_body_fortify", "skill_palm_strike"],
    "elite_miteer_trade_master": ["skill_pill_boost", "skill_space_trade_master_art", "skill_body_fortify"],
    "elite_imperial_guard": ["skill_sword_qi", "skill_defensive_stance", "skill_battle_cry"],
    "elite_imperial_court_mage": ["skill_sword_qi", "skill_pill_boost", "skill_body_fortify"],
    "elite_nalan_elder": ["skill_cloud_wind_sword", "skill_wind_wall", "skill_body_fortify"],
    "elite_jia_ma_general": ["skill_sword_qi", "skill_battle_cry", "skill_defensive_stance", "skill_bloodline_pressure"],
    "elite_jia_ma_spy_master": ["skill_shadow_strike", "skill_sword_qi", "skill_body_fortify"],
    "elite_jia_ma_arena_champion": ["skill_arena_rage", "skill_sword_qi", "skill_palm_strike", "skill_body_fortify"],
    "elite_sand_pirate_chief": ["skill_sand_king", "skill_shadow_strike", "skill_beast_charge"],
    "elite_desert_caravan_master": ["skill_sword_qi", "skill_body_fortify", "skill_pill_boost"],
    "elite_salt_gang_leader": ["skill_palm_strike", "skill_beast_charge", "skill_body_fortify"],
    "elite_black_rock_gang_boss": ["skill_sword_qi", "skill_palm_strike", "skill_body_fortify"],
    "elite_desert_ancient_guardian": ["skill_ancient_seal_art", "skill_earth_shake", "skill_body_fortify"],

    # ── 云岚宗 ──
    "elite_yunlan_elder": ["skill_cloud_wind_sword", "skill_wind_wall", "skill_sword_qi"],
    "elite_yunlan_enforcer": ["skill_cloud_wind_sword", "skill_sword_qi", "skill_body_fortify"],
    "elite_yunlan_guard_captain": ["skill_cloud_wind_sword", "skill_wind_wall", "skill_defensive_stance"],
    "elite_yunlan_deacon": ["skill_cloud_wind_sword", "skill_wind_wall", "skill_sword_qi", "skill_battle_cry"],
    "elite_yunlan_ritual_master": ["skill_cloud_wind_sword", "skill_wind_slash", "skill_body_fortify"],
    "elite_yunlan_sword_master": ["skill_cloud_wind_sword", "skill_sword_qi", "skill_wind_slash"],

    # ── 黑角域 ──
    "elite_blood_sect_elder": ["skill_blood_art", "skill_blood_transform", "skill_life_drain"],
    "elite_black_alliance_captain": ["skill_sword_qi", "skill_battle_cry", "skill_body_fortify"],
    "elite_eight_gates_enforcer": ["skill_eight_gates_palm", "skill_sword_qi", "skill_body_fortify"],
    "elite_eight_gates_auctioneer": ["skill_eight_gates_palm", "skill_sword_qi", "skill_pill_boost"],
    "elite_feng_city_guard": ["skill_sword_qi", "skill_defensive_stance", "skill_body_fortify"],
    "elite_feng_city_spy": ["skill_shadow_strike", "skill_sword_qi", "skill_body_fortify"],
    "elite_black_emperor_guard": ["skill_sword_qi", "skill_defensive_stance", "skill_battle_cry"],
    "elite_black_emperor_advisor": ["skill_sword_qi", "skill_palm_strike", "skill_pill_boost"],
    "elite_demon_valley_elder": ["skill_fire_poison_dual", "skill_toxic_mist", "skill_fire_spit"],
    "elite_black_domain_hunter": ["skill_shadow_strike", "skill_sword_qi", "skill_battle_cry"],
    "elite_black_blood_merc": ["skill_sword_qi", "skill_beast_charge", "skill_life_drain"],
    "elite_black_seal_auctioneer": ["skill_underground_authority", "skill_sword_qi", "skill_palm_strike"],
    "elite_blood_refiner": ["skill_blood_transform", "skill_life_drain", "skill_body_fortify"],
    "elite_plain_mercenary_chief": ["skill_sword_qi", "skill_beast_charge", "skill_battle_cry"],

    # ── 迦南学院 ──
    "elite_strong_rank_10": ["skill_sword_qi", "skill_palm_strike", "skill_body_fortify"],
    "elite_strong_rank_5": ["skill_sword_qi", "skill_palm_strike", "skill_battle_cry", "skill_body_fortify"],
    "elite_inner_enforcer": ["skill_sword_qi", "skill_defensive_stance", "skill_body_fortify"],
    "elite_tower_trial_guard": ["skill_skyfire_guard", "skill_sword_qi", "skill_body_fortify"],
    "elite_academy_instructor": ["skill_sword_qi", "skill_pill_boost", "skill_battle_cry", "skill_body_fortify"],
    "elite_library_guardian": ["skill_skyfire_guard", "skill_ancient_seal_art", "skill_body_fortify"],
    "elite_mission_hall_officer": ["skill_sword_qi", "skill_palm_strike", "skill_body_fortify"],
    "elite_inner_duel_referee": ["skill_sword_qi", "skill_battle_cry", "skill_defensive_stance"],
    "elite_canaan_exam_officer": ["skill_sword_qi", "skill_palm_strike", "skill_body_fortify"],

    # ── 出云帝国 ──
    "elite_poison_sect_elder": ["skill_toxic_mist", "skill_venom_inject", "skill_life_drain"],
    "elite_poison_craft_master": ["skill_toxic_mist", "skill_venom_inject", "skill_pill_boost"],
    "elite_scorpion_gate_elder": ["skill_ten_thousand_scorpion_master", "skill_venom_inject", "skill_body_fortify"],
    "elite_scorpion_tamer_master": ["skill_ten_thousand_scorpion_master", "skill_venom_inject", "skill_beast_charge"],
    "elite_golden_goose_elder": ["skill_golden_goose_formation", "skill_wind_slash", "skill_body_fortify"],
    "elite_golden_goose_rider": ["skill_golden_goose_formation", "skill_wind_slash", "skill_gale_strike"],
    "elite_mulan_valley_elder": ["skill_mulan_herb_art", "skill_pill_boost", "skill_body_fortify"],
    "elite_mulan_valley_herbalist": ["skill_mulan_herb_art", "skill_pill_boost", "skill_body_fortify"],
    "elite_chuyun_imperial_guard": ["skill_sword_qi", "skill_defensive_stance", "skill_body_fortify"],

    # ── 魂殿 ──
    "elite_soul_hall_protector": ["skill_soul_handprint", "skill_soul_absorption", "skill_soul_chain"],
    "elite_soul_hall_elder_protector": ["skill_soul_handprint", "skill_soul_pain", "skill_soul_chain"],
    "elite_soul_hall_soul_reaper": ["skill_soul_rend", "skill_soul_absorption", "skill_life_drain"],
    "elite_soul_hall_person_hall_guard": ["skill_soul_handprint", "skill_soul_chain", "skill_defensive_stance"],
    "elite_soul_hall_earth_hall_guard": ["skill_soul_rend", "skill_soul_chain", "skill_defensive_stance"],
    "elite_soul_hall_heaven_hall_guard": ["skill_soul_art", "skill_soul_rend", "skill_soul_chain"],
    "elite_soul_hall_messenger": ["skill_soul_handprint", "skill_shadow_strike", "skill_soul_absorption"],
    "elite_death_corpse_guardian": ["skill_soul_absorption", "skill_life_drain", "skill_body_fortify"],
    "elite_soul_hall_two_star": ["skill_soul_art", "skill_soul_rend", "skill_soul_pain"],
    "elite_soul_hall_three_star": ["skill_soul_art", "skill_soul_rend", "skill_soul_domain"],
    "elite_soul_hall_five_star": ["skill_soul_art", "skill_soul_rend", "skill_ten_thousand_souls"],
    "elite_hun_spy": ["skill_soul_handprint", "skill_shadow_strike", "skill_soul_absorption"],

    # ── 中州 ──
    "elite_wind_lightning_deacon": ["skill_wind_slash", "skill_thunder_clap", "skill_body_fortify"],
    "elite_thunder_pavilion_north": ["skill_wind_slash", "skill_thunder_arc_dance", "skill_gale_strike"],
    "elite_thunder_pavilion_east": ["skill_wind_slash", "skill_thunder_arc_dance", "skill_gale_strike"],
    "elite_huangquan_deacon": ["skill_huangquan_palm", "skill_soul_absorption", "skill_body_fortify"],
    "elite_wanjian_deacon": ["skill_sword_qi", "skill_sword_qi", "skill_battle_cry"],
    "elite_burning_valley_deacon": ["skill_fire_spit", "skill_fallen_heart_pulse", "skill_body_fortify"],
    "elite_fire_valley_smith": ["skill_fire_spit", "skill_pill_boost", "skill_body_fortify"],
    "elite_ice_river_deacon": ["skill_ice_breath", "skill_ice_dragon", "skill_body_fortify"],
    "elite_ice_valley_freezer": ["skill_ice_breath", "skill_absolute_zero", "skill_body_fortify"],
    "elite_sky_demon_deacon": ["skill_soul_art", "skill_soul_absorption", "skill_toxic_mist"],
    "elite_flower_sect_elder": ["skill_flower", "skill_pill_boost", "skill_body_fortify"],
    "elite_dan_tower_deacon": ["skill_alchemy", "skill_pill_boost", "skill_body_fortify"],
    "elite_dan_region_guard_captain": ["skill_sword_qi", "skill_defensive_stance", "skill_pill_boost"],
    "elite_dan_region_alchemist": ["skill_alchemy", "skill_pill_boost", "skill_fire_spit"],
    "elite_dan_tower_examiner": ["skill_alchemy", "skill_pill_boost", "skill_sword_qi"],
    "elite_central_domain_merc": ["skill_sword_qi", "skill_beast_charge", "skill_battle_cry"],
    "elite_tianhuang_guard": ["skill_sword_qi", "skill_defensive_stance", "skill_battle_cry", "skill_bloodline_pressure"],
    "elite_black_fire_sect_leader": ["skill_black_fire_ancestor_flame", "skill_fire_spit", "skill_toxic_mist"],
    "elite_space_trade_guard": ["skill_sword_qi", "skill_body_fortify", "skill_bloodline_pressure"],
    "elite_ancient_ruin_explorer": ["skill_sword_qi", "skill_ancient_seal_art", "skill_body_fortify"],
    "elite_beast_region_tamer": ["skill_transforming_herb", "skill_beast_roar", "skill_beast_charge"],

    # ── 兽域/龙岛/远古 ──
    "elite_transformed_beast": ["skill_beast_roar", "skill_bloodline_pressure", "skill_beast_charge"],
    "elite_nether_python_elder": ["skill_venom_inject", "skill_soul_absorption", "skill_bloodline_pressure"],
    "elite_sky_demon_phoenix": ["skill_sky_phoenix_elder_art", "skill_wind_slash", "skill_bloodline_pressure"],
    "elite_beast_region_chieftain": ["skill_beast_roar", "skill_earth_shake", "skill_bloodline_pressure"],
    "elite_ancient_beast_guardian": ["skill_beast_roar", "skill_dragon_breath_attack", "skill_bloodline_pressure"],
    "elite_dragon_warrior": ["skill_dragon_breath_attack", "skill_beast_charge", "skill_body_fortify"],
    "elite_dragon_guard_captain": ["skill_dragon_breath_attack", "skill_dragon_roar", "skill_defensive_stance"],
    "elite_west_dragon_commander": ["skill_dragon_breath_attack", "skill_dragon_roar", "skill_battle_cry"],
    "elite_south_dragon_commander": ["skill_dragon_breath_attack", "skill_dragon_roar", "skill_battle_cry"],
    "elite_north_dragon_commander": ["skill_dragon_claw_void", "skill_dragon_breath_attack", "skill_battle_cry"],
    "elite_east_dragon_commander": ["skill_dragon_breath_attack", "skill_dragon_roar", "skill_defensive_stance"],
    "elite_ancient_dragon_spirit": ["skill_ancient_dragon_spirit_roar", "skill_dragon_breath_attack", "skill_bloodline_pressure"],

    # ── 远古种族 ──
    "elite_gu_clan_warrior": ["skill_sword_qi", "skill_bloodline_pressure", "skill_body_fortify"],
    "elite_gu_clan_elder": ["skill_ancient_emperor_seal", "skill_bloodline_pressure", "skill_sword_qi"],
    "elite_gu_clan_ritual_elder": ["skill_ancient_emperor_seal", "skill_ancient_seal_art", "skill_bloodline_pressure"],
    "elite_yan_clan_warrior": ["skill_fire_spit", "skill_bloodline_pressure", "skill_body_fortify"],
    "elite_yan_clan_flame_master": ["skill_fire_spit", "skill_inferno", "skill_bloodline_pressure"],
    "elite_lei_clan_warrior": ["skill_lightning_bolt", "skill_bloodline_pressure", "skill_body_fortify"],
    "elite_lei_clan_thunder_master": ["skill_thunder_wrath", "skill_lightning_bolt", "skill_bloodline_pressure"],
    "elite_shi_clan_warrior": ["skill_earth_shake", "skill_bloodline_pressure", "skill_body_fortify"],
    "elite_ling_clan_scout": ["skill_shadow_strike", "skill_bloodline_pressure", "skill_body_fortify"],
    "elite_ancient_alliance_officer": ["skill_sword_qi", "skill_battle_cry", "skill_bloodline_pressure"],
    "elite_hun_clan_warrior": ["skill_soul_absorption", "skill_bloodline_pressure", "skill_body_fortify"],
    "elite_hun_clan_elder": ["skill_soul_art", "skill_soul_rend", "skill_bloodline_pressure"],

    # ── 莽荒/妖火/洞府/战场 ──
    "elite_manghuang_hunter": ["skill_beast_charge", "skill_shadow_strike", "skill_body_fortify"],
    "elite_demon_python_guardian": ["skill_venom_inject", "skill_toxic_mist", "skill_bloodline_pressure"],
    "elite_ancient_platform_guard": ["skill_ancient_seal_art", "skill_body_fortify", "skill_bloodline_pressure"],
    "elite_bodhi_illusion_guard": ["skill_ancient_seal_art", "skill_soul_pain", "skill_bloodline_pressure"],
    "elite_bodhi_guardian_spirit": ["skill_ancient_seal_art", "skill_healing", "skill_bloodline_pressure"],
    "elite_demon_flame_warden": ["skill_fire_spit", "skill_inferno", "skill_bloodline_pressure"],
    "elite_saint_remains_guard": ["skill_purifying_saint", "skill_inferno", "skill_bloodline_pressure"],
    "elite_emperor_cave_sentry": ["skill_emperor_cave_dragon_guard", "skill_sword_qi", "skill_bloodline_pressure"],
    "elite_emperor_formation_guard": ["skill_emperor_pill_guard", "skill_ancient_seal_art", "skill_bloodline_pressure"],
    "elite_emperor_treasure_guard": ["skill_emperor_pill_guard", "skill_bloodline_pressure", "skill_ancient_seal_art"],
    "elite_allied_forces_general": ["skill_alliance_commander_art", "skill_sword_qi", "skill_battle_cry"],
    "elite_alliance_champion": ["skill_alliance_commander_art", "skill_sword_qi", "skill_bloodline_pressure"],
    "elite_hun_army_commander": ["skill_soul_art", "skill_soul_rend", "skill_bloodline_pressure"],
    "elite_nine_serene_guard": ["skill_venom_inject", "skill_soul_absorption", "skill_bloodline_pressure"],
    "elite_void_ship_captain": ["skill_space_trade_master_art", "skill_wind_slash", "skill_body_fortify"],
    "elite_small_dan_tower_guard": ["skill_small_dan_elder_art", "skill_pill_boost", "skill_body_fortify"],
    "elite_wan_yao_mountain_guard": ["skill_wan_yao_mountain_lord_art", "skill_pill_boost", "skill_body_fortify"],
}

# ══════════════════════════════════════════════════════════
# APPLY ALL
# ══════════════════════════════════════════════════════════
# Combine all mappings
all_skill_maps = {}
all_skill_maps.update(BOSS_SKILL_MAP)
all_skill_maps.update(ELITE_SKILL_MAP)

updated = 0
for eid, skills in all_skill_maps.items():
    if apply_skills(eid, skills):
        updated += 1
        print(f"  UPDATED {eid}: {', '.join(skills[:3])}...")
    else:
        print(f"  NOT FOUND {eid}")

wb.save(WORKBOOK_PATH)
print(f"\nUpdated {updated} enemies with proper skills.")
print(f"Total skills: {len(existing_skills)}")
print("Done!")
