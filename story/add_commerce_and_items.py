"""
Complete commercial sub-areas for ALL cities + region-specific specialty items.
"""
from pathlib import Path
import openpyxl

WORKBOOK_PATH = Path("story/text_game_event_schema_v4.xlsx")

wb = openpyxl.load_workbook(WORKBOOK_PATH)


def find_sheet(name_prefix):
    for name in wb.sheetnames:
        if name.startswith(name_prefix):
            return wb[name]
    raise KeyError(f"Sheet starting with '{name_prefix}' not found")


def get_existing_ids(ws):
    ids = set()
    for row in ws.iter_rows(min_row=2):
        val = row[0].value
        if val:
            ids.add(str(val))
    return ids


def append_row(ws, data: dict):
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    row_data = []
    for h in headers:
        val = data.get(h, "")
        row_data.append(val if val is not None else "")
    ws.append(row_data)


ws_maps = find_sheet("Maps_")
ws_encounters = find_sheet("Map_Encounters_")
ws_items = find_sheet("Items_")
ws_gathering = find_sheet("Map_Gathering")
existing_maps = get_existing_ids(ws_maps)
existing_encounters = get_existing_ids(ws_encounters)
existing_items = get_existing_ids(ws_items)
existing_gathering = get_existing_ids(ws_gathering)

# ══════════════════════════════════════════════════════════
# PART 1: REGION-SPECIALTY ITEMS
# Each region gets unique items found only there
# ══════════════════════════════════════════════════════════
REGION_ITEMS = [
    # 沧澜帝国特产
    ("item_wutan_specialty", "青石城特产药材包", "consumable", "common", "false", "99", "50", "25",
     "", "hp:+20,stamina:+10",
     "青石城本地采集的混装药材，含凝血草、活气果等基础材料。仅在沧澜帝国境内可获得。", "沧澜帝国特产"),
    ("item_jia_ma_passport", "沧澜通关文牒", "quest", "common", "false", "1", "100", "50",
     "", "reputation:+5",
     "沧澜帝国官方颁发的通关文书，可免检通过帝国境内关隘。边境城市获取。", "沧澜特产"),

    # 赤沙荒漠特产
    ("item_desert_compass", "沙漠星盘", "equipment", "uncommon", "false", "1", "300", "150",
     "", "soul:+3",
     "以沙晶打磨的导航星盘，能在沙暴中指引方向。赤沙荒漠独有。", "沙漠特产"),
    ("item_snake_venom_vial", "蛇人毒液瓶", "material", "uncommon", "true", "10", "200", "100",
     "", "alchemy:+2",
     "蛇人族提炼的剧毒液，可用作毒丹材料或武器淬毒。仅蛇人族领地可获得。", "蛇人族特产"),

    # 魔兽山脉特产
    ("item_beast_hide_bundle", "魔兽皮料包", "material", "common", "true", "20", "80", "40",
     "", "silver:+30",
     "魔兽山脉猎获的皮料，包括狼皮、蛇鳞等。是制作防具的基础材料。", "魔兽山脉特产"),
    ("item_mountain_herb_kit", "山脉采药包", "consumable", "common", "true", "10", "60", "30",
     "", "hp:+15",
     "魔兽山脉特有的野生药材混合包，止血效果显著。限于魔兽山脉区域获得。", "魔兽山脉特产"),

    # 暗角域特产
    ("item_black_market_pass", "暗角域黑市令牌", "quest", "rare", "false", "1", "500", "250",
     "", "reputation:+3",
     "黑印城八扇门发放的黑市通行令牌，持有者可进入暗角域各大地下拍卖会。暗角域独有。", "暗角域特产"),
    ("item_forbidden_pill_fragment", "禁药配方残页", "material", "rare", "false", "1", "1000", "500",
     "", "alchemy:+5",
     "暗角域黑市中偶尔出现的禁药配方残页，记录着以寿命换取实力的禁忌炼药术。暗角域独有。", "暗角域特产"),

    # 迦南学院特产
    ("item_canaan_token", "火能卡", "currency", "uncommon", "true", "50", "0", "100",
     "", "exp:+10",
     "迦南学院内院流通的火能储值卡，可用于天焚炼气塔修炼或内院交易。迦南学院独有。", "迦南学院特产"),

    # 中州特产
    ("item_zhongzhou_map", "中州势力地图", "quest", "uncommon", "false", "1", "400", "200",
     "", "soul:+5",
     "标注中州各大势力分布与虫洞路线的详细地图。中州各大城池有售。", "中州特产"),
    ("item_wormhole_pass", "虫洞通行令", "quest", "rare", "false", "1", "800", "400",
     "", "exp:+20",
     "中州虫洞驿站发行的通行令牌，可优先使用特定虫洞线路。天涯城与天黄城可获取。", "中州特产"),

    # 丹域特产
    ("item_dan_herb_box", "丹域药材精装", "consumable", "rare", "true", "5", "600", "300",
     "", "alchemy:+3,hp:+50",
     "丹域特产的精选药材合集，包含数种中州特有的炼药灵材。仅丹域可获取。", "丹域特产"),
    ("item_alchemy_handbook", "炼药手札", "book", "uncommon", "false", "1", "500", "250",
     "", "alchemy:+5",
     "丹阁发行的炼药师入门手札，记录着基础丹方与控火技巧。丹域各城有售。", "丹域特产"),

    # 兽域特产
    ("item_beast_blood_essence", "魔兽精血瓶", "material", "rare", "true", "5", "500", "250",
     "", "douqi:+10",
     "高阶魔兽的精炼血液，蕴含浓郁灵力。可用于炼药或直接吸收。兽域独有。", "兽域特产"),
    ("item_transforming_herb", "化形草", "herb", "rare", "false", "1", "800", "400",
     "", "douqi:+5,hp:+30",
     "兽域深处生长的奇异灵草，是炼制化形丹的主材料。兽域独有。", "兽域特产"),

    # 龙岛特产
    ("item_dragon_scale", "龙鳞碎片", "material", "legendary", "false", "1", "3000", "1500",
     "", "def:+20,douqi:+30",
     "虚空龙族蜕落的鳞片碎片，蕴含龙族威压。可用于炼制顶级护甲。龙岛独有。", "龙岛特产"),
    ("item_dragon_saliva_herb", "龙涎灵草", "herb", "legendary", "false", "1", "2500", "1200",
     "", "alchemy:+8,hp:+100",
     "经龙涎长期浸润的灵草，药力远超寻常药材。龙岛独有。", "龙岛特产"),

    # 莽荒古域特产
    ("item_ancient_seed", "古域灵种", "material", "legendary", "false", "1", "2000", "1000",
     "", "soul:+15",
     "莽荒古域远古植物遗留的种子，蕴含庞大的生命精华。莽荒古域独有。", "莽荒古域特产"),
    ("item_bodhi_leaf", "菩提叶", "material", "legendary", "false", "1", "5000", "2500",
     "", "soul:+20,douqi:+20",
     "菩提古树上飘落的一片灵叶，蕴含菩提树的悟道之力。古域台附近可得。", "菩提古树特产"),

    # 妖火空间特产
    ("item_demon_flame_crystal", "妖火结晶", "material", "legendary", "false", "1", "4000", "2000",
     "", "douqi:+40",
     "净世白莲火焚烧空间形成的火晶，蕴含精纯的源火能量。妖火空间独有。", "妖火空间特产"),

    # 远古遗迹特产
    ("item_ancient_relic", "远古遗物碎片", "material", "legendary", "false", "1", "3500", "1800",
     "", "soul:+10,douqi:+10",
     "远古遗迹中残存的不知名器物碎片，仍保留着微弱的远古威压。远古遗迹独有。", "远古遗迹特产"),
]

# ══════════════════════════════════════════════════════════
# PART 2: COMMERCIAL SUB-AREAS FOR ALL CITIES
# ══════════════════════════════════════════════════════════
NEW_COMMERCE_MAPS = [
    # ── 沧澜帝国城市 ──
    ("map_black_rock_inn", "黑岩城客栈", "黑岩城", "", 4, 3, "yes", "",
     "黑岩城往来商旅落脚之处，大厅里流传着帝国东部的最新消息。", "商业"),
    ("map_mo_trade_post", "漠城商栈", "漠城", "", 6, 4, "yes", "",
     "漠城最大的货物集散地，沙漠商队在此装卸来自帝国各地的物资。", "商业"),
    ("map_stone_mo_inn", "石漠城客栈", "石漠城", "", 7, 3, "yes", "",
     "深入沙漠的最后一站，客栈墙上写满了前人留下的沙漠生存经验。", "商业"),
    ("map_salt_inn", "盐城客栈", "盐城", "", 5, 3, "yes", "",
     "盐帮经营的客栈，住客多是盐贩与码头工人。大堂里永远有盐帮的人在喝酒。", "商业"),

    # ── 暗角域城市 ──
    ("map_black_emperor_market", "黑皇城坊市", "黑皇城", "", 25, 5, "yes", "",
     "黑皇宗管辖的大型坊市，暗角域最正规的交易场所。丹药、功法、魔核齐全——价格也比别处公道。", "商业"),
    ("map_black_seal_inn", "黑印城客栈", "黑印城", "", 20, 3, "yes", "",
     "八扇门旗下的客栈，拍卖会期间一房难求。大堂是最好的情报交换地——只要你请得起酒。", "商业"),
    ("map_feng_market", "枫城坊市", "枫城", "", 25, 4, "yes", "",
     "冷煜经营时期留下的交易网络依然活跃。坊市以药材与丹药交易见长。", "商业"),
    ("map_peace_town_inn", "和平镇客栈", "和平镇", "", 15, 3, "yes", "",
     "进入迦南学院前的最后一站，客栈里挤满了赶考学员与护送家长。墙上贴着学院的招生公告。", "商业"),
    ("map_peace_town_market", "和平镇集市", "和平镇", "", 15, 4, "yes", "",
     "和平镇唯一的主街上，小贩们向未来的学员兜售(保过秘籍)和(考前丹药)。真假难辨。", "商业"),

    # ── 迦南学院 ──
    ("map_canaan_trade_street", "迦南外院商业街", "迦南学院", "", 20, 3, "yes", "",
     "外院学生日常采购的街区，药材、丹药与简易装备一应俱全。火能兑换处门前排着长队。", "商业"),
    ("map_inner_market", "内院坊市", "迦南内院", "", 22, 4, "yes", "",
     "内院学生以火能交易的自由市场。摊位简陋，但偶尔能淘到强榜高手淘汰的好货。", "商业"),

    # ── 中州城市 ──
    ("map_tianbei_market", "天北城商业街", "天北城", "", 35, 4, "yes", "",
     "天北城最繁华的街道，韩家与洪家的商铺各占半条街。两家的竞争从货架延伸到价格。", "商业"),
    ("map_tianbei_inn", "天北城客栈", "天北城", "", 35, 3, "yes", "",
     "旅人抵达天北城后的第一落脚点。老板与四方阁的弟子混得熟，打听中州局势的最佳场所。", "商业"),
    ("map_ye_market", "叶城坊市", "叶城", "", 40, 4, "yes", "",
     "叶城中心交易区，叶家丹药铺占据了最好的位置。丹阁期间，这里变成了炼药师的小型交流会场。", "商业"),
    ("map_tianhuang_market", "天黄城坊市", "天黄城", "", 52, 4, "yes", "",
     "中域枢纽天黄城的繁华坊市，四方货物汇集。价格不菲，但品类齐全——在中域没有买不到的东西。", "商业"),
    ("map_tianhuang_inn", "天黄城客栈", "天黄城", "", 52, 3, "yes", "",
     "常年满员的客栈，住客多是等待虫洞的旅人。大堂留言板上钉满了找队友和换路线的便条。", "商业"),

    # ── 莽荒古域 ──
    ("map_manghuang_inn", "莽荒镇客栈", "莽荒镇", "", 62, 3, "yes", "",
     "古域外唯一的客栈，墙壁被各路强者的笔迹刻满——有人留下古域地图，有人写下遗言。", "商业"),

    # ── 大岭城 ──
    ("map_daling_market", "大岭城集市", "大岭城", "", 8, 4, "yes", "",
     "边陲重镇的露天集市，商队在此进行进出帝国的最后补给。边境特产与走私货混杂其间。", "商业"),
    ("map_daling_inn", "大岭城客栈", "大岭城", "", 8, 3, "yes", "",
     "赶往暗角域的旅人在此歇脚，客栈里的每一张桌子都在讨论暗角域的凶险与机遇。", "商业"),

    # ── 宗门交易区 ──
    ("map_burning_valley_market", "焚炎谷坊市", "焚炎谷", "", 40, 5, "yes", "",
     "焚炎谷外围的交易区，谷中炼器师出售火属性武器与护甲。偶尔能买到唐震亲手淬炼的火纹钢。", "商业"),
    ("map_star_pavilion_market", "星落阁坊市", "星落阁", "", 40, 4, "yes", "",
     "星落阁为弟子与访客设立的交易区，丹药品质上乘——毕竟是玄炉老人的门派。", "商业"),
    ("map_flower_sect_market", "花宗坊市", "花宗", "", 50, 4, "yes", "",
     "花宗外围的灵花交易市场，宗内弟子在此出售以灵花炼制的养颜丹与回春露。在中州女修中极受欢迎。", "商业"),
]

# ══════════════════════════════════════════════════════════
# PART 3: ENCOUNTERS FOR NEW COMMERCE AREAS
# ══════════════════════════════════════════════════════════
def enc(evt_id, map_id, text, opt1_text="", opt1_effect="",
        opt2_text="", opt2_effect="", opt3_text="", opt3_effect="", weight=1):
    return {
        "Event_ID": evt_id, "Map_ID": map_id, "Weight": weight,
        "Text": text, "Pre_Condition": "",
        "Opt1_Text": opt1_text, "Opt1_Condition": "", "Opt1_Next": "", "Opt1_Effect": opt1_effect,
        "Opt2_Text": opt2_text, "Opt2_Condition": "", "Opt2_Next": "", "Opt2_Effect": opt2_effect,
        "Opt3_Text": opt3_text, "Opt3_Condition": "", "Opt3_Next": "", "Opt3_Effect": opt3_effect,
        "Notes": "",
    }

NEW_ENCOUNTERS = [
    # ── 黑岩城客栈 ──
    enc("enc_br_inn_1", "map_black_rock_inn",
        "客栈大堂里，一个商队领队正在招募前往魔兽山脉的护卫。他的商队满载准备运往帝都的皮毛和魔核。",
        "应聘护卫", "exp:+40,silver:+15", "打听魔兽山脉情报", "soul:+2"),
    # ── 漠城商栈 ──
    enc("enc_mo_trade_1", "map_mo_trade_post",
        "商栈老板正在清点一批从沙漠深处运来的古物。其中一件陶罐底下压着半张残缺的兽皮地图。",
        "买下残图", "silver:-20,soul:+3", "帮老板清点货物", "silver:+10"),
    # ── 石漠城客栈 ──
    enc("enc_sm_inn_1", "map_stone_mo_inn",
        "客栈墙上密密麻麻写满了前人的沙漠生存笔记。其中一行字迹格外醒目：'蛇人族绿洲以东二十里，古井见火莲。'",
        "记下这条线索", "soul:+3", "向老板打听笔记作者", "soul:+2"),
    # ── 盐城客栈 ──
    enc("enc_salt_inn_1", "map_salt_inn",
        "盐帮的账房先生喝醉了酒，正在客栈角落里吹嘘自己知道盐帮帮主的秘密交易。周围几个人竖起了耳朵。",
        "给账房再添一壶酒", "silver:-5,soul:+5", "假装没听见", ""),

    # ── 黑皇城坊市 ──
    enc("enc_be_market_1", "map_black_emperor_market",
        "坊市中一家老字号丹药铺的老板神秘地拉开抽屉：'这可是黑皇宗长老私下炼制的破宗丹——虽然只成功了七成，但药效绝对不差。'",
        "高价购买", "silver:-40,item:+item_elixir", "要求验货", "alchemy:+2"),
    # ── 黑印城客栈 ──
    enc("enc_bs_inn_1", "map_black_seal_inn",
        "客栈角落里，两个刚参加完拍卖的买家正在低声争吵——他们合伙拍下了一件异宝，却在分赃时闹翻。",
        "居中调停", "reputation:+5,soul:+2", "趁机提出购买", "silver:-30,exp:+40"),
    # ── 枫城坊市 ──
    enc("enc_feng_market_1", "map_feng_market",
        "坊市中有人在收购海心焰的消息，报酬丰厚但来路可疑。告示上画着一枚六品丹药作为酬劳。",
        "提供假消息套取丹药", "soul:+3,item:+item_elixir", "无视可疑告示", ""),
    # ── 和平镇客栈 ──
    enc("enc_pt_inn_1", "map_peace_town_inn",
        "一个刚从迦南学院入学考试中被刷下来的青年正对着酒壶发愁。他看到你，眼中燃起希望：'前辈，能指点一下修炼吗？'",
        "指点修炼", "reputation:+5,soul:+2", "鼓励他明年再考", "reputation:+3"),
    # ── 和平镇集市 ──
    enc("enc_pt_market_1", "map_peace_town_market",
        "小贩热情地推销一瓶'迦南学院内部考题'——价格不菲，但周围已经围了好几个心动的考生。",
        "买下考题", "silver:-15,soul:+1", "拆穿骗局", "reputation:+5"),

    # ── 迦南外院商业街 ──
    enc("enc_canaan_trade_1", "map_canaan_trade_street",
        "火能兑换处前排着长队。一个内院学长正在高价收购火能——他需要在天焚炼气塔闭关冲击瓶颈。",
        "出售多余火能", "exp:+30,silver:+20", "拒绝出售", ""),
    # ── 内院坊市 ──
    enc("enc_inner_market_1", "map_inner_market",
        "一个强榜前十的学长在摊位上出售他淘汰的玄阶灵技卷轴。'修炼到圆满就没什么用了，便宜卖了。'",
        "买下灵技", "silver:-30,exp:+50", "讨价还价", "silver:-20,exp:+40"),

    # ── 天北城商业街 ──
    enc("enc_tb_market_1", "map_tianbei_market",
        "韩家与洪家的商铺面对面打擂台——韩家的丹药刚降价，洪家就推出了买三赠一。顾客们乐见其成。",
        "趁降价囤货", "silver:-25,item:+item_elixir", "观察两家竞争", "soul:+2"),
    # ── 天北城客栈 ──
    enc("enc_tb_inn_1", "map_tianbei_inn",
        "客栈老板透露，风雷阁最近在招募散修参与一项秘密任务——据说与天目山脉的能量潮汐有关。",
        "打听任务详情", "soul:+3", "直接前往风雷阁", "exp:+30"),

    # ── 叶城坊市 ──
    enc("enc_ye_market_1", "map_ye_market",
        "丹阁期间，坊市变成了小型交流会。几个炼药师正围着一尊药鼎争论丹方中火候的控制技巧。",
        "加入讨论", "alchemy:+3,soul:+2", "旁听学习", "alchemy:+1"),
    # ── 天黄城坊市 ──
    enc("enc_th_market_1", "map_tianhuang_market",
        "天黄城坊市的一个云族摊位吸引了你的注意——摊主在出售远古遗迹中出土的功法残片，价格令人咋舌。",
        "购买功法残片", "silver:-50,soul:+5", "与摊主攀谈远古秘闻", "soul:+3"),
    # ── 天黄城客栈 ──
    enc("enc_th_inn_1", "map_tianhuang_inn",
        "客栈留言板上钉满了便条：有人找队友去丹域，有人转让半张藏宝图，还有人悬赏仇家下落。",
        "仔细阅读留言板", "soul:+3", "揭下一张感兴趣的任务", "exp:+60"),

    # ── 莽荒镇客栈 ──
    enc("enc_mh_inn_1", "map_manghuang_inn",
        "客栈墙上最醒目的位置刻着一行大字：'菩提树下，莫回头。'字迹陈旧，却散发着若有若无的灵魂威压。",
        "感悟字迹中的威压", "soul:+8", "向老板打听刻字之人", "soul:+3"),

    # ── 大岭城集市/客栈 ──
    enc("enc_dl_market_1", "map_daling_market",
        "边境集市上，一个从暗角域逃回来的伤者正在出售他的'生存经验'——一卷手绘的暗角域势力分布图。",
        "买下地图", "silver:-15,soul:+3", "向伤者打听详情", "soul:+2"),
    enc("enc_dl_inn_1", "map_daling_inn",
        "客栈中一个刚离开暗角域的佣兵正在讲述他的遭遇：'那鬼地方根本没有规则！实力就是一切！'",
        "请他喝酒详谈", "silver:-5,soul:+3", "默默旁听", "soul:+1"),

    # ── 焚炎谷坊市 ──
    enc("enc_bv_market_1", "map_burning_valley_market",
        "坊市铁匠铺中陈列着一柄以火山钢锻造的重剑，剑身上天然形成的火焰纹路让它在所有兵器中脱颖而出。",
        "购买火纹重剑", "silver:-40,atk:+5", "观摩锻造过程", "alchemy:+2"),
    # ── 星落阁坊市 ──
    enc("enc_sp_market_1", "map_star_pavilion_market",
        "玄炉老人亲传的几名弟子在坊市中出售高阶丹药——品质远超市面流通品，价格也相当公道。",
        "购买丹药", "silver:-20,item:+item_elixir", "请教炼丹心得", "alchemy:+2"),
    # ── 花宗坊市 ──
    enc("enc_fs_market_1", "map_flower_sect_market",
        "花宗弟子出售的养颜丹引发了一场竞价。一位男修红着脸为道侣抢购——周围的女修们善意地起哄。",
        "也买一瓶备用", "silver:-10,reputation:+1", "笑看热闹", ""),
]

# ══════════════════════════════════════════════════════════
# PART 4: GATHERING SPOTS WITH REGIONAL SPECIALTIES
# ══════════════════════════════════════════════════════════
def gather(gid, map_id, item_id, chance, min_q=1, max_q=3, pre="", effect="", limit="", exclusive="", notes=""):
    return {
        "Gather_ID": gid, "Map_ID": map_id, "Item_ID": item_id,
        "Chance_Percent": chance, "Min_Qty": min_q, "Max_Qty": max_q,
        "Pre_Condition": pre, "Explore_Effect": effect,
        "Unique_Limit": limit, "Location_Exclusive": exclusive, "Notes": notes,
    }

NEW_GATHERING = [
    # ── 沧澜帝国特产采集 ──
    gather("gather_wutan_specialty", "map_wutan_commercial_street", "item_wutan_specialty", 35, 1, 3, "", "", "", "沧澜帝国", "青石城特产"),
    gather("gather_rock_specialty", "map_black_rock_market", "item_wutan_specialty", 30, 1, 2, "", "", "", "沧澜帝国", "黑岩城特产"),
    gather("gather_passport_border", "map_jia_ma_border", "item_jia_ma_passport", 15, 1, 1, "", "", "", "沧澜帝国", "通关文牒"),
    gather("gather_passport_ghost", "map_ghost_pass", "item_jia_ma_passport", 20, 1, 1, "", "", "", "沧澜帝国", "通关文牒"),

    # ── 沙漠特产采集 ──
    gather("gather_desert_compass", "map_desert_camp", "item_desert_compass", 15, 1, 1, "", "", "", "赤沙荒漠", "沙漠星盘"),
    gather("gather_snake_venom", "map_snake_oasis", "item_snake_venom_vial", 25, 1, 2, "", "", "", "蛇人族", "毒液瓶"),
    gather("gather_snake_venom_temple", "map_snake_temple_outer", "item_snake_venom_vial", 30, 1, 3, "", "", "", "蛇人族", "毒液瓶"),

    # ── 魔兽山脉特产采集 ──
    gather("gather_beast_hide", "map_magic_inner", "item_beast_hide_bundle", 30, 1, 3, "", "", "", "魔兽山脉", "皮料包"),
    gather("gather_mountain_herb", "map_magic_herb_valley", "item_mountain_herb_kit", 25, 1, 2, "", "", "", "魔兽山脉", "采药包"),

    # ── 暗角域特产采集 ──
    gather("gather_black_market_pass", "map_black_seal_auction", "item_black_market_pass", 10, 1, 1, "", "", "", "暗角域", "黑市令牌"),
    gather("gather_forbidden_pill", "map_black_herb_market", "item_forbidden_pill_fragment", 5, 1, 1, "", "", "", "暗角域", "禁药残页"),

    # ── 迦南学院特产 ──
    gather("gather_canaan_token", "map_canaan_mission_hall", "item_canaan_token", 40, 5, 20, "", "", "", "迦南学院", "火能卡"),
    gather("gather_inner_token", "map_inner_trade_district", "item_canaan_token", 35, 5, 15, "", "", "", "迦南学院", "火能卡"),

    # ── 中州特产采集 ──
    gather("gather_zhongzhou_map", "map_zhongzhou_transfer_square", "item_zhongzhou_map", 20, 1, 1, "", "", "", "中州", "势力地图"),
    gather("gather_wormhole_pass_tianya", "map_tianya_wormhole_square", "item_wormhole_pass", 15, 1, 1, "", "", "", "中州", "虫洞通行令"),
    gather("gather_wormhole_pass_th", "map_tianhuang_market", "item_wormhole_pass", 15, 1, 1, "", "", "", "中州", "虫洞通行令"),

    # ── 丹域特产采集 ──
    gather("gather_dan_herb_box", "map_dan_herb_street", "item_dan_herb_box", 20, 1, 2, "", "", "", "丹域", "药材精装"),
    gather("gather_alchemy_handbook", "map_sacred_dan_market", "item_alchemy_handbook", 15, 1, 1, "", "", "", "丹域", "炼药手札"),

    # ── 兽域特产采集 ──
    gather("gather_beast_blood", "map_beast_market", "item_beast_blood_essence", 25, 1, 2, "", "", "", "兽域", "魔兽精血"),
    gather("gather_transforming_herb", "map_beast_bone_mountains", "item_transforming_herb", 10, 1, 1, "", "", "", "兽域", "化形草"),

    # ── 龙岛特产采集 ──
    gather("gather_dragon_scale", "map_dragon_island_harbor", "item_dragon_scale", 15, 1, 1, "", "", "", "龙岛", "龙鳞碎片"),
    gather("gather_dragon_saliva", "map_east_dragon_island", "item_dragon_saliva_herb", 12, 1, 1, "", "", "", "龙岛", "龙涎灵草"),

    # ── 莽荒古域特产 ──
    gather("gather_ancient_seed", "map_wilderness_outpost", "item_ancient_seed", 10, 1, 1, "", "", "", "莽荒古域", "古域灵种"),
    gather("gather_bodhi_leaf", "map_bodhi_tree", "item_bodhi_leaf", 5, 1, 1, "", "", "", "菩提古树", "菩提叶"),

    # ── 妖火空间 ──
    gather("gather_demon_crystal", "map_demon_flame_core", "item_demon_flame_crystal", 15, 1, 2, "", "", "", "妖火空间", "妖火结晶"),

    # ── 远古遗迹 ──
    gather("gather_ancient_relic", "map_ancient_ruins_core", "item_ancient_relic", 12, 1, 1, "", "", "", "远古遗迹", "远古遗物"),
]

# ══════════════════════════════════════════════════════════
# APPLY ALL
# ══════════════════════════════════════════════════════════
print(f"Adding {len(REGION_ITEMS)} regional specialty items...")
for data in REGION_ITEMS:
    item_id = data[0]
    if item_id in existing_items:
        print(f"  SKIP item: {item_id}")
        continue
    row = {
        "Item_ID": data[0], "Name": data[1], "Type": data[2],
        "Rarity": data[3], "Stackable": data[4], "Max_Stack": data[5],
        "Buy_Price": data[6], "Sell_Price": data[7],
        "Use_Condition": data[8], "Use_Effect": data[9],
        "Description": data[10], "Notes": data[11],
    }
    append_row(ws_items, row)
    existing_items.add(item_id)
    print(f"  + item: {item_id}: {data[1]}")

print(f"\nAdding {len(NEW_COMMERCE_MAPS)} new commerce sub-areas...")
for data in NEW_COMMERCE_MAPS:
    map_id = data[0]
    if map_id in existing_maps:
        print(f"  SKIP map: {map_id}")
        continue
    row = {
        "Map_ID": data[0], "Name": data[1], "Region": data[2],
        "Unlock_Condition": data[3], "Recommend_Level": data[4],
        "Explore_Stamina_Cost": data[5], "Default_Encounter_Weight": 1,
        "Safe_Zone": data[6], "Exit_Event": data[7],
        "Description": data[8], "Notes": data[9],
    }
    append_row(ws_maps, row)
    existing_maps.add(map_id)
    print(f"  + map: {map_id}: {data[1]}")

print(f"\nAdding {len(NEW_ENCOUNTERS)} new encounters...")
for data in NEW_ENCOUNTERS:
    evt_id = data["Event_ID"]
    if evt_id in existing_encounters:
        print(f"  SKIP enc: {evt_id}")
        continue
    append_row(ws_encounters, data)
    existing_encounters.add(evt_id)
    print(f"  + enc: {evt_id}")

print(f"\nAdding {len(NEW_GATHERING)} new gathering spots...")
for data in NEW_GATHERING:
    gid = data["Gather_ID"]
    if gid in existing_gathering:
        print(f"  SKIP gather: {gid}")
        continue
    append_row(ws_gathering, data)
    existing_gathering.add(gid)
    print(f"  + gather: {gid}")

wb.save(WORKBOOK_PATH)
print(f"\nSaved! Totals:")
print(f"  Items: +{len(REGION_ITEMS)} -> {len(existing_items)}")
print(f"  Maps: +{len(NEW_COMMERCE_MAPS)} -> {len(existing_maps)}")
print(f"  Encounters: +{len(NEW_ENCOUNTERS)} -> {len(existing_encounters)+len(NEW_ENCOUNTERS)}")
print(f"  Gathering: +{len(NEW_GATHERING)}")
print("Done!")
