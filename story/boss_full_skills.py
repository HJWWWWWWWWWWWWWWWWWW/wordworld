"""
Boss full skill design: minimum 10 skills per boss.
Skills assigned by: level tier + faction + element + unique signature.
"""
from pathlib import Path
import openpyxl

WORKBOOK_PATH = Path("story/text_game_event_schema_v4.xlsx")
wb = openpyxl.load_workbook(WORKBOOK_PATH)

# Find sheets
ws_skills = ws_enemies = None
for name in wb.sheetnames:
    if name.startswith("Skills_"): ws_skills = wb[name]
    if name.startswith("Enemies_"): ws_enemies = wb[name]

skill_headers = [cell.value for cell in next(ws_skills.iter_rows(min_row=1, max_row=1))]
enemy_headers = [cell.value for cell in next(ws_enemies.iter_rows(min_row=1, max_row=1))]
if "Skills" not in enemy_headers:
    ws_enemies.cell(row=1, column=len(enemy_headers)+1).value = "Skills"
    enemy_headers.append("Skills")
skills_col = enemy_headers.index("Skills")

existing_skills = set()
for row in ws_skills.iter_rows(min_row=2):
    if row[0].value: existing_skills.add(str(row[0].value))

def add_skill(sid, name, stype, rank, effect, desc):
    if sid in existing_skills: return
    row = {"Skill_ID": sid, "Name": name, "Type": stype, "Rank": rank,
           "Effect": effect, "Description": desc}
    ws_skills.append([str(row.get(h, "")) for h in skill_headers])
    existing_skills.add(sid)

# ══════════════════════════════════════════════════════════
# ADDITIONAL GENERIC SKILLS (for boss skill pools)
# ══════════════════════════════════════════════════════════
NEW_GENERIC = [
    # 战斗基础 (Lv 1-20) - every cultivator learns these
    ("skill_qi_shield", "斗气盾", "防御", "黄阶高级", "def:+30",
     "以斗气凝聚护盾格挡伤害。修炼者的基础防御技能。"),
    ("skill_qi_burst", "斗气爆发", "辅助", "玄阶低级", "atk:+25,spd:+15",
     "瞬间爆发斗气提升战力。中阶修炼者通用的战斗技巧。"),
    ("skill_iron_body", "铜体术", "防御", "玄阶低级", "def:+35,spd:-10",
     "以斗气强化肉身防御。佣兵和体修偏好的防御技能。"),
    ("skill_spirit_sense", "灵觉", "辅助", "玄阶低级", "spd:+20,soul:+5",
     "释放灵魂感知预判对手行动。炼药师和魂修的基础能力。"),

    # 身法 (Lv 10-35)
    ("skill_shadow_step", "影步", "身法", "玄阶中级", "spd:+35",
     "高速移动留下残影。各大势力中层弟子的标配身法。"),
    ("skill_soaring_sky", "踏空步", "身法", "玄阶高级", "spd:+40",
     "借助斗气在空中短暂行走。斗王以上强者的标志性能力。"),
    ("skill_lightning_flash", "电光一闪", "身法", "地阶低级", "spd:+55",
     "以雷电之力驱动极致速度。风雷阁及雷系强者的高级身法。"),

    # 攻击进阶 (Lv 20-50)
    ("skill_qi_explosion", "斗气炸裂", "物理", "玄阶高级", "atk:+40",
     "将斗气压缩后引爆造成范围伤害。斗灵以上修炼者的强力攻击。"),
    ("skill_mountain_splitter", "开山裂石", "物理", "玄阶高级", "atk:+45,spd:-5",
     "蓄力发出开山裂石般的重击。力量型修炼者的偏爱。"),
    ("skill_hundred_fists", "百拳连打", "物理", "玄阶中级", "atk:+30,spd:+10",
     "高速连打百拳压制对手。近战型修炼者的常见连招。"),
    ("skill_air_slash", "空斩", "物理", "地阶低级", "atk:+50",
     "凝聚斗气于兵刃发出隔空斩击。斗王以上强者可轻易施展。"),

    # 元素进阶 (Lv 25-60)
    ("skill_fire_storm", "火焰风暴", "火系", "地阶低级", "atk:+55",
     "召唤大范围火焰风暴。斗皇以上火系强者的战场杀招。"),
    ("skill_ice_prison", "冰封牢笼", "冰系", "地阶低级", "atk:+45,spd:-20",
     "以极寒斗气冻结目标。冰系强者的控制技能。"),
    ("skill_wind_explosion", "风爆", "风系", "地阶低级", "atk:+50,spd:+15",
     "压缩风属性斗气产生剧烈爆炸。风系强者的进阶攻击。"),
    ("skill_earth_wall", "大地之壁", "土系", "地阶低级", "def:+50",
     "从大地升起岩壁防御。土系修炼者的强大防御技。"),
    ("skill_thunder_rain", "雷霆暴雨", "雷系", "地阶中级", "atk:+65,spd:+10",
     "召唤雷霆暴雨覆盖战场。雷系强者的范围攻击。"),
    ("skill_toxic_explosion", "毒爆", "毒系", "地阶低级", "atk:+50,poison:+20",
     "引爆毒雾产生剧毒冲击。毒师的高阶杀伤技能。"),

    # 暗/灵魂进阶 (Lv 30-70)
    ("skill_dark_erosion", "暗蚀", "暗系", "地阶低级", "atk:+50,soul:+15",
     "以黑暗斗气侵蚀目标防御。魂殿及暗系修炼者的进阶技能。"),
    ("skill_soul_blast", "灵魂爆破", "灵魂", "地阶中级", "atk:+60,soul:+25",
     "引爆灵魂能量造成毁灭性伤害。魂殿高阶护法的杀招。"),
    ("skill_mind_break", "精神碎裂", "灵魂", "地阶中级", "atk:+55,soul:+20",
     "以灵魂之力粉碎目标精神。魂殿尊者的强力攻击。"),
    ("skill_shadow_dance", "暗影之舞", "暗系", "地阶中级", "spd:+60,atk:+20",
     "融入暗影中高速穿梭攻击。暗杀系强者的高级身法。"),

    # 龙系进阶 (Lv 50-90)
    ("skill_dragon_scale_armor", "龙鳞甲", "龙系", "地阶高级", "def:+60,hp:+100",
     "以太虚古龙鳞片凝聚防御甲。龙族强者的防御技能。"),
    ("skill_dragon_wing_slash", "龙翼斩", "龙系", "天阶低级", "atk:+75,spd:+20",
     "化为龙翼发动高速斩击。高阶龙族的强力技能。"),
    ("skill_void_tear", "虚空撕裂", "龙系", "天阶中级", "atk:+100",
     "以龙爪撕裂虚空造成空间伤害。龙皇级别的终极攻击。"),

    # 远古进阶 (Lv 55-95)
    ("skill_ancient_roar", "远古咆哮", "远古", "天阶低级", "atk:+70,soul:+20",
     "发出远古咆哮震慑灵魂。远古种族和远古魔兽的能力。"),
    ("skill_emperor_pressure", "帝威", "远古", "天阶高级", "atk:+90,spd:-20",
     "释放斗帝级别的威压。陀舍古帝和魂天帝级别的终极能力。"),
    ("skill_time_space_lock", "时空封锁", "远古", "天阶高级", "atk:+85,soul:+40",
     "封锁时空使目标无法行动。远古斗帝的逆天能力。"),

    # 治疗/辅助进阶 (Lv 30-80)
    ("skill_self_regen", "自愈术", "辅助", "地阶低级", "hp:+150",
     "以斗气催动肉身高速自愈。高阶修炼者的续航技能。"),
    ("skill_battle_meditation", "战意冥想", "辅助", "地阶中级", "atk:+30,soul:+20,hp:+50",
     "战斗中进入冥想状态全面提升。药族和丹塔高层的独有技能。"),

    # 体力管理
    ("skill_stamina_save", "省力战法", "辅助", "玄阶中级", "atk:+10,def:+10",
     "以最小消耗维持战斗的实用技巧。经验丰富的佣兵和冒险者常用。"),
]

print(f"Adding {len(NEW_GENERIC)} generic skills...")
for data in NEW_GENERIC:
    add_skill(*data)
print(f"Total skills: {len(existing_skills)}")

# ══════════════════════════════════════════════════════════
# SKILL TIER SYSTEM
# Each tier maps to a level range. Bosses get all tiers up to their level.
# ══════════════════════════════════════════════════════════

# TIER 1: Basic (all bosses Lv 10+)
TIER1_BASIC = [
    "skill_body_fortify", "skill_qi_shield", "skill_sword_qi",
    "skill_palm_strike", "skill_qi_burst", "skill_spirit_sense",
]

# TIER 2: Movement/Defense (Lv 20+)
TIER2_MOVEMENT = [
    "skill_shadow_step", "skill_defensive_stance", "skill_iron_body",
    "skill_stamina_save", "skill_battle_cry",
]

# TIER 3: Advanced Combat (Lv 35+)
TIER3_ADVANCED = [
    "skill_qi_explosion", "skill_air_slash", "skill_soaring_sky",
    "skill_hundred_fists", "skill_counter_strike",
]

# TIER 4: High Combat (Lv 50+)
TIER4_HIGH = [
    "skill_mountain_splitter", "skill_self_regen", "skill_battle_meditation",
    "skill_lightning_flash", "skill_soul_blast",
]

# TIER 5: Ultimate (Lv 70+)
TIER5_ULTIMATE = [
    "skill_ancient_roar", "skill_emperor_pressure", "skill_time_space_lock",
    "skill_void_tear", "skill_mind_break", "skill_shadow_dance",
]

# ELEMENT TIERS by attribute
FIRE_SKILLS = ["skill_fire_spit", "skill_fire_serpent", "skill_fire_storm", "skill_inferno", "skill_green_lotus_burn"]
ICE_SKILLS = ["skill_ice_breath", "skill_ice_dragon", "skill_ice_prison", "skill_absolute_zero"]
WIND_SKILLS = ["skill_wind_slash", "skill_wind_push", "skill_wind_explosion", "skill_cyclone"]
EARTH_SKILLS = ["skill_earth_shake", "skill_earth_wall", "skill_earthquake", "skill_sand_king"]
THUNDER_SKILLS = ["skill_lightning_bolt", "skill_thunder_clap", "skill_thunder_rain", "skill_thunder_wrath"]
POISON_SKILLS = ["skill_venom_inject", "skill_toxic_mist", "skill_toxic_explosion", "skill_corrosive_acid"]
SOUL_SKILLS = ["skill_soul_absorption", "skill_soul_shock", "skill_soul_pain", "skill_soul_rend", "skill_life_drain"]
DRAGON_SKILLS = ["skill_dragon_breath_attack", "skill_dragon_roar", "skill_dragon_scale_armor", "skill_dragon_claw_void"]
ANCIENT_SKILLS = ["skill_bloodline_pressure", "skill_ancient_seal_art", "skill_emperor_will"]

# FACTION-SPECIFIC skills
YUNLAN_SKILLS = ["skill_wind_wall", "skill_cloud_wind_sword", "skill_wind_extreme_fall", "skill_wind_bind", "skill_wind_sword"]
BLOOD_SKILLS = ["skill_blood_art", "skill_blood_transform", "skill_life_drain"]
SOUL_HALL_SKILLS = ["skill_soul_handprint", "skill_soul_chain", "skill_soul_art", "skill_ten_thousand_souls", "skill_soul_domain"]
DRAGON_TRIBE_SKILLS = ["skill_dragon_power", "skill_dragon_phoenix_armor", "skill_ancient_dragon_roar", "skill_dragon_wing_slash"]

def assign_boss_skills(eid, name, level, faction, element, unique_skills):
    """Assign 10+ skills to a boss based on level, faction, element, and unique skills."""
    lv = int(level)
    pool = list(unique_skills)  # Start with unique skills

    # Add faction skills (all)
    for fs in faction:
        pool.extend(fs)

    # Add element skills by level
    for es in element:
        if lv >= 15: pool.append(es[0])  # Basic element
        if lv >= 25: pool.extend(es[1:2])  # Mid element
        if lv >= 40: pool.extend(es[2:3])  # High element
        if lv >= 60: pool.extend(es[3:4])  # Ultimate element

    # Add tiered generic skills
    if lv >= 10: pool.extend(TIER1_BASIC)
    if lv >= 20: pool.extend(TIER2_MOVEMENT)
    if lv >= 35: pool.extend(TIER3_ADVANCED)
    if lv >= 50: pool.extend(TIER4_HIGH)
    if lv >= 70: pool.extend(TIER5_ULTIMATE)

    # Deduplicate and limit
    seen = []
    result = []
    for s in pool:
        if s not in seen:
            seen.append(s)
            result.append(s)

    # Ensure at least 10
    if len(result) < 10:
        extras = ["skill_beast_charge", "skill_beast_roar", "skill_pill_boost",
                  "skill_meditate", "skill_heavy_strike", "skill_arena_rage"]
        for ex in extras:
            if ex not in seen:
                result.append(ex)
                if len(result) >= 10:
                    break

    return result

# ══════════════════════════════════════════════════════════
# BOSS SKILL ASSIGNMENTS (10+ each)
# ══════════════════════════════════════════════════════════
BOSS_SKILLS = {
    # ── 加玛帝国 (Lv 12-24) ──
    "boss_mu_she": (12, [], [EARTH_SKILLS], ["skill_lion_mountain_split"]),
    "enemy_nalan": (15, [YUNLAN_SKILLS], [WIND_SKILLS], ["skill_wind_sword", "skill_cloud_wind_sword", "skill_wind_push"]),
    "boss_twin_head_fire_snake": (15, [], [FIRE_SKILLS, POISON_SKILLS], ["skill_fire_serpent"]),
    "boss_amethyst_wing_lion": (18, [], [WIND_SKILLS], ["skill_beast_roar"]),
    "boss_sand_thief_king": (18, [], [EARTH_SKILLS], ["skill_sand_king", "skill_shadow_strike"]),
    "enemy_hai": (18, [], [ICE_SKILLS], ["skill_ice", "skill_ice_dragon", "skill_ice_sword_art"]),
    "boss_desert_ancient_scorpion": (20, [], [POISON_SKILLS, EARTH_SKILLS], ["skill_ten_thousand_scorpion_master"]),
    "boss_guard_beast_of_temple": (20, [], [EARTH_SKILLS], ["skill_bloodline_pressure", "skill_ancient_seal_art"]),
    "enemy_yun_leng": (22, [YUNLAN_SKILLS], [WIND_SKILLS], ["skill_wind_wall", "skill_cloud_wind_sword"]),
    "boss_snake_high_priest": (22, [], [POISON_SKILLS], ["skill_snake_seal", "skill_snake_eye"]),
    "boss_fire_ape_king": (22, [], [FIRE_SKILLS], ["skill_beast_roar", "skill_earth_shake"]),
    "enemy_fan_ling": (24, [BLOOD_SKILLS], [], ["skill_blood_art", "skill_blood_transform"]),

    # ── 黑角域 (Lv 25-35) ──
    "enemy_medusa": (25, [], [POISON_SKILLS], ["skill_snake_seal", "skill_snake_eye", "skill_colorful_snake_scale"]),
    "boss_eight_gates_master": (26, [], [], ["skill_eight_gates_palm", "skill_underground_authority"]),
    "boss_inner_academy_guardian": (26, [], [FIRE_SKILLS], ["skill_skyfire_guard"]),
    "enemy_fan_lao": (28, [BLOOD_SKILLS], [], ["skill_blood_art", "skill_blood_transform", "skill_life_drain"]),
    "boss_blood_sect_leader": (28, [BLOOD_SKILLS], [], ["skill_blood_art", "skill_blood_transform"]),
    "enemy_yun_shan": (30, [YUNLAN_SKILLS], [WIND_SKILLS], ["skill_wind_sword", "skill_wind_extreme_fall", "skill_wind_wall", "skill_wind_bind"]),
    "boss_black_corner_arena_king": (30, [], [], ["skill_arena_rage", "skill_beast_charge"]),
    "boss_skyfire_tower_beast": (30, [], [FIRE_SKILLS], ["skill_fallen_heart_pulse", "skill_beast_roar"]),
    "enemy_wu_hufa": (32, [SOUL_HALL_SKILLS], [SOUL_SKILLS], ["skill_soul_handprint", "skill_soul_chain"]),
    "boss_black_alliance_commander": (32, [], [], ["skill_battle_cry", "skill_bloodline_pressure"]),
    "boss_black_seal_underground_lord": (32, [], [], ["skill_underground_authority", "skill_eight_gates_palm"]),
    "boss_skyfire_beast_king": (32, [], [FIRE_SKILLS], ["skill_fallen_heart_pulse", "skill_beast_roar"]),
    "enemy_han_feng": (35, [], [FIRE_SKILLS], ["skill_sea_flame", "skill_sea_flame_tide", "skill_flame_splitting_ruler", "skill_pill_boost"]),
    "enemy_mo_tianxing": (35, [], [SOUL_SKILLS], ["skill_soul_chain", "skill_soul_shock"]),

    # ── 出云帝国 (Lv 32-42) ──
    "enemy_xie_biyan": (32, [], [POISON_SKILLS], ["skill_ten_thousand_scorpion_master", "skill_ten_thousand_scorpion_array"]),
    "boss_golden_goose_leader": (36, [], [WIND_SKILLS], ["skill_golden_goose_formation"]),
    "boss_mulan_valley_leader": (36, [], [], ["skill_mulan_herb_art", "skill_pill_boost"]),
    "enemy_di_mo_laogui": (38, [], [SOUL_SKILLS], ["skill_demon_puppet", "skill_life_drain"]),

    # ── 魂殿中层 (Lv 42-65) ──
    "enemy_hun_yao": (65, [SOUL_HALL_SKILLS], [SOUL_SKILLS], ["skill_soul_handprint"]),
    "enemy_hun_xuzi": (75, [SOUL_HALL_SKILLS], [SOUL_SKILLS], ["skill_soul_art"]),
    "enemy_hun_yu": (70, [SOUL_HALL_SKILLS], [SOUL_SKILLS], ["skill_soul_art", "skill_soul_handprint"]),
    "enemy_hun_feng": (75, [SOUL_HALL_SKILLS], [SOUL_SKILLS], ["skill_soul_art", "skill_soul_handprint", "skill_soul_rend"]),
    "enemy_hunmo": (80, [SOUL_HALL_SKILLS], [SOUL_SKILLS, FIRE_SKILLS], ["skill_fire_poison_dual"]),

    # ── 中州 (Lv 48-72) ──
    "boss_thunder_emperor": (48, [], [WIND_SKILLS, THUNDER_SKILLS], ["skill_wind_thunder", "skill_thunder_arc_dance", "skill_thunder"]),
    "boss_sword_emperor": (48, [], [], ["skill_sword", "skill_sword_qi"]),
    "enemy_mugu": (50, [], [FIRE_SKILLS], ["skill_bone_chilling_flame", "skill_pill_boost", "skill_alchemy"]),
    "boss_black_fire_ancestor": (50, [], [FIRE_SKILLS, POISON_SKILLS], ["skill_black_fire_ancestor_flame"]),
    "boss_sky_demon_sect_leader": (52, [SOUL_HALL_SKILLS], [SOUL_SKILLS], ["skill_soul_art"]),
    "enemy_zhai_xing": (55, [SOUL_HALL_SKILLS], [SOUL_SKILLS], ["skill_soul_handprint", "skill_soul_chain"]),
    "boss_dan_tower_guardian": (55, [], [FIRE_SKILLS], ["skill_pill_boost", "skill_alchemy", "skill_ancient_seal_art"]),
    "boss_wan_yao_mountain_lord": (56, [], [], ["skill_wan_yao_mountain_lord_art", "skill_pill_boost"]),
    "boss_nether_python_usurper": (58, [], [POISON_SKILLS, SOUL_SKILLS], ["skill_nether_python_usurp"]),
    "boss_space_trade_master": (58, [], [], ["skill_space_trade_master_art"]),
    "boss_manghuang_mayor": (60, [], [], ["skill_manghuang_survival"]),
    "boss_sky_phoenix_elder": (62, [], [WIND_SKILLS], ["skill_sky_phoenix_elder_art"]),

    # ── 龙岛 (Lv 65-95) ──
    "boss_south_dragon_king": (65, [DRAGON_TRIBE_SKILLS], [DRAGON_SKILLS], ["skill_dragon_power"]),
    "boss_ancient_domain_beast": (66, [], [ANCIENT_SKILLS], ["skill_ancient_domain_beast_king"]),
    "boss_west_dragon_king": (66, [DRAGON_TRIBE_SKILLS], [DRAGON_SKILLS], ["skill_dragon_power"]),
    "boss_demon_python_king": (68, [], [POISON_SKILLS, DRAGON_SKILLS[:2]], ["skill_bloodline_pressure"]),
    "boss_north_dragon_king": (68, [DRAGON_TRIBE_SKILLS], [DRAGON_SKILLS], ["skill_dragon_power", "skill_dragon_claw_void"]),
    "enemy_fang_yan": (30, [], [FIRE_SKILLS, POISON_SKILLS], ["skill_fire_poison_dual", "skill_demon_fire_array"]),
    "boss_bodhi_tree_guardian": (70, [], [ANCIENT_SKILLS], ["skill_healing", "skill_ancient_seal_art"]),
    "boss_demon_flame_saint_spirit": (70, [], [FIRE_SKILLS, ANCIENT_SKILLS], ["skill_purifying_saint", "skill_bloodline_pressure"]),
    "boss_yao_clan_guardian": (70, [], [], ["skill_alchemy", "skill_pill_boost", "skill_ancient_seal_art"]),
    "boss_small_dan_tower_elder": (72, [], [FIRE_SKILLS], ["skill_small_dan_elder_art", "skill_alchemy", "skill_pill_boost"]),
    "boss_ancient_dragon_spirit_king": (72, [DRAGON_TRIBE_SKILLS], [DRAGON_SKILLS], ["skill_ancient_dragon_spirit_roar"]),

    # ── 远古 (Lv 72-86) ──
    "boss_gu_clan_general": (72, [], [ANCIENT_SKILLS], ["skill_gu_general_art"]),
    "boss_yan_clan_elder_chief": (74, [], [FIRE_SKILLS, ANCIENT_SKILLS], ["skill_yan_clan_elder_art"]),
    "boss_lei_clan_elder_chief": (74, [], [THUNDER_SKILLS, ANCIENT_SKILLS], ["skill_lei_clan_elder_art"]),
    "enemy_hun_miesheng": (70, [SOUL_HALL_SKILLS], [SOUL_SKILLS], ["skill_soul_handprint", "skill_ten_thousand_souls", "skill_soul_domain", "skill_soul_chain"]),
    "boss_soul_hall_three_elders": (76, [SOUL_HALL_SKILLS], [SOUL_SKILLS], ["skill_three_heavenly_elders_art", "skill_soul_pain"]),
    "boss_soul_hall_vice_leader": (78, [SOUL_HALL_SKILLS], [SOUL_SKILLS], ["skill_soul_hall_vice_art", "skill_soul_handprint", "skill_ten_thousand_souls"]),
    "enemy_hun_shengtian": (82, [SOUL_HALL_SKILLS], [SOUL_SKILLS, ANCIENT_SKILLS], ["skill_soul_art", "skill_soul_rend"]),
    "boss_emperor_cave_dragon": (82, [DRAGON_TRIBE_SKILLS], [DRAGON_SKILLS], ["skill_emperor_cave_dragon_guard"]),
    "boss_hun_clan_ritual_master": (82, [SOUL_HALL_SKILLS], [SOUL_SKILLS, ANCIENT_SKILLS], ["skill_soul_art", "skill_soul_chain"]),
    "enemy_hun_yuantian": (85, [SOUL_HALL_SKILLS], [SOUL_SKILLS, ANCIENT_SKILLS], ["skill_soul_art", "skill_soul_rend"]),
    "boss_gu_clan_three_immortals": (85, [], [ANCIENT_SKILLS], ["skill_ancient_emperor_seal", "skill_open_mountain_seal"]),
    "boss_yan_clan_flame_emperor": (86, [], [FIRE_SKILLS, ANCIENT_SKILLS], ["skill_gold_flame", "skill_gold_emperor_burn"]),
    "boss_lei_clan_thunder_emperor": (86, [], [THUNDER_SKILLS, ANCIENT_SKILLS], ["skill_thunder"]),
    "boss_burial_sky_guardian": (86, [], [SOUL_SKILLS, ANCIENT_SKILLS], ["skill_burial_sky_art", "skill_ancient_seal_art"]),

    # ── 四魔圣/最终 (Lv 88-99) ──
    "enemy_four_demon_saints": (88, [SOUL_HALL_SKILLS], [SOUL_SKILLS, ANCIENT_SKILLS], ["skill_soul_art"]),
    "boss_emperor_pill_guardian": (88, [], [ANCIENT_SKILLS], ["skill_emperor_pill_guard", "skill_bloodline_pressure"]),
    "boss_hun_army_general": (90, [SOUL_HALL_SKILLS], [SOUL_SKILLS], ["skill_soul_art", "skill_soul_rend"]),
    "boss_alliance_commander": (90, [], [ANCIENT_SKILLS], ["skill_alliance_commander_art"]),
    "boss_tuoshe_emperor_remnant": (92, [], [FIRE_SKILLS, ANCIENT_SKILLS], ["skill_emperor_flame", "skill_emperor_unity_flame"]),
    "enemy_xuwu": (95, [], [SOUL_SKILLS, ANCIENT_SKILLS], ["skill_nihility_flame", "skill_nihility_swallow"]),
    "boss_dragon_emperor_zhu_kun": (95, [DRAGON_TRIBE_SKILLS], [DRAGON_SKILLS, ANCIENT_SKILLS], ["skill_dragon_power", "skill_dragon_phoenix_armor"]),
    "boss_five_emperor_gate_guard": (95, [], [ANCIENT_SKILLS], ["skill_five_emperor_gate_art", "skill_emperor_flame"]),
    "boss_hun_tiandi": (99, [SOUL_HALL_SKILLS], [SOUL_SKILLS, ANCIENT_SKILLS], ["skill_hun_di", "skill_death_gate", "skill_ten_thousand_souls_devour"]),
}

# ══════════════════════════════════════════════════════════
# APPLY
# ══════════════════════════════════════════════════════════
updated = 0
for eid, (level, faction, element, uniques) in BOSS_SKILLS.items():
    skills = assign_boss_skills(eid, eid, level, faction, element, uniques)
    # Find and update
    for row in ws_enemies.iter_rows(min_row=2):
        if str(row[0].value) == eid:
            ws_enemies.cell(row=row[0].row, column=skills_col+1).value = ",".join(skills)
            updated += 1
            print(f"  {eid}: {len(skills)} skills")
            break
    else:
        print(f"  NOT FOUND: {eid}")

wb.save(WORKBOOK_PATH)
print(f"\nUpdated {updated} bosses. Total skills: {len(existing_skills)}")
print("Done!")
