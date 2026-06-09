"""
Fix level 99→100: 斗圣高阶 Lv95-99, 斗帝 Lv100.
Also update attribute level max from 99 to 100.
"""
import shutil
from pathlib import Path
import openpyxl

WORKBOOK_PATH = Path("story/text_game_event_schema_v4.xlsx")
BACKUP_PATH = Path("story/text_game_event_schema_v4_backup.xlsx")

shutil.copy2(WORKBOOK_PATH, BACKUP_PATH)
print(f"备份已创建：{BACKUP_PATH}")

wb = openpyxl.load_workbook(WORKBOOK_PATH)


def find_sheet(name_prefix: str):
    for name in wb.sheetnames:
        if name.startswith(name_prefix):
            return wb[name]
    return None


# 1. Update Realms_境界成长: 斗圣高阶 Lv95-99, 斗帝 Lv100
ws_realms = find_sheet("Realms")
if ws_realms:
    # Find the 斗圣高阶 row (should be row with min 95) and 斗帝 row (min 99)
    for row in range(2, ws_realms.max_row + 1):
        min_val = ws_realms.cell(row=row, column=3).value  # Min_Level column
        if min_val == 95:
            ws_realms.cell(row=row, column=4).value = 99  # Max_Level: 98→99
            print(f"斗圣高阶 max_level: 98→99")
        if min_val == 99:
            ws_realms.cell(row=row, column=3).value = 100  # Min_Level: 99→100
            ws_realms.cell(row=row, column=4).value = 100  # Max_Level: 99→100
            print(f"斗帝: 99→100")

# 2. Update Level_Progression: add Lv100 row
ws_lp = find_sheet("Level_Progression")
if ws_lp:
    # Find headers
    col_map = {}
    for col in range(1, ws_lp.max_column + 1):
        val = ws_lp.cell(row=1, column=col).value
        if val:
            col_map[str(val).strip()] = col

    # Update Lv99 row: change range to "100" and set Progress_Exp=0 (special event)
    for row in range(2, ws_lp.max_row + 1):
        lr = ws_lp.cell(row=row, column=col_map["Level_Range"]).value
        if lr and str(lr).strip() == "99":
            ws_lp.cell(row=row, column=col_map["Level_Range"]).value = "100"
            ws_lp.cell(row=row, column=col_map["Notes"]).value = "斗帝 — 特殊事件"
            print(f"Level_Progression: 99→100 (斗帝)")

# 3. Update Attributes_属性库: level max 99→100
ws_attr = find_sheet("Attributes")
if ws_attr:
    for row in range(2, ws_attr.max_row + 1):
        attr_id = ws_attr.cell(row=row, column=1).value
        if attr_id and str(attr_id).strip() == "level":
            ws_attr.cell(row=row, column=4).value = 100  # Max column
            print(f"Attributes level max: 99→100")

wb.save(WORKBOOK_PATH)
print("完成！")
