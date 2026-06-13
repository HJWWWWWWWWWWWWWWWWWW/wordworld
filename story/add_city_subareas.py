"""
Add commercial sub-areas for new cities.
Each city gets: market/commercial street + inn (for hub cities).
"""
from pathlib import Path
import openpyxl

WORKBOOK_PATH = Path("story/text_game_event_schema_v4.xlsx")

wb = openpyxl.load_workbook(WORKBOOK_PATH)


def find_sheet(name_prefix):
    for name in wb.sheetnames:
        if name.startswith(name_prefix):
            return wb[name]
    raise KeyError(f"Sheet starting with '{name_prefix}' not found")


def get_existing_ids(ws):
    ids = set()
    for row in ws.iter_rows(min_row=2):
        val = row[0].value
        if val:
            ids.add(str(val))
    return ids


def append_row(ws, data: dict):
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    row_data = []
    for h in headers:
        val = data.get(h, "")
        row_data.append(val if val is not None else "")
    ws.append(row_data)


ws_maps = find_sheet("Maps_")
ws_encounters = find_sheet("Map_Encounters_")

existing_maps = get_existing_ids(ws_maps)
existing_encounters = get_existing_ids(ws_encounters)

# ══════════════════════════════════════════════════════════
# NEW SUB-AREA MAPS
# Format: (map_id, name, region, unlock_cond, rec_lvl, stamina, safe_zone, exit, desc, notes)
# ══════════════════════════════════════════════════════════
NEW_SUB_MAPS = [
    # ── 漠城 ──
    ("map_mo_market", "漠城集市", "漠城", "", 6, 4, "yes", "",
     "漠城最热闹的交易区，沙漠商队卸下远方货物在此叫卖。偶有探险者出售从沙漠深处带出的古物残片。",
     "漠城商业"),
    ("map_mo_inn", "漠城客栈", "漠城", "", 6, 3, "yes", "",
     "往来沙漠的旅人落脚处，大厅里永远飘着沙枣茶的香气。佣兵们在此交换沙漠情报与雇主的消息。",
     "漠城休整"),

    # ── 石漠城 ──
    ("map_stone_mo_market", "石漠城坊市", "石漠城", "", 7, 4, "yes", "",
     "以砂岩垒砌的露天坊市，蛇人族偶尔携沙漠特产前来以物易物。城中佣兵大厅就设在坊市入口。",
     "石漠城商业"),
    ("map_stone_mo_mercenary", "石漠城佣兵大厅", "石漠城", "", 7, 4, "yes", "",
     "石漠城佣兵任务发布与结算的中心，木制柜台后排满了等候接取悬赏的灵者。青鳞在此被人刁难。",
     "青鳞剧情关键节点"),

    # ── 黑岩城 ──
    ("map_black_rock_market", "黑岩城商业街", "黑岩城", "", 4, 4, "yes", "",
     "黑岩城最繁华的街道，炼药师公会分会位于街尾。沿街药铺鳞次栉比，暗巷通往地下黑市。",
     "黑岩城商业"),
    ("map_black_rock_black_market", "黑岩城地下黑市", "黑岩城", "", 5, 5, "no", "",
     "藏在废弃铁匠铺下的隐秘交易点，禁售药材与来路不明的功法在此暗流涌动。帮派眼线混杂其中。",
     "黑岩城黑市"),

    # ── 盐城 ──
    ("map_salt_market", "盐城码头集市", "盐城", "", 5, 4, "yes", "",
     "紧邻码头的露天集市，成袋盐矿堆积如山。盐帮的账房先生在此收税，搬运工的号子声此起彼伏。",
     "盐城商业"),

    # ── 镇鬼关 ──
    ("map_ghost_pass_market", "镇鬼关边市", "镇鬼关", "", 6, 4, "yes", "",
     "关隘南侧自发形成的边贸市场，军士与私贩在此暗中交易边境缴获的魔兽材料与违禁品。",
     "镇鬼关商业"),
    ("map_ghost_pass_barracks", "镇鬼关军营", "镇鬼关", "", 6, 5, "no", "",
     "镇鬼关守军的驻扎营地，瞭望塔上的哨兵日夜监视北方魔兽动向。守将在此发布边境清剿任务。",
     "镇鬼关军事"),

    # ── 天涯城 ──
    ("map_tianya_wormhole_square", "天涯城虫洞广场", "天涯城", "", 26, 5, "yes", "",
     "天涯城核心——罗家世代守护的空间虫洞坐落于此。广场上永远挤满等待虫洞开启的各路强者，罗家护卫穿梭巡查。",
     "虫洞枢纽核心"),
    ("map_tianya_market", "天涯城街市", "天涯城", "", 26, 4, "yes", "",
     "虫洞带来的商机让这条街市异常繁荣。功法、丹药与各地特产琳琅满目，但价格比别处贵上三成。",
     "天涯城商业"),
    ("map_tianya_inn", "天涯城客栈", "天涯城", "", 26, 3, "yes", "",
     "虫洞维护期间被滞留的强者们聚集于此。客栈大堂是最好的情报交换地——只要请对酒。",
     "天涯城休整"),

    # ── 黑印城 ──
    ("map_black_seal_auction", "黑印城拍卖行", "黑印城", "", 20, 5, "yes", "",
     "八扇门掌控的地下拍卖场，暗角域最隐秘的交易盛会。珍品功法、禁药与高阶魔核在此轮番登场。",
     "暗角域拍卖核心"),
    ("map_black_seal_market", "黑印城地下集市", "黑印城", "", 20, 4, "yes", "",
     "拍卖会外围自发形成的集市，竞拍失败的买家在此互相交易。真假货混杂，眼力就是身家。",
     "黑印城商业"),

    # ── 魔焰谷 ──
    ("map_demon_valley_hall", "魔焰谷大殿", "魔焰谷", "", 36, 8, "no", "",
     "方言三老议事的大殿，谷中毒火从地缝中渗出。殿内悬着历代谷主的画像，阴森中透着一股邪气。",
     "魔焰谷核心"),
    ("map_demon_valley_archive", "魔焰谷藏功阁", "魔焰谷", "", 36, 6, "no", "",
     "存放魔焰谷百年积累的功法灵技与冷煜炼药笔记的石阁。禁制森严，擅闯者遭毒火反噬。",
     "魔焰谷功法"),

    # ── 万蝎门 ──
    ("map_scorpion_hall", "万蝎门大殿", "万蝎门", "", 34, 8, "no", "",
     "蝎毕岩踞坐于万蝎王座之上，殿中蝎毒雾气弥漫。墙壁上爬满拇指大的赤红毒蝎。",
     "万蝎门核心"),
    ("map_scorpion_cave", "万蝎门毒窟", "万蝎门", "", 34, 7, "no", "",
     "万蝎门培育毒蝎与提炼蝎毒的阴暗洞窟，洞壁上布满蝎尾划出的沟痕。剧毒材料俯拾皆是——只要不怕被蛰。",
     "万蝎门毒窟"),

    # ── 天冥宗 ──
    ("map_sky_demon_gate", "天冥宗山门", "天冥宗", "", 42, 5, "no", "",
     "幽暗雾气笼罩的天冥宗入口，守门弟子以冥气凝聚鬼面审视来者。未得允许擅入者，会被冥气吞噬。",
     "天冥宗门户"),
    ("map_sky_demon_hall", "天冥宗冥殿", "天冥宗", "", 42, 8, "no", "",
     "宗主与长老议事之所，冥气化为实质在殿中流转。殿中幽暗的光线下，隐约可见黑渊护法的身影。",
     "天冥宗核心"),

    # ── 天罡殿 ──
    ("map_heavenly_gang_prison", "天罡殿魂牢", "天罡殿", "", 62, 10, "no", "",
     "囚禁被收集灵魂体的阴冷牢狱，铁链拖曳声与灵魂哀嚎不绝于耳。玄炉老人曾被囚于此。",
     "黑渊殿囚牢"),
    ("map_heavenly_gang_origin", "天罡殿本源殿", "天罡殿", "", 63, 10, "no", "",
     "黑渊殿囤积灵魂本源的秘殿，殿中央的灵魂井深不见底。魂灭生在此抽取本源之力强化自身。",
     "黑渊殿本源"),
]

# ══════════════════════════════════════════════════════════
# ENCOUNTERS for sub-areas
# ══════════════════════════════════════════════════════════
def enc(evt_id, map_id, text, opt1_text="", opt1_effect="",
        opt2_text="", opt2_effect="", opt3_text="", opt3_effect="", weight=1):
    return {
        "Event_ID": evt_id, "Map_ID": map_id, "Weight": weight,
        "Text": text, "Pre_Condition": "",
        "Opt1_Text": opt1_text, "Opt1_Condition": "", "Opt1_Next": "", "Opt1_Effect": opt1_effect,
        "Opt2_Text": opt2_text, "Opt2_Condition": "", "Opt2_Next": "", "Opt2_Effect": opt2_effect,
        "Opt3_Text": opt3_text, "Opt3_Condition": "", "Opt3_Next": "", "Opt3_Effect": opt3_effect,
        "Notes": "",
    }

NEW_ENCOUNTERS = [
    # ── 漠城集市 ──
    enc("enc_mo_market_1", "map_mo_market",
        "集市角落的旧货摊上，一堆沙尘覆盖的杂物中露出一角泛黄的残破地图。残图入手微温，绝非寻常纸张。",
        "仔细翻找残图", "soul:+3,item:+item_elixir",
        "向摊主打听来历", "soul:+2"),
    enc("enc_mo_market_2", "map_mo_market",
        "一个沙漠商贩正在兜售据称从古殿废墟中挖掘出的古物。围观者中有人嗤之以鼻，有人目光灼灼。",
        "购买古物碰碰运气", "silver:-15,soul:+3",
        "鉴定古物真伪", "alchemy:+2",
        "当热闹看", ""),

    # ── 漠城客栈 ──
    enc("enc_mo_inn_1", "map_mo_inn",
        "客栈角落里，一个披斗篷的老者独自饮酒。杯中酒水在他指尖凝出薄冰——沙漠中绝不该出现的景象。",
        "上前攀谈", "rel:npc_hai_bodong:+5,soul:+3",
        "远远观察", "soul:+2"),
    enc("enc_mo_inn_2", "map_mo_inn",
        "几个佣兵在大堂高声谈论沙漠深处的奇遇：沙暴中浮现的古城、地穴中的火焰异象……",
        "请他们喝酒套话", "silver:-10,soul:+3",
        "旁听记录情报", "soul:+2"),

    # ── 石漠城坊市 ──
    enc("enc_stone_mo_market_1", "map_stone_mo_market",
        "蛇人族商贩摆出沙漠独有的药材与矿石，围观者不少但敢上前交易的寥寥无几。",
        "大方交易", "silver:-15,item:+herb_desert_mandala",
        "旁观学习", "alchemy:+1"),
    enc("enc_stone_mo_market_2", "map_stone_mo_market",
        "坊市告示板上贴满了寻人启事——大多是深入沙漠后失踪的探险者。一张新贴的悬赏格外引人注目。",
        "揭下悬赏", "exp:+40,silver:+20",
        "记下情报", "soul:+2"),

    # ── 石漠城佣兵大厅 ──
    enc("enc_stone_mo_merc_1", "map_stone_mo_mercenary",
        "大厅中，几个佣兵围着一个怯生生的少女嬉笑刁难。少女眼中碧绿光芒一闪即逝——那是蛇人族的特征。",
        "出手解围", "rel:npc_qing_lin:+10,reputation:+5",
        "冷眼旁观", "soul:+1"),
    enc("enc_stone_mo_merc_2", "map_stone_mo_mercenary",
        "佣兵大厅发布了新的沙漠护送任务，目的地是蛇人族领地边缘的石漠。报酬不菲但风险极高。",
        "接取护送任务", "exp:+50,silver:+20",
        "招募佣兵同行", "exp:+30,reputation:+3"),

    # ── 黑岩城商业街 ──
    enc("enc_black_rock_market_1", "map_black_rock_market",
        "炼药师公会分会门前排着长队。今日是月度考核日，执事正逐一核验报名者的资格。",
        "报名参加考核", "alchemy:+2,reputation:+3",
        "观摩他人考核", "alchemy:+1"),
    enc("enc_black_rock_market_2", "map_black_rock_market",
        "街角药铺的老板神秘兮兮地拉开抽屉，露出几株被禁售的黑骨花——疗伤圣品，但采集者多死于魔兽之口。",
        "高价购买", "silver:-20,item:+herb_blood_essence_fruit",
        "举报违禁品", "reputation:+3"),

    # ── 黑岩城地下黑市 ──
    enc("enc_black_rock_black_1", "map_black_rock_black_market",
        "黑市深处正在举行小型地下拍卖。一件来历不明的玄阶功法残卷被推上前台，竞价声此起彼伏。",
        "参与竞价", "silver:-30,exp:+40",
        "观察买家身份", "soul:+2"),
    enc("enc_black_rock_black_2", "map_black_rock_black_market",
        "暗巷中两个身影低声交谈，隐约听到「青岚宗」和「黑渊殿」的字眼。其中一人袖口绣着陌生的标志。",
        "暗中偷听", "soul:+5",
        "跟踪可疑人物", "exp:+30"),

    # ── 盐城码头集市 ──
    enc("enc_salt_market_1", "map_salt_market",
        "盐帮的账房先生正在集市收税，一个老实的盐贩因交不出份额被盐帮打手围住。",
        "替盐贩解围", "reputation:+5,silver:-5",
        "找账房理论", "reputation:+3"),
    enc("enc_salt_market_2", "map_salt_market",
        "码头上停着一艘满载货物的商船，船长正在招募前往魔兽山脉东麓的护卫。",
        "接受护卫任务", "exp:+40,silver:+15",
        "打听魔兽山脉情报", "soul:+2"),

    # ── 镇鬼关边市 ──
    enc("enc_ghost_pass_market_1", "map_ghost_pass_market",
        "边市角落里，军士正以极低价格出售一批从魔兽身上缴获的材料。品相完好但来路显然不符合军规。",
        "趁便宜收购", "silver:-10,item:+core_magic",
        "举报私售军资", "reputation:+3"),
    enc("enc_ghost_pass_market_2", "map_ghost_pass_market",
        "一个疲惫的旅人在边市打听北上的安全路线。他自称要穿越边境前往出云帝国寻找失散的亲人。",
        "分享情报", "reputation:+3,soul:+1",
        "劝阻他不要冒险", "reputation:+2"),

    # ── 镇鬼关军营 ──
    enc("enc_ghost_pass_barracks_1", "map_ghost_pass_barracks",
        "守关将领站在沙盘前眉头紧锁——北境魔兽的活动范围正在向南扩张，几个前沿哨站已经失联。",
        "主动请缨侦查", "exp:+50,reputation:+5",
        "协助制定防御计划", "soul:+3,reputation:+3"),

    # ── 天涯城虫洞广场 ──
    enc("enc_tianya_wormhole_1", "map_tianya_wormhole_square",
        "虫洞出口光芒大盛，一位刚从虫洞中踏出的灵尊强者面色苍白地扶墙喘息。罗家护卫立刻上前递上恢复丹药。",
        "向灵尊请教中州见闻", "soul:+3",
        "观察虫洞运转", "soul:+2"),
    enc("enc_tianya_wormhole_2", "map_tianya_wormhole_square",
        "罗家族长亲自来到广场，宣布虫洞因不稳定波动需要紧急维修。滞留的人群发出一片哀叹——维修至少需要三天。",
        "主动帮忙维修", "soul:+5,exp:+50",
        "耐心等待", "exp:+10"),

    # ── 天涯城街市 ──
    enc("enc_tianya_market_1", "map_tianya_market",
        "街市上的物价高得离谱——一株普通凝血草的价格是外界的五倍。但虫洞就在眼前，再贵也有人买账。",
        "咬牙采购补给", "silver:-20,item:+item_elixir",
        "只逛不买打探情报", "soul:+2"),
    enc("enc_tianya_market_2", "map_tianya_market",
        "一个商人正在收购从中州带回的消息。他说中州各大势力的动向在天涯城能卖好价钱。",
        "出售已知情报", "silver:+15,reputation:+2",
        "向商人购买中州消息", "silver:-10,soul:+3"),

    # ── 天涯城客栈 ──
    enc("enc_tianya_inn_1", "map_tianya_inn",
        "客栈大堂里聚满了因虫洞维修而滞留的旅人。一个自称曾在中州四方阁修炼过的中年灵者正在高谈阔论。",
        "请他喝酒打听详情", "silver:-5,soul:+3",
        "默默旁听", "soul:+1"),
    enc("enc_tianya_inn_2", "map_tianya_inn",
        "客栈老板悄悄告诉你，后院住着一位刚从虫洞那头逃回来的伤者——他声称中州那边出了大事。",
        "去后院探视", "exp:+30,soul:+3",
        "请老板喝一杯套话", "silver:-5,soul:+2"),

    # ── 黑印城拍卖行 ──
    enc("enc_black_seal_auction_1", "map_black_seal_auction",
        "八扇门的地下拍卖厅座无虚席。今晚的压轴拍品是一卷地阶灵技残卷，起拍价已让大半在场者倒吸一口凉气。",
        "参与竞拍", "silver:-50,exp:+60",
        "记录竞拍者身份", "soul:+3"),
    enc("enc_black_seal_auction_2", "map_black_seal_auction",
        "拍卖会休息间隙，八扇门的管事向贵宾包厢送去了几件未上拍品清单的'私货'。",
        "设法混入贵宾区", "soul:+5,exp:+30",
        "在外围打探私货消息", "soul:+2"),

    # ── 黑印城地下集市 ──
    enc("enc_black_seal_market_1", "map_black_seal_market",
        "竞拍失败的买家们在这个自发集市中互相交易——有人急于出手，有人想捡漏。真假混杂，全凭眼力。",
        "淘宝捡漏", "silver:-15,soul:+3",
        "观察交易寻找线索", "soul:+2"),

    # ── 魔焰谷大殿 ──
    enc("enc_demon_valley_hall_1", "map_demon_valley_hall",
        "大殿中方言三老正在商议对策。'冷煜那家伙……把烂摊子丢给我们，自己灵魂跑路了。'",
        "现身质问", "exp:+150,reputation:+10",
        "继续偷听", "soul:+5"),

    # ── 魔焰谷藏功阁 ──
    enc("enc_demon_valley_archive_1", "map_demon_valley_archive",
        "藏功阁石壁上刻满功法口诀，冷煜在此留下了海心焰的控火心得。禁制在源火面前形同虚设。",
        "抄录功法", "exp:+100,alchemy:+3",
        "专注研究控火心得", "alchemy:+5"),

    # ── 万蝎门大殿 ──
    enc("enc_scorpion_hall_1", "map_scorpion_hall",
        "蝎毕岩的万蝎王座之下，数百只毒蝎组成了一道活体屏障。他冷笑着把玩手中的蝎毒晶。",
        "挑战蝎毕岩", "exp:+180,reputation:+12",
        "以解毒丹交换情报", "alchemy:+3,soul:+2"),

    # ── 万蝎门毒窟 ──
    enc("enc_scorpion_cave_1", "map_scorpion_cave",
        "洞窟深处堆积着数以千计的蝎蜕，其中几具泛着诡异的紫光——那是变异蝎王的遗蜕，入药可解百毒。",
        "收集蝎王蜕", "item:+herb_blood_lotus_essence,alchemy:+2",
        "小心避开深入探索", "exp:+50"),

    # ── 天冥宗山门 ──
    enc("enc_sky_demon_gate_1", "map_sky_demon_gate",
        "守门弟子的冥气鬼面在空中凝聚成形，空洞的眼眶中冥火闪烁。'来者止步，报上名来。'",
        "递交拜帖", "reputation:+3",
        "展示实力闯入", "exp:+80"),

    # ── 天冥宗冥殿 ──
    enc("enc_sky_demon_hall_1", "map_sky_demon_hall",
        "殿中冥气流转，宗主与黑渊护法的密谈正在进行。你的闯入让两人同时转头，目光中杀意凛然。",
        "正面质问勾结黑渊殿之事", "exp:+150,reputation:+10",
        "暗中收集证据", "soul:+8"),

    # ── 天罡殿魂牢 ──
    enc("enc_heavenly_gang_prison_1", "map_heavenly_gang_prison",
        "魂牢中铁链拖曳声此起彼伏，数百个灵魂体被锁在特制的魂锁中。其中一个苍老的灵魂体让你心头一颤——那是玄炉老人被囚的位置。",
        "释放被囚灵魂", "soul:+12,reputation:+10",
        "寻找玄炉老人遗留的痕迹", "soul:+8"),

    # ── 天罡殿本源殿 ──
    enc("enc_heavenly_gang_origin_1", "map_heavenly_gang_origin",
        "本源殿中央的灵魂井深不见底，井口涌出的灵魂本源化为实质性的雾霭。魂灭生正在井边吸收这些力量。",
        "阻止魂灭生吸收本源", "exp:+250,soul:+15",
        "试图封印灵魂井", "soul:+10,exp:+150"),
]

# ══════════════════════════════════════════════════════════
# GATHERING SPOTS for new sub-areas
# ══════════════════════════════════════════════════════════
# First check if Map_Gathering sheet exists
gathering_exists = any(name.startswith("Map_Gathering") for name in wb.sheetnames)
if not gathering_exists:
    print("Creating Map_Gathering sheet...")
    ws_gathering = wb.create_sheet("Map_Gathering_采集点")
    ws_gathering.append([
        "Gather_ID", "Map_ID", "Item_ID", "Chance_Percent",
        "Min_Qty", "Max_Qty", "Pre_Condition", "Explore_Effect",
        "Unique_Limit", "Location_Exclusive", "Notes"
    ])
else:
    ws_gathering = find_sheet("Map_Gathering")

existing_gathering = get_existing_ids(ws_gathering) if ws_gathering else set()

def gather(gid, map_id, item_id, chance, min_q=1, max_q=3, pre="", effect="", limit="", exclusive="", notes=""):
    return {
        "Gather_ID": gid, "Map_ID": map_id, "Item_ID": item_id,
        "Chance_Percent": chance, "Min_Qty": min_q, "Max_Qty": max_q,
        "Pre_Condition": pre, "Explore_Effect": effect,
        "Unique_Limit": limit, "Location_Exclusive": exclusive, "Notes": notes,
    }

NEW_GATHERING = [
    # 漠城
    gather("gather_mo_market", "map_mo_market", "herb_coagulation", 20),
    gather("gather_mo_market_map", "map_mo_market", "item_elixir", 8),
    # 石漠城
    gather("gather_stone_mo_market", "map_stone_mo_market", "herb_desert_mandala", 22),
    gather("gather_stone_mo_merc", "map_stone_mo_mercenary", "core_magic", 15),
    # 黑岩城
    gather("gather_black_rock_market", "map_black_rock_market", "herb_spirit_gathering", 25),
    gather("gather_black_rock_black", "map_black_rock_black_market", "herb_green_flame_grass", 12),
    # 盐城
    gather("gather_salt_market", "map_salt_market", "core_earth", 18),
    # 镇鬼关
    gather("gather_ghost_pass_market", "map_ghost_pass_market", "core_magic", 20),
    # 天涯城
    gather("gather_tianya_market", "map_tianya_market", "item_elixir", 10),
    # 黑印城
    gather("gather_black_seal_market", "map_black_seal_market", "herb_blood_essence_fruit", 10),
    # 魔焰谷
    gather("gather_demon_valley_archive", "map_demon_valley_archive", "core_fire", 25),
    # 万蝎门
    gather("gather_scorpion_cave", "map_scorpion_cave", "herb_desert_mandala", 28),
]

# ══════════════════════════════════════════════════════════
# APPLY
# ══════════════════════════════════════════════════════════
print(f"Adding {len(NEW_SUB_MAPS)} new sub-area maps...")
for data in NEW_SUB_MAPS:
    map_id = data[0]
    if map_id in existing_maps:
        print(f"  SKIP (exists): {map_id}")
        continue
    row = {
        "Map_ID": data[0], "Name": data[1], "Region": data[2],
        "Unlock_Condition": data[3], "Recommend_Level": data[4],
        "Explore_Stamina_Cost": data[5], "Default_Encounter_Weight": 1,
        "Safe_Zone": data[6], "Exit_Event": data[7],
        "Description": data[8], "Notes": data[9],
    }
    append_row(ws_maps, row)
    existing_maps.add(map_id)
    print(f"  + {map_id}: {data[1]}")

print(f"\nAdding {len(NEW_ENCOUNTERS)} new encounters...")
for data in NEW_ENCOUNTERS:
    evt_id = data["Event_ID"]
    if evt_id in existing_encounters:
        print(f"  SKIP (exists): {evt_id}")
        continue
    append_row(ws_encounters, data)
    existing_encounters.add(evt_id)
    print(f"  + {evt_id}")

print(f"\nAdding {len(NEW_GATHERING)} gathering spots...")
for data in NEW_GATHERING:
    gid = data["Gather_ID"]
    if gid in existing_gathering:
        print(f"  SKIP (exists): {gid}")
        continue
    append_row(ws_gathering, data)
    existing_gathering.add(gid)
    print(f"  + {gid}")

wb.save(WORKBOOK_PATH)
print(f"\nSaved to {WORKBOOK_PATH}")
print("Done!")
