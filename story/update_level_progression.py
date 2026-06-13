"""
Update Level_Progression sheet with new Progress_Exp column.
Each level range gets a per-progress-point exp cost.
Backup is created before modification.
"""
import shutil
from pathlib import Path
import openpyxl

WORKBOOK_PATH = Path("story/text_game_event_schema_v4.xlsx")
BACKUP_PATH = Path("story/text_game_event_schema_v4_backup.xlsx")

# Backup
shutil.copy2(WORKBOOK_PATH, BACKUP_PATH)
print(f"备份已创建：{BACKUP_PATH}")

wb = openpyxl.load_workbook(WORKBOOK_PATH)


def find_sheet(name_prefix: str):
    for name in wb.sheetnames:
        if name.startswith(name_prefix):
            return wb[name]
    return None


ws = find_sheet("Level_Progression")
if ws is None:
    raise ValueError("找不到 Level_Progression 表")

# Read headers (row 1)
headers = {}
for col in range(1, ws.max_column + 1):
    val = ws.cell(row=1, column=col).value
    if val:
        headers[col] = str(val).strip()

# Find the last used column
last_col = max(headers.keys())
next_col = last_col + 1

# Add new header
ws.cell(row=1, column=next_col, value="Progress_Exp")
print(f"在列 {next_col} 添加 Progress_Exp 表头")

# Define progress exp per level range
# These match the new level design in engine.py
progress_exp_map = {
    "1-9": 25,
    "10-19": 60,
    "20-29": 150,
    "30-39": 300,
    "40-49": 600,
    "50-59": 1200,
    "60-69": 2500,
    "70-79": 5000,
    "80-89": 10000,
    "90-94": 20000,
    "95-98": 40000,
    "99": 0,  # 灵帝 — 特殊事件
}

# Update existing rows and add Progress_Exp values
# First, update Level_Range to match the new finer-grained design
new_ranges = [
    ("1-9", "25", "斗之气"),
    ("10-19", "60", "灵者"),
    ("20-29", "150", "斗师"),
    ("30-39", "300", "大斗师"),
    ("40-49", "600", "灵师"),
    ("50-59", "1200", "灵王"),
    ("60-69", "2500", "灵皇"),
    ("70-79", "5000", "灵宗"),
    ("80-89", "10000", "灵尊"),
    ("90-94", "20000", "灵圣"),
    ("95-98", "40000", "灵圣高阶"),
    ("99", "0", "灵帝 — 特殊事件"),
]

# Clear existing data rows (rows 2 onwards) and rewrite
# First, delete all rows from 2 to max_row
if ws.max_row > 1:
    ws.delete_rows(2, ws.max_row - 1)

# Write new data
# We need columns: Level_Range, Exp_Formula (keep for reference), Douqi_Gain, HP_Gain, ATK_Gain, DEF_Gain, SPD_Gain, Notes, Progress_Exp
# Let's find the column indices
col_map = {}
for col, name in headers.items():
    col_map[name] = col
col_map["Progress_Exp"] = next_col

# Stat gains per level range (matching original design, scaled)
stat_gains = {
    "1-9":     ("level*25", 5, 10, 2, 1, 1),
    "10-19":   ("level*60", 8, 15, 3, 2, 1),
    "20-29":   ("level*150", 12, 25, 5, 3, 2),
    "30-39":   ("level*300", 18, 40, 8, 5, 3),
    "40-49":   ("level*600", 25, 60, 12, 8, 5),
    "50-59":   ("level*1200", 35, 90, 18, 12, 7),
    "60-69":   ("level*2500", 50, 130, 25, 18, 10),
    "70-79":   ("level*5000", 70, 180, 35, 25, 14),
    "80-89":   ("level*10000", 100, 250, 50, 35, 20),
    "90-94":   ("level*20000", 140, 350, 70, 50, 28),
    "95-98":   ("level*40000", 200, 500, 100, 70, 40),
    "99":      ("special_event", 0, 0, 0, 0, 0),
}

for row_idx, (level_range, progress_exp, notes) in enumerate(new_ranges, start=2):
    exp_formula, douqi, hp, atk, def_, spd = stat_gains[level_range]
    ws.cell(row=row_idx, column=col_map["Level_Range"], value=level_range)
    ws.cell(row=row_idx, column=col_map["Exp_Formula"], value=exp_formula)
    ws.cell(row=row_idx, column=col_map["Douqi_Gain"], value=douqi)
    ws.cell(row=row_idx, column=col_map["HP_Gain"], value=hp)
    ws.cell(row=row_idx, column=col_map["ATK_Gain"], value=atk)
    ws.cell(row=row_idx, column=col_map["DEF_Gain"], value=def_)
    ws.cell(row=row_idx, column=col_map["SPD_Gain"], value=spd)
    ws.cell(row=row_idx, column=col_map["Notes"], value=notes)
    ws.cell(row=row_idx, column=col_map["Progress_Exp"], value=progress_exp)

wb.save(WORKBOOK_PATH)
print(f"已更新 Level_Progression 表：{len(new_ranges)} 行数据")
print("完成！")
