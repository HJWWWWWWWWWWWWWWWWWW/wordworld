"""
Enemy skill system design:
1. Add ~120 monster/NPC skills to Skills sheet
2. Assign skills to all 363 enemies based on tier/type/level
3. Boss-unique skills based on original novel
"""
from pathlib import Path
import openpyxl

WORKBOOK_PATH = Path("story/text_game_event_schema_v4.xlsx")
wb = openpyxl.load_workbook(WORKBOOK_PATH)

def find_sheet(name_prefix):
    for name in wb.sheetnames:
        if name.startswith(name_prefix):
            return wb[name]
    raise KeyError

ws_skills = find_sheet("Skills_")
ws_enemies = find_sheet("Enemies_")

# Get headers and existing IDs
skill_headers = [cell.value for cell in next(ws_skills.iter_rows(min_row=1, max_row=1))]
enemy_headers = [cell.value for cell in next(ws_enemies.iter_rows(min_row=1, max_row=1))]

existing_skills = set()
for row in ws_skills.iter_rows(min_row=2):
    if row[0].value: existing_skills.add(str(row[0].value))

def append_row(ws, data, headers):
    ws.append([str(data.get(h, "")) for h in headers])

# ══════════════════════════════════════════════════════════
# PART 1: MONSTER SKILLS (~120 new skills)
# ══════════════════════════════════════════════════════════
NEW_SKILLS = [
    # ── 物理通用 (Lv 1-20 mobs) ──
    ("ms_bite", "撕咬", "物理", "黄阶低级", "atk:+8", "野兽以利齿撕咬目标。"),
    ("ms_claw", "爪击", "物理", "黄阶低级", "atk:+10", "以利爪横扫目标。"),
    ("ms_tackle", "冲撞", "物理", "黄阶低级", "atk:+8,spd:+5", "以身体猛撞目标。"),
    ("ms_tail_swipe", "尾击", "物理", "黄阶中级", "atk:+12", "以尾巴猛烈抽击。"),
    ("ms_stomp", "践踏", "物理", "黄阶中级", "atk:+15,spd:-5", "沉重践踏造成范围伤害。"),
    ("ms_pierce", "穿刺", "物理", "黄阶中级", "atk:+18", "以角/刺贯穿目标防御。"),
    ("ms_constrict", "绞杀", "物理", "黄阶高级", "atk:+20,def:-5", "以身体缠绞目标。"),
    ("ms_crush", "碾压", "物理", "黄阶高级", "atk:+25,spd:-10", "以巨大身躯碾压目标。"),
    ("ms_frenzy", "狂暴", "物理", "玄阶低级", "atk:+30,def:-10", "陷入狂暴状态提升攻击。"),
    ("ms_rend", "撕裂", "物理", "玄阶低级", "atk:+28", "利爪撕裂目标造成持续伤害。"),

    # ── 火系 (Lv 3-40) ──
    ("ms_fire_ball", "火球", "火系", "黄阶低级", "atk:+10", "凝聚火焰成球攻击。"),
    ("ms_fire_breath", "火焰吐息", "火系", "黄阶中级", "atk:+16", "喷吐火焰灼烧前方。"),
    ("ms_flame_wave", "烈焰波", "火系", "黄阶高级", "atk:+22", "掀起火焰波浪席卷战场。"),
    ("ms_inferno", "烈焰风暴", "火系", "玄阶低级", "atk:+30,spd:-5", "召唤火焰风暴。"),
    ("ms_flame_armor", "火焰护体", "火系", "玄阶中级", "def:+15,atk:+10", "火焰缠绕自身形成护盾。"),
    ("ms_magma_burst", "岩浆喷发", "火系", "玄阶高级", "atk:+40,def:-10", "从地底召唤岩浆喷发。"),
    ("ms_meteor", "陨石降世", "火系", "地阶低级", "atk:+55", "召唤陨石从天而降。"),

    # ── 冰系 (Lv 3-40) ──
    ("ms_ice_shard", "冰锥", "冰系", "黄阶低级", "atk:+10", "凝聚冰锥射击目标。"),
    ("ms_frost_breath", "寒霜吐息", "冰系", "黄阶中级", "atk:+16,spd:-5", "喷吐寒气冻结目标。"),
    ("ms_ice_barrier", "冰墙", "冰系", "黄阶高级", "def:+20", "凝聚冰墙格挡攻击。"),
    ("ms_blizzard", "暴风雪", "冰系", "玄阶低级", "atk:+28,spd:-10", "召唤暴风雪覆盖战场。"),
    ("ms_frozen_nova", "冰霜新星", "冰系", "玄阶中级", "atk:+35,spd:-15", "释放环形冰霜冲击波。"),
    ("ms_absolute_zero", "绝对零度", "冰系", "地阶低级", "atk:+50,spd:-20", "将周围化为极寒炼狱。"),

    # ── 风系 (Lv 3-40) ──
    ("ms_wind_blade", "风刃", "风系", "黄阶低级", "atk:+10,spd:+5", "凝聚风刃远程攻击。"),
    ("ms_gale_strike", "疾风突袭", "风系", "黄阶中级", "atk:+15,spd:+10", "借助风力加速突袭。"),
    ("ms_cyclone", "旋风斩", "风系", "黄阶高级", "atk:+22,spd:+5", "旋转攻击周围目标。"),
    ("ms_thunder_gust", "雷风突进", "风系", "玄阶低级", "atk:+30,spd:+15", "风雷交加的迅猛冲击。"),
    ("ms_tempest", "风暴之眼", "风系", "玄阶高级", "atk:+42", "制造风暴漩涡。"),
    ("ms_heavenly_wind", "天罡风煞", "风系", "地阶低级", "atk:+52,spd:+10", "引天罡之风化为煞气攻击。"),

    # ── 土系 (Lv 3-40) ──
    ("ms_earth_spike", "地刺", "土系", "黄阶低级", "atk:+10", "从地面突起岩刺攻击。"),
    ("ms_rock_throw", "投石", "土系", "黄阶中级", "atk:+14", "投掷巨石砸向目标。"),
    ("ms_stone_skin", "石肤术", "土系", "黄阶高级", "def:+25", "石化皮肤大幅提升防御。"),
    ("ms_earthquake", "地震术", "土系", "玄阶低级", "atk:+28,spd:-10", "震动大地造成范围伤害。"),
    ("ms_sandstorm", "沙暴", "土系", "玄阶中级", "atk:+32,spd:-10", "制造沙尘暴遮蔽视线。"),
    ("ms_mountain_smash", "山崩地裂", "土系", "地阶低级", "atk:+50,spd:-20", "引山崩之力碾压。"),

    # ── 水系 (Lv 5-40) ──
    ("ms_water_jet", "水箭", "水系", "黄阶低级", "atk:+10", "高压水箭射击目标。"),
    ("ms_water_whip", "水鞭", "水系", "黄阶中级", "atk:+15,spd:+5", "水流化为鞭索。"),
    ("ms_whirlpool", "漩涡", "水系", "黄阶高级", "atk:+20,spd:-10", "制造水流漩涡困住目标。"),
    ("ms_tidal_wave", "巨浪滔天", "水系", "玄阶中级", "atk:+35", "掀起巨浪冲击。"),

    # ── 毒系 (Lv 7-50) ──
    ("ms_poison_spit", "毒液喷射", "毒系", "黄阶低级", "atk:+8,poison:+5", "喷射毒液使目标中毒。"),
    ("ms_venom_fang", "毒牙", "毒系", "黄阶中级", "atk:+14,poison:+8", "注入毒液造成持续伤害。"),
    ("ms_toxic_mist", "毒雾弥漫", "毒系", "黄阶高级", "atk:+18,poison:+10", "释放毒雾覆盖区域。"),
    ("ms_paralyze_venom", "麻痹毒液", "毒系", "玄阶低级", "atk:+22,spd:-15", "麻痹毒液降低目标速度。"),
    ("ms_corrosive_acid", "腐蚀酸液", "毒系", "玄阶中级", "atk:+30,def:-10", "喷射强酸腐蚀防御。"),
    ("ms_death_miasma", "死亡瘴气", "毒系", "地阶低级", "atk:+45,poison:+20", "释放致命瘴气。"),

    # ── 雷系 (Lv 8-50) ──
    ("ms_lightning_bolt", "雷电术", "雷系", "黄阶高级", "atk:+22,spd:+10", "召唤雷电劈击。"),
    ("ms_thunder_clap", "雷霆一击", "雷系", "玄阶低级", "atk:+32,spd:+5", "雷属性强力一击。"),
    ("ms_lightning_chain", "雷电锁链", "雷系", "玄阶中级", "atk:+38", "雷电在多个目标间跳跃。"),
    ("ms_thunder_wrath", "天雷怒降", "雷系", "地阶低级", "atk:+55,spd:+10", "召唤天雷之怒。"),

    # ── 暗系/灵魂系 (Lv 15-70) ──
    ("ms_shadow_strike", "暗影突袭", "暗系", "玄阶低级", "atk:+25,spd:+15", "从暗影中突袭目标。"),
    ("ms_soul_drain", "灵魂吸取", "灵魂", "玄阶低级", "atk:+20,soul:+5", "吸取目标灵魂力量。"),
    ("ms_life_drain", "生命汲取", "暗系", "玄阶中级", "atk:+25,hp:+20", "汲取目标生命恢复自身。"),
    ("ms_soul_shock", "灵魂冲击", "灵魂", "玄阶中级", "atk:+30,soul:+10", "灵魂冲击造成精神伤害。"),
    ("ms_dark_bind", "暗之束缚", "暗系", "玄阶高级", "atk:+35,spd:-10", "黑暗能量束缚目标。"),
    ("ms_soul_rend", "灵魂撕裂", "灵魂", "地阶低级", "atk:+48,soul:+15", "撕裂目标灵魂。"),
    ("ms_void_consumption", "虚空吞噬", "暗系", "地阶中级", "atk:+65", "虚空之力吞噬周围。"),

    # ── 龙系 (Lv 50-95) ──
    ("ms_dragon_roar", "龙吟", "龙系", "地阶低级", "atk:+40,soul:+10", "龙吟震慑目标灵魂。"),
    ("ms_dragon_breath", "龙息", "龙系", "地阶中级", "atk:+55", "喷吐毁灭性的龙息。"),
    ("ms_dragon_tail_sweep", "神龙摆尾", "龙系", "地阶中级", "atk:+50,spd:+10", "龙尾横扫千军。"),
    ("ms_dragon_might", "龙威", "龙系", "地阶高级", "atk:+60,spd:-15", "龙威压制使目标战力下降。"),
    ("ms_ancient_dragon_roar", "远虚空龙吟", "龙系", "天阶低级", "atk:+80,soul:+30", "远虚空龙吟震荡灵魂。"),

    # ── 远古系 (Lv 55-95) ──
    ("ms_ancient_seal", "远古封印", "远古", "地阶高级", "atk:+55,spd:-20", "远古封印禁锢目标。"),
    ("ms_bloodline_suppress", "血脉压制", "远古", "地阶高级", "atk:+50,spd:-10", "以远古血脉压制对手。"),
    ("ms_emperor_will", "帝之意志", "远古", "天阶低级", "atk:+70,soul:+25", "帝之意志不可违抗。"),

    # ── 源火系技能 ──
    ("ms_green_lotus_burn", "青莲焚天", "源火", "地阶高级", "atk:+70", "青莲源火焚尽万物。"),
    ("ms_fallen_heart_pulse", "陨心源火脉冲", "源火", "地阶高级", "atk:+65,soul:+15", "陨心源火脉冲灼烧灵魂。"),
    ("ms_three_thousand_starfire", "三千焱炎", "源火", "天阶低级", "atk:+80", "三千星空火星火燎原。"),
    ("ms_bone_chilling_ice", "玄冥冷火", "源火", "天阶低级", "atk:+75,soul:+20", "极寒源火冻结灵魂。"),
    ("ms_purifying_lotus", "净莲净化", "源火", "天阶中级", "atk:+90,soul:+30", "净世白莲火净化一切。"),
    ("ms_nihility_devour", "虚无吞噬", "源火", "天阶高级", "atk:+120,soul:+50", "虚无吞炎吞噬万物。"),

    # ── 人类/NPC通用技能 ──
    ("ms_sword_slash", "剑气斩", "剑技", "黄阶中级", "atk:+14", "凝聚剑气远程斩击。"),
    ("ms_heavy_strike", "重击", "物理", "黄阶中级", "atk:+16,spd:-5", "蓄力重击造成额外伤害。"),
    ("ms_quick_steps", "疾步", "身法", "黄阶中级", "spd:+20", "加快移动速度。"),
    ("ms_defensive_stance", "防御姿态", "防御", "黄阶高级", "def:+25,atk:-10", "进入防御姿态提升生存。"),
    ("ms_counter_strike", "反击", "物理", "玄阶低级", "atk:+25", "格挡后瞬间反击。"),
    ("ms_battle_cry", "战吼", "物理", "玄阶低级", "atk:+18,def:+10", "战吼鼓舞自身战力。"),
    ("ms_meditate", "冥想", "辅助", "黄阶高级", "soul:+10,hp:+20", "冥想恢复灵魂与生命。"),
    ("ms_potion_master", "丹药增幅", "辅助", "玄阶低级", "atk:+15,hp:+30", "服用丹药临时增幅战力。"),

    # ── 宗门专属技能 ──
    ("ms_cloud_wind_sword", "云岚剑诀", "剑技", "玄阶高级", "atk:+35,spd:+10", "青岚宗剑法精髓。"),
    ("ms_blood_refine", "血炼术", "秘法", "玄阶高级", "atk:+40,hp:-20", "血宗以血换力的秘术。"),
    ("ms_soul_lock", "锁魂链", "灵魂", "地阶低级", "atk:+45,soul:+15", "黑渊殿锁魂链攻击。"),
    ("ms_thunder_pavilion_art", "风雷诀", "风系", "地阶低级", "atk:+45,spd:+20", "风雷阁镇阁功法。"),
    ("ms_huangquan_art", "黄泉诀", "水/暗", "地阶低级", "atk:+45,soul:+12", "黄泉阁镇阁功法。"),
    ("ms_wanjian_art", "万剑归宗", "剑技", "地阶中级", "atk:+55", "万剑阁绝学。"),
    ("ms_fire_valley_art", "焚炎天功", "火系", "地阶中级", "atk:+55,soul:+10", "焚炎谷镇谷功法。"),
    ("ms_ice_valley_art", "冰河天功", "冰系", "地阶中级", "atk:+52,def:+15", "冰河谷镇谷功法。"),
]

print(f"Adding {len(NEW_SKILLS)} new monster skills...")
added = 0
for data in NEW_SKILLS:
    sid = data[0]
    if sid in existing_skills:
        continue
    row = {"Skill_ID": sid, "Name": data[1], "Type": data[2], "Rank": data[3],
           "Effect": data[4], "Description": data[5]}
    append_row(ws_skills, row, skill_headers)
    existing_skills.add(sid)
    added += 1
print(f"  + {added} skills added. Total skills: {len(existing_skills)}")

# ══════════════════════════════════════════════════════════
# PART 2: ASSIGN SKILLS TO ENEMIES
# ══════════════════════════════════════════════════════════
# Skill assignment rules by type/level/theme
def get_enemy_skills(eid, name, etype, level):
    skills = []
    lv = int(level)

    # ── PHYSICAL MOBS ──
    if 'wolf' in eid or '狼' in name:
        skills = ['ms_bite','ms_claw'] if lv<15 else ['ms_frenzy','ms_bite','ms_claw']
    elif 'snake' in eid or 'serpent' in eid or 'viper' in eid or 'python' in eid or '蟒' in name or '蛇' in name:
        skills = ['ms_bite','ms_venom_fang'] if lv<20 else ['ms_constrict','ms_venom_fang','ms_poison_spit']
    elif 'bear' in eid or '熊' in name:
        skills = ['ms_tackle','ms_stomp'] if lv<15 else ['ms_crush','ms_stomp','ms_battle_cry']
    elif 'tiger' in eid or '虎' in name or 'lion' in eid or '狮' in name:
        skills = ['ms_claw','ms_rend'] if lv<20 else ['ms_frenzy','ms_rend','ms_dragon_roar']
    elif 'eagle' in eid or 'hawk' in eid or 'bird' in eid or 'vulture' in eid or '鹰' in name or '鸟' in name:
        skills = ['ms_wind_blade','ms_gale_strike'] if lv<20 else ['ms_cyclone','ms_gale_strike']
    elif 'scorpion' in eid or '蝎' in name:
        skills = ['ms_pierce','ms_venom_fang','ms_paralyze_venom']
    elif 'spider' in eid or '蛛' in name:
        skills = ['ms_poison_spit','ms_dark_bind']
    elif 'boar' in eid or 'rhino' in eid or 'elephant' in eid or '猪' in name or '象' in name or '犀' in name:
        skills = ['ms_tackle','ms_stomp'] if lv<15 else ['ms_crush','ms_tackle']
    elif 'ape' in eid or '猿' in name:
        skills = ['ms_claw','ms_battle_cry'] if lv<15 else ['ms_frenzy','ms_crush']
    elif 'crocodile' in eid or '鳄' in name or 'lizard' in eid or '蜥' in name:
        skills = ['ms_bite','ms_tail_swipe'] if lv<15 else ['ms_constrict','ms_bite']
    elif 'bat' in eid or '蝙蝠' in name:
        skills = ['ms_shadow_strike','ms_life_drain']
    elif 'mantis' in eid or '螳' in name:
        skills = ['ms_claw','ms_quick_steps']

    # ── ELEMENTAL MOBS ──
    elif 'fire' in eid or 'flame' in eid or 'magma' in eid or 'scorching' in eid or '火' in name or '炎' in name or '焰' in name:
        skills = ['ms_fire_ball','ms_fire_breath'] if lv<20 else (['ms_flame_wave','ms_flame_armor'] if lv<40 else ['ms_inferno','ms_flame_armor'])
    elif 'ice' in eid or 'frost' in eid or 'frozen' in eid or '冰' in name or '霜' in name or '寒' in name:
        skills = ['ms_ice_shard'] if lv<15 else (['ms_frost_breath','ms_ice_barrier'] if lv<30 else ['ms_blizzard','ms_ice_barrier'])
    elif 'wind' in eid or 'gale' in eid or '风' in name:
        skills = ['ms_wind_blade'] if lv<15 else (['ms_gale_strike'] if lv<25 else ['ms_cyclone','ms_gale_strike'])
    elif 'earth' in eid or 'rock' in eid or 'stone' in eid or 'sand' in eid or '土' in name or '岩' in name or '沙' in name:
        skills = ['ms_earth_spike','ms_rock_throw'] if lv<20 else (['ms_stone_skin','ms_earthquake'] if lv<35 else ['ms_sandstorm','ms_stone_skin'])
    elif 'water' in eid or 'water' in eid or '水' in name or 'whirlpool' in eid:
        skills = ['ms_water_jet'] if lv<15 else ['ms_water_whip','ms_whirlpool']
    elif 'thunder' in eid or 'lightning' in eid or '雷' in name or '电' in name:
        skills = ['ms_lightning_bolt'] if lv<25 else (['ms_thunder_clap'] if lv<40 else ['ms_lightning_chain'])

    # ── POISON/DARK ──
    elif 'poison' in eid or 'toxic' in eid or 'venom' in eid or 'plague' in eid or '毒' in name:
        skills = ['ms_poison_spit','ms_toxic_mist'] if lv<25 else (['ms_paralyze_venom','ms_toxic_mist'] if lv<40 else ['ms_corrosive_acid','ms_death_miasma'])
    elif 'shadow' in eid or 'dark' in eid or '暗' in name or '影' in name:
        skills = ['ms_shadow_strike'] if lv<25 else ['ms_dark_bind','ms_shadow_strike']

    # ── SOUL/UNDEAD ──
    elif 'soul' in eid or 'ghost' in eid or 'wraith' in eid or 'skeleton' in eid or 'undead' in eid or 'zombie' in eid or '魂' in name or '鬼' in name or '尸' in name:
        skills = ['ms_soul_shock'] if lv<30 else (['ms_soul_drain','ms_life_drain'] if lv<50 else ['ms_soul_rend','ms_life_drain'])

    # ── DRAGON ──
    elif 'dragon' in eid or '龙' in name:
        skills = ['ms_dragon_roar','ms_dragon_tail_sweep'] if lv<60 else (['ms_dragon_breath','ms_dragon_might'] if lv<80 else ['ms_ancient_dragon_roar','ms_dragon_might'])

    # ── HUMANOID ENEMIES ──
    elif 'bandit' in eid or 'thug' in eid or 'thief' in eid or 'killer' in eid or 'raider' in eid or '盗' in name or '匪' in name or '徒' in name:
        skills = ['ms_heavy_strike'] if lv<20 else (['ms_shadow_strike','ms_quick_steps'] if lv<30 else ['ms_counter_strike','ms_sword_slash'])
    elif 'disciple' in eid or '弟子' in name:
        skills = ['ms_sword_slash','ms_defensive_stance'] if lv<25 else (['ms_heavy_strike','ms_meditate'] if lv<40 else ['ms_counter_strike','ms_sword_slash'])
    elif 'guard' in eid or 'guardian' in eid or 'patrol' in eid or 'sentry' in eid or 'enforcer' in eid or '守卫' in name or '护卫' in name:
        skills = ['ms_defensive_stance','ms_counter_strike'] if lv<30 else ['ms_battle_cry','ms_defensive_stance']
    elif 'elder' in eid or 'deacon' in eid or '长老' in name or '执事' in name:
        skills = ['ms_sword_slash','ms_meditate'] if lv<40 else ['ms_counter_strike','ms_battle_cry','ms_meditate']
    elif 'captain' in eid or 'general' in eid or 'commander' in eid or 'leader' in eid or 'chief' in eid or '队长' in name or '将军' in name or '统领' in name or '首领' in name:
        skills = ['ms_battle_cry','ms_heavy_strike'] if lv<35 else ['ms_counter_strike','ms_battle_cry','ms_sword_slash']

    # ── CULTIVATOR/HUMANOID GENERIC ──
    elif 'cultivator' in eid or 'fighter' in eid or 'warrior' in eid or 'merc' in eid or 'hunter' in eid or 'soldier' in eid or 'spy' in eid or 'master' in eid or 'examiner' in eid or 'champion' in eid or 'officer' in eid or 'instructor' in eid or 'referee' in eid:
        skills = ['ms_sword_slash','ms_heavy_strike'] if lv<30 else (['ms_counter_strike','ms_sword_slash','ms_meditate'] if lv<50 else ['ms_battle_cry','ms_counter_strike','ms_sword_slash'])
    elif 'slave' in eid or '奴隶' in name:
        skills = ['ms_heavy_strike','ms_battle_cry']

    # ── ELEMENTAL SPIRITS ──
    elif 'elemental' in eid or 'spirit' in eid or 'wisp' in eid or 'golem' in eid or 'puppet' in eid or '傀儡' in name or '像' in name:
        skills = ['ms_tackle'] if lv<30 else (['ms_stone_skin','ms_earthquake'] if lv<50 else ['ms_ancient_seal','ms_emperor_will'])

    # ── DEFAULT (by level) ──
    else:
        if lv <= 10:
            skills = ['ms_tackle']
        elif lv <= 20:
            skills = ['ms_tackle','ms_heavy_strike']
        elif lv <= 30:
            skills = ['ms_heavy_strike','ms_defensive_stance']
        elif lv <= 40:
            skills = ['ms_sword_slash','ms_battle_cry']
        elif lv <= 50:
            skills = ['ms_counter_strike','ms_sword_slash','ms_meditate']
        elif lv <= 65:
            skills = ['ms_battle_cry','ms_counter_strike','ms_sword_slash']
        else:
            skills = ['ms_emperor_will','ms_battle_cry','ms_counter_strike']

    return skills

# ══════════════════════════════════════════════════════════
# PART 3: BOSS UNIQUE SKILLS (from original novel)
# ══════════════════════════════════════════════════════════
BOSS_UNIQUE = {
    # ── 沧澜帝国 ──
    "enemy_nalan": ["skill_wind_sword", "skill_wind_push"],  # 苏婉清 - 风之极
    "enemy_medusa": ["skill_snake_seal", "skill_snake_eye", "ms_bloodline_suppress"],  # 赤鳞
    "enemy_hai": ["skill_ice", "ms_absolute_zero"],  # 海波东 - 冰皇
    "enemy_yun_leng": ["skill_wind_bind", "ms_cloud_wind_sword"],  # 青棱
    "enemy_yun_shan": ["skill_wind_sword", "skill_wind_bind", "ms_cloud_wind_sword"],  # 青山

    # ── 暗角域 ──
    "enemy_fan_lao": ["skill_blood_art", "ms_blood_refine", "ms_life_drain"],  # 范痨
    "enemy_fan_ling": ["skill_blood_art", "ms_blood_refine"],
    "enemy_han_feng": ["skill_sea_flame", "skill_flame_splitting_ruler", "ms_potion_master"],  # 冷煜 - 海心焰
    "enemy_mo_tianxing": ["ms_dark_bind", "ms_soul_shock", "skill_soul_chain"],

    # ── 出云帝国 ──
    "enemy_xie_biyan": ["ms_paralyze_venom", "ms_corrosive_acid", "ms_death_miasma"],  # 蝎毕岩

    # ── 黑渊殿 ──
    "enemy_wu_hufa": ["skill_soul_handprint", "skill_soul_chain", "ms_soul_drain"],  # 鹜护法
    "enemy_zhai_xing": ["skill_soul_chain", "ms_soul_rend", "ms_dark_bind"],  # 摘星老鬼
    "enemy_hun_yu": ["skill_soul_art", "ms_soul_rend", "ms_life_drain"],
    "enemy_hun_feng": ["skill_soul_art", "ms_soul_rend", "ms_void_consumption"],
    "enemy_hun_shengtian": ["skill_soul_art", "ms_soul_rend", "ms_ancient_seal", "ms_emperor_will"],
    "enemy_hun_yuantian": ["skill_soul_art", "ms_soul_rend", "ms_ancient_seal"],
    "enemy_hun_miesheng": ["skill_soul_art", "ms_soul_rend", "ms_void_consumption", "skill_soul_chain"],
    "enemy_hun_yao": ["skill_soul_handprint", "ms_soul_drain", "ms_soul_shock"],
    "enemy_hunmo": ["skill_soul_art", "ms_demon_valley_art", "ms_corrosive_acid"],
    "enemy_hun_xuzi": ["skill_soul_art", "ms_soul_drain", "ms_soul_rend"],
    "boss_hun_tiandi": ["skill_hun_di", "ms_nihility_devour", "ms_emperor_will", "ms_soul_rend"],  # 玄冥帝
    "enemy_xuwu": ["skill_nihility_flame", "ms_void_consumption", "ms_nihility_devour", "ms_emperor_will"],  # 虚无吞炎

    # ── 迦南学院 ──
    "boss_skyfire_beast_king": ["skill_fallen_heart_flame", "ms_inferno", "skill_five_wheel_fire"],

    # ── 中州 ──
    "enemy_mugu": ["skill_bone_chilling_flame", "ms_bone_chilling_ice", "ms_potion_master"],  # 慕骨老人 - 玄冥冷火
    "enemy_fang_yan": ["ms_fire_valley_art", "ms_inferno", "ms_blood_refine"],  # 方言-魔焰谷
    "boss_thunder_emperor": ["skill_wind_thunder", "ms_thunder_wrath", "ms_heavenly_wind"],  # 雷尊者
    "boss_sword_emperor": ["skill_sword", "ms_wanjian_art", "ms_sword_slash"],  # 剑尊者
    "boss_demon_flame_saint_spirit": ["ms_purifying_lotus", "ms_emperor_will"],  # 净莲妖圣

    # ── 龙岛 ──
    "boss_south_dragon_king": ["skill_dragon_power", "ms_ancient_dragon_roar", "ms_dragon_might"],
    "boss_north_dragon_king": ["skill_dragon_power", "ms_ancient_dragon_roar", "ms_dragon_might", "ms_void_consumption"],
    "boss_dragon_emperor_zhu_kun": ["skill_dragon_power", "skill_dragon_phoenix_armor", "ms_ancient_dragon_roar", "ms_emperor_will"],
    "boss_west_dragon_king": ["skill_dragon_power", "ms_ancient_dragon_roar", "ms_dragon_might"],

    # ── 远古种族 ──
    "boss_gu_clan_three_immortals": ["skill_ancient_emperor_seal", "skill_open_mountain_seal", "ms_bloodline_suppress", "ms_emperor_will"],
    "boss_yan_clan_flame_emperor": ["skill_gold_flame", "ms_green_lotus_burn", "ms_emperor_will"],
    "boss_lei_clan_thunder_emperor": ["skill_thunder", "ms_thunder_wrath", "ms_emperor_will"],
    "boss_hun_clan_ritual_master": ["skill_soul_art", "skill_soul_chain", "ms_soul_rend", "ms_bloodline_suppress"],

    # ── 莽荒古域 ──
    "boss_demon_python_king": ["ms_constrict", "ms_death_miasma", "ms_bloodline_suppress"],
    "boss_bodhi_tree_guardian": ["ms_emperor_will", "ms_ancient_seal", "skill_healing"],

    # ── 古帝洞府 ──
    "boss_tuoshe_emperor_remnant": ["skill_emperor_flame", "skill_emperor_seal", "ms_emperor_will", "ms_purifying_lotus"],
    "boss_emperor_pill_guardian": ["ms_emperor_will", "ms_ancient_seal", "ms_green_lotus_burn"],

    # ── 药族 ──
    "elite_yao_clan_alchemist": ["skill_alchemy", "ms_potion_master", "ms_meditate"],
    "enemy_yao_tian": ["skill_alchemy", "ms_potion_master"],
}
# Note: skill_nihility_flame and skill_emperor_flame are referenced but may need to be created
# Let me add them if missing
MISSING_BOSS_SKILLS = [
    ("skill_nihility_flame", "虚无吞炎", "源火能力", "天阶高级", "atk:+150,soul:+80", "源火榜第二，吞噬万物。"),
    ("skill_emperor_flame", "帝炎", "源火能力", "帝阶", "atk:+300,soul:+100", "万火归一，炎帝之焰。"),
]
for data in MISSING_BOSS_SKILLS:
    if data[0] not in existing_skills:
        row = {"Skill_ID": data[0], "Name": data[1], "Type": data[2], "Rank": data[3],
               "Effect": data[4], "Description": data[5]}
        append_row(ws_skills, row, skill_headers)
        existing_skills.add(data[0])
        print(f"  + missing boss skill: {data[0]}")

# ══════════════════════════════════════════════════════════
# PART 4: APPLY SKILLS TO ALL ENEMIES
# ══════════════════════════════════════════════════════════
# Check if "Skills" column exists in enemies sheet
if "Skills" not in enemy_headers:
    ws_enemies.cell(row=1, column=len(enemy_headers)+1).value = "Skills"
    enemy_headers.append("Skills")
    print("Added Skills column to Enemies sheet")

skills_col = enemy_headers.index("Skills")

updated = 0
for row_idx, row in enumerate(ws_enemies.iter_rows(min_row=2), start=2):
    eid = str(row[0].value)
    if not eid:
        continue
    name = str(row[1].value) if row[1].value else ""
    etype = str(row[2].value) if row[2].value else ""
    level = int(row[3].value) if row[3].value else 1

    # Boss unique skills take priority
    if eid in BOSS_UNIQUE:
        skills = BOSS_UNIQUE[eid]
    else:
        skills = get_enemy_skills(eid, name, etype, level)

    skill_str = ",".join(skills)
    ws_enemies.cell(row=row_idx, column=skills_col+1).value = skill_str
    updated += 1

wb.save(WORKBOOK_PATH)
print(f"Skills assigned to {updated} enemies.")
print(f"Total skills: {len(existing_skills)}")
print("Done!")
