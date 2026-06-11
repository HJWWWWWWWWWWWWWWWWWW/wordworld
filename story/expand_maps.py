"""
Expand map data: add missing main maps, sub-areas, encounters, and gathering spots.
Run this once to regenerate the xlsx with comprehensive map data.
"""
import copy
import sys
from pathlib import Path
from zipfile import ZipFile
import xml.etree.ElementTree as ET

import openpyxl

WORKBOOK_PATH = Path("story/text_game_event_schema_v4.xlsx")
BACKUP_PATH = Path("story/text_game_event_schema_v4_backup.xlsx")

# ── backup ──────────────────────────────────────────────────
if not BACKUP_PATH.exists():
    import shutil
    shutil.copy2(WORKBOOK_PATH, BACKUP_PATH)
    print(f"Backed up to {BACKUP_PATH}")

wb = openpyxl.load_workbook(WORKBOOK_PATH)

# ── sheet helpers ────────────────────────────────────────────
def sheet_rows(ws):
    """Yield dicts from an openpyxl worksheet. First row is headers."""
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    for row in ws.iter_rows(min_row=2):
        values = [cell.value for cell in row]
        yield {h: (v if v is not None else "") for h, v in zip(headers, values)}

def append_row(ws, data: dict):
    """Append a row to the worksheet, matching header order."""
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    row_data = []
    for h in headers:
        val = data.get(h, "")
        row_data.append(val if val is not None else "")
    ws.append(row_data)

def get_existing_ids(ws):
    """Get all existing IDs from the first column."""
    ids = set()
    for row in ws.iter_rows(min_row=2):
        val = row[0].value
        if val:
            ids.add(str(val))
    return ids

def find_sheet(name_prefix):
    """Find a sheet by prefix."""
    for name in wb.sheetnames:
        if name.startswith(name_prefix):
            return wb[name]
    raise KeyError(f"Sheet starting with '{name_prefix}' not found")

# ── locate sheets ────────────────────────────────────────────
ws_maps = find_sheet("Maps_")
ws_encounters = find_sheet("Map_Encounters_")
ws_gathering = find_sheet("Map_Gathering_")

existing_maps = get_existing_ids(ws_maps)
existing_encounters = get_existing_ids(ws_encounters)
existing_gathering = get_existing_ids(ws_gathering)

print(f"Existing maps: {len(existing_maps)}")
print(f"Existing encounters: {len(existing_encounters)}")
print(f"Existing gathering: {len(existing_gathering)}")

# ══════════════════════════════════════════════════════════════
# NEW MAP DATA
# ══════════════════════════════════════════════════════════════

NEW_MAIN_MAPS = [
    # ── yunlan_war: 重返加玛与云岚宗大战 ──
    ("map_cloud_mountain_peak", "云山巅峰", "云岚宗", "", 35, 10, "no", "", "云岚宗最高峰，云山修炼与决战之地。"),
    ("map_yan_alliance_hq", "炎盟总部", "加玛帝国", "", 35, 5, "yes", "", "萧炎在加玛帝国建立联盟的总部。"),
    ("map_jia_ma_battle_front", "加玛战线", "加玛帝国", "", 33, 8, "no", "", "炎盟与云岚宗残党交战的边境战场。"),

    # ── poison_sect_war: 出云帝国与毒宗之战 ──
    ("map_chuyun_empire", "出云帝国", "西北大陆", "", 35, 8, "no", "", "毒宗统治的出云帝国疆域。"),
    ("map_poison_sect", "毒宗", "出云帝国", "", 36, 10, "no", "", "出云帝国毒师宗门，阴毒功法汇聚之地。"),
    ("map_golden_goose_sect", "金雁宗", "西北大陆", "", 35, 8, "no", "", "金雁宗驻地，三宗联军之一。"),
    ("map_mulan_valley", "慕兰谷", "西北大陆", "", 35, 8, "no", "", "慕兰谷驻地，三宗联军之一。"),

    # ── revisit_tower: 再探塔底与天火尊者 ──
    ("map_skyfire_magma_world", "炼气塔岩浆世界", "天焚炼气塔", "", 42, 12, "no", "", "炼气塔底的岩浆空间，深处有火焰蜥蜴人与令古帝玉异动的神秘存在。"),

    # ── northwest_fortress_war: 玄黄要塞与西北大陆大战 ──
    ("map_xuanhuang_fortress", "玄黄要塞", "西北大陆", "", 56, 10, "no", "", "拱卫西北大陆的庞大军事要塞。"),
    ("map_northwest_battle_front", "西北战线", "西北大陆", "", 55, 10, "no", "", "魂殿联军与炎盟对峙的前线战场。"),

    # ── tianfu_alliance: 建立天府联盟 ──
    ("map_tianfu_council_hall", "天府议事厅", "星陨阁", "", 60, 5, "yes", "", "天府联盟高层议事的中枢场所。"),
    ("map_alliance_war_room", "联盟战备室", "星陨阁", "", 60, 5, "yes", "", "联盟情报汇总与战略推演的密室。"),

    # ── nether_spring: 九幽黄泉与妖暝 ──
    ("map_nether_spring", "九幽黄泉", "兽域", "", 62, 10, "no", "", "九幽地冥蟒族禁地，蕴含黄泉精血与远古传承。"),
    ("map_nether_python_tribe", "九幽地冥蟒族地", "兽域", "", 60, 8, "no", "", "九幽地冥蟒族的聚居领地。"),
    ("map_nether_underground_palace", "九幽地宫", "兽域", "", 63, 10, "no", "", "地冥蟒族地下宫殿，妖暝被囚之处。"),

    # ── post_demon_wars: 魂殿殿主与北龙王终战 ──
    ("map_soul_emperor_throne", "魂殿主殿", "魂殿", "", 72, 10, "no", "", "魂殿殿主所在的核心殿堂。"),

    # ── five_emperors: 五帝破空 ──
    ("map_world_gate", "世界通道", "斗气大陆", "", 85, 5, "yes", "", "斗帝强者合力打开的新世界入口。"),
    ("map_emperor_memorial_peak", "帝陨峰", "斗气大陆", "", 85, 5, "yes", "", "双帝之战遗迹与后世修炼者朝圣之地。"),
]

NEW_SUB_MAPS = [
    # ── black_corner_war: 收服陨落心炎与清算韩枫 ──
    ("map_feng_merchant_hall", "枫城商会", "枫城", "", 30, 5, "yes", "", "枫城商业交易与消息汇聚的中心。"),
    ("map_feng_alchemy_room", "枫城炼药密室", "枫城", "", 30, 5, "yes", "", "韩枫留下的炼药设施，残留异火痕迹。"),
    ("map_feng_defense_wall", "枫城城墙", "枫城", "", 30, 6, "no", "", "枫城外墙防线，黑盟入侵时的前沿阵地。"),

    # ── revive_mentor: 复活药老 ──
    ("map_star_realm_core", "星界核心", "星界", "", 48, 6, "yes", "", "星界空间枢纽，药老复活仪式的举行地。"),
    ("map_star_realm_training_ground", "星界修炼场", "星界", "", 48, 6, "yes", "", "星界内的灵气修炼场所。"),
    ("map_star_pavilion_council", "星陨阁议事大殿", "星陨阁", "", 48, 5, "yes", "", "重建后的星陨阁核心议事场所。"),

    # ── flower_sect: 花宗与云韵传承 ──
    ("map_flower_sect_gate", "花宗山门", "花宗", "", 52, 5, "yes", "", "花宗入口，灵气环绕的花海山道。"),
    ("map_flower_sect_garden", "花宗灵花园", "花宗", "", 52, 5, "yes", "", "培育珍稀灵花的宗内禁园。"),
    ("map_flower_sect_heritage_hall", "花宗传承殿", "花宗", "", 53, 6, "yes", "", "花宗历代宗主传承功法与信物的殿堂。"),

    # ── dragon_island_war: 古龙岛三岛大战 ──
    ("map_west_dragon_palace", "西龙岛王殿", "西龙岛", "", 62, 8, "no", "", "西龙王统治的宫殿区域。"),
    ("map_south_dragon_battlefield", "南龙岛战场", "南龙岛", "", 62, 10, "no", "", "三岛大战中立势力与南岛的交战区域。"),
    ("map_north_dragon_throne", "北龙王座", "北龙岛", "", 64, 10, "no", "", "北龙王的龙座大殿，龙族统一最后战场。"),

    # ── demon_flame: 净莲妖火 ──
    ("map_demon_flame_illusion_realm", "妖火幻境", "妖火空间", "", 66, 8, "no", "", "净莲妖火创造的幻境领域，考验闯入者心智。"),
    ("map_demon_flame_core", "妖火核心", "妖火空间", "", 68, 10, "no", "", "净莲妖火本体所在的核心区域。"),
    ("map_demon_flame_saint_remains", "净莲妖圣残像", "妖火空间", "", 68, 8, "no", "", "净莲妖圣残存意志守护的禁地。"),

    # ── medicine_ceremony: 药典与药族灭族战 ──
    ("map_yao_realm_ceremony_square", "药典广场", "药界", "", 68, 5, "yes", "", "药族举办药典大会的广场。"),
    ("map_yao_realm_herb_garden", "药族药园", "药界", "", 68, 5, "yes", "", "药族培育万年灵药的圣地。"),
    ("map_yao_realm_survivor_camp", "药族幸存者营地", "药界", "", 70, 6, "yes", "", "魂族灭族战后药族幸存者的藏身营地。"),

    # ── ancient_clan_war: 远古种族联盟战 ──
    ("map_soul_realm_battlefield", "魂界战场", "魂界", "", 85, 12, "no", "", "远古种族联盟与魂族决战的虚空战场。"),
    ("map_hun_clan_ritual_site", "魂族祭坛", "魂族空间", "", 85, 10, "no", "", "魂族用于夺取帝玉与血脉的仪式祭坛。"),
    ("map_ancient_alliance_camp", "远古联盟营地", "古界", "", 82, 5, "yes", "", "古族、炎族、雷族等远古种族的联军营地。"),

    # ── ancient_emperor: 古帝洞府 ──
    ("map_emperor_cave_gate", "古帝石门", "古帝洞府", "", 75, 8, "no", "", "陀舍古帝洞府入口的巨大石门。"),
    ("map_emperor_cave_inner", "洞府内殿", "古帝洞府", "", 78, 10, "no", "", "古帝核心传承所在的内殿空间。"),
    ("map_emperor_cave_treasure_room", "古帝丹室", "古帝洞府", "", 80, 10, "no", "", "帝品雏丹与斗帝功法存放之处。"),

    # ── final_war: 双帝之战 ──
    ("map_double_emperor_peak", "双帝峰顶", "双帝山脉", "", 90, 12, "no", "", "萧炎与魂天帝最终决战的双帝峰顶端。"),
    ("map_allied_forces_camp", "联军大营", "双帝山脉", "", 88, 5, "yes", "", "大陆联军在最终战场前的集结营地。"),

    # ── dan_meeting_flame: 星域 ──
    ("map_star_domain", "星域", "丹塔", "", 45, 10, "no", "", "三千焱炎火所在的星空领域，是丹会冠军的试炼终点。"),

    # ── chuyun_empire sub-areas ──
    ("map_chuyun_border", "出云边塞", "出云帝国", "", 34, 6, "no", "", "出云帝国与加玛帝国交界的边境防线。"),
    ("map_poison_sect_hall", "毒宗大殿", "毒宗", "", 37, 8, "no", "", "毒宗宗主议事与炼毒的阴暗大殿。"),
    ("map_poison_sect_herb_cave", "毒宗药窟", "毒宗", "", 36, 6, "no", "", "毒宗培育毒草与炼制毒丹的地下洞窟。"),

    # ── xuanhuang_fortress sub-areas ──
    ("map_xuanhuang_war_hall", "玄黄战备厅", "玄黄要塞", "", 56, 5, "yes", "", "玄黄要塞指挥作战与调度的核心厅堂。"),
    ("map_xuanhuang_defense_wall", "玄黄要塞城墙", "玄黄要塞", "", 57, 8, "no", "", "玄黄要塞面向魂殿联军的防御城墙。"),

    # ── nether_spring sub-areas ──
    ("map_nether_spring_pool", "黄泉血池", "九幽黄泉", "", 63, 8, "no", "", "蕴含九幽黄泉精血的血色深池。"),
    ("map_nether_python_throne", "冥蟒王座", "九幽地冥蟒族地", "", 62, 8, "no", "", "九幽地冥蟒族王座所在，曾属于妖暝。"),

    # ── world_gate sub-areas ──
    ("map_emperor_ascension_platform", "飞升台", "世界通道", "", 90, 5, "yes", "", "五帝准备破空而去的升天平台。"),

    # ── soul_hall sub-area expansions ──
    ("map_soul_hall_person_hall", "魂殿人殿", "魂殿", "", 65, 10, "no", "", "魂殿收集与囚禁灵魂体的核心分殿。"),
    ("map_soul_hall_soul_well", "魂殿灵魂井", "魂殿", "", 66, 8, "no", "", "魂殿囤积灵魂本源的深井，阴冷刺骨。"),

    # ── tianfu_alliance sub-area ──
    ("map_star_pavilion_alliance_hub", "星陨阁联盟枢纽", "星陨阁", "", 60, 5, "yes", "", "各大势力在星陨阁设立的联络与情报交换枢纽。"),

    # ── ancient_ruins more sub-areas ──
    ("map_ancient_ruins_core", "遗迹核心殿", "远古遗迹", "", 55, 10, "no", "", "远古遗迹最深处的核心殿堂，斗圣骸骨存放之处。"),
    ("map_beast_region_trade_hub", "兽域交易区", "兽域", "", 52, 5, "yes", "", "魔兽家族与人类势力交换物资的中立交易区。"),
]

# ══════════════════════════════════════════════════════════════
# ENCOUNTERS for new maps (at least 1 per map)
# ══════════════════════════════════════════════════════════════

def enc(evt_id, map_id, text, opt1_text="", opt1_effect="", opt2_text="", opt2_effect="", opt3_text="", opt3_effect="", weight=1):
    """Build an encounter row."""
    return {
        "Event_ID": evt_id,
        "Map_ID": map_id,
        "Weight": weight,
        "Text": text,
        "Pre_Condition": "",
        "Opt1_Text": opt1_text,
        "Opt1_Condition": "",
        "Opt1_Next": "",
        "Opt1_Effect": opt1_effect,
        "Opt2_Text": opt2_text,
        "Opt2_Condition": "",
        "Opt2_Next": "",
        "Opt2_Effect": opt2_effect,
        "Opt3_Text": opt3_text,
        "Opt3_Condition": "",
        "Opt3_Next": "",
        "Opt3_Effect": opt3_effect,
        "Notes": "",
    }

# Encounters for each new map
NEW_ENCOUNTERS = [
    # yunlan_war
    enc("enc_cloud_peak_1", "map_cloud_mountain_peak", "云山巅上，云雾翻涌。云山负手立于崖边，斗气威压如实质般铺开。", "正面迎战", "exp:+100,reputation:+10", "暂避锋芒", "exp:+30", "观察地形", "soul:+3"),
    enc("enc_yan_hq_1", "map_yan_alliance_hq", "炎盟大厅中人声鼎沸，各方将领正为战线部署争论不休。", "参与议事", "reputation:+5,rel:npc_xiao_zhan:+3", "暗中查探情报", "soul:+2", "整顿军备", "item:+item_elixir"),
    enc("enc_jia_ma_front_1", "map_jia_ma_battle_front", "前线告急，云岚宗残党与魂殿护法联手发动突袭。", "率队迎击", "exp:+80,reputation:+8", "设伏反击", "exp:+60,soul:+3"),

    # poison_sect_war
    enc("enc_chuyun_1", "map_chuyun_empire", "出云帝国的街市飘散着淡淡药味，毒师装扮的行人步履匆匆。", "打听毒宗情报", "soul:+2", "购买解毒药材", "item:+herb_coagulation", "低调通行", ""),
    enc("enc_poison_sect_1", "map_poison_sect", "毒宗大殿阴森潮湿，墙壁上爬满诡异的毒藤，空气中弥漫着令人晕眩的气息。", "闯入大殿", "exp:+100", "破解毒阵", "soul:+5,alchemy:+2"),
    enc("enc_golden_goose_1", "map_golden_goose_sect", "金雁宗弟子乘金雁盘旋于空，羽翼破风声尖锐刺耳。", "迎击金雁骑士", "exp:+80", "谈判周旋", "reputation:+3"),
    enc("enc_mulan_1", "map_mulan_valley", "慕兰谷中奇花异草遍布，但每一株都暗藏致命陷阱。", "穿越花谷", "exp:+70,soul:+2", "采集毒草", "item:+herb_desert_mandala"),

    # revisit_tower
    enc("enc_magma_world_1", "map_skyfire_magma_world", "岩浆翻涌，热浪扭曲了视线。岩浆深处似有巨物游动，古帝的气息若隐若现。", "深入岩浆", "exp:+120,soul:+5", "在边缘搜寻", "exp:+50,item:+core_fire", "感知古帝残念", "soul:+8"),

    # northwest_fortress_war
    enc("enc_xuanhuang_1", "map_xuanhuang_fortress", "玄黄要塞烽火连天，守军面色凝重。魂殿联军的黑旗在地平线上蔓延。", "督战鼓舞士气", "reputation:+10", "加固防御", "exp:+50", "出击迎敌", "exp:+100"),
    enc("enc_northwest_front_1", "map_northwest_battle_front", "西北战线尸横遍野，空气中弥漫着血腥与燃烧的斗气。", "冲锋陷阵", "exp:+120,reputation:+8", "救治伤员", "reputation:+5,rel:npc_xiao_yixian:+3"),

    # tianfu_alliance
    enc("enc_tianfu_hall_1", "map_tianfu_council_hall", "天府议事厅内悬挂着大陆地图，各势力首领围坐长桌，面色凝重。", "主持会议", "reputation:+8", "私下斡旋", "soul:+3,reputation:+3"),
    enc("enc_alliance_war_1", "map_alliance_war_room", "战备室的沙盘上标注着魂殿各个据点的位置，情报官员低声汇报。", "分析情报", "soul:+5", "部署兵力", "reputation:+5"),

    # nether_spring
    enc("enc_nether_spring_1", "map_nether_spring", "九幽黄泉深处传来低沉的龙吟，血池中的液体翻涌不止。", "跳入血池", "douqi:+15", "在岸边观察", "soul:+5"),
    enc("enc_nether_python_1", "map_nether_python_tribe", "地冥蟒族战士盘踞在入口，蛇瞳中闪烁着警惕的光芒。", "正面交涉", "reputation:+5", "潜入绕过", "exp:+60"),
    enc("enc_nether_palace_1", "map_nether_underground_palace", "地下宫殿的石壁上刻满蛇形符文，妖暝的封印在深处闪耀。", "破除封印", "flag:yaoming_restored=1,exp:+150", "先清除守卫", "exp:+100"),

    # post_demon_wars
    enc("enc_soul_throne_1", "map_soul_emperor_throne", "魂殿主殿深处，殿主端坐于王座之上，阴冷的魂力充斥整个空间。", "正面挑战", "exp:+200,reputation:+15", "联手围攻", "exp:+180"),

    # five_emperors
    enc("enc_world_gate_1", "map_world_gate", "世界通道缓缓打开，通道那头传来全新的气息。无数强者仰望天空。", "踏入通道", "flag:story_finished=1,exp:+500", "留下传承", "reputation:+50"),
    enc("enc_emperor_peak_1", "map_emperor_memorial_peak", "帝陨峰上铭刻着双帝之战的痕迹，后世修炼者在此感悟斗帝境界。", "感悟帝境", "soul:+20,douqi:+30", "祭拜先烈", "reputation:+10"),

    # black_corner_war sub-areas
    enc("enc_feng_hall_1", "map_feng_merchant_hall", "枫城商会中商人神色紧张，黑盟的消息让他们坐立不安。", "询问情报", "soul:+2", "购买物资", "item:+item_elixir"),
    enc("enc_feng_alchemy_1", "map_feng_alchemy_room", "韩枫的炼药室残留着海心焰的焦痕，未完成的丹方散落一地。", "研究丹方", "alchemy:+3", "收集残料", "item:+herb_green_flame_grass"),
    enc("enc_feng_wall_1", "map_feng_defense_wall", "城墙上的守军严阵以待，远处黑盟的旗帜正在逼近。", "协助防守", "exp:+80,reputation:+5", "出击突袭", "exp:+100"),

    # revive_mentor sub-areas
    enc("enc_star_core_1", "map_star_realm_core", "星界核心处，生骨融血丹的炼制已进入最关键阶段。", "主持炼制", "alchemy:+5,soul:+5", "护法守卫", "exp:+80"),
    enc("enc_star_train_1", "map_star_realm_training_ground", "星界修炼场灵气充沛，适合突破瓶颈。", "闭关修炼", "exp:+150,douqi:+5"),
    enc("enc_star_council_1", "map_star_pavilion_council", "议事大殿中，药老与风尊者正在讨论星陨阁的未来规划。", "参与议事", "reputation:+5,rel:npc_yao_lao:+3"),

    # flower_sect sub-areas
    enc("enc_flower_gate_1", "map_flower_sect_gate", "花宗山门两侧种满灵花，空气中弥漫着醉人芬芳。", "通报来意", "reputation:+3", "强行闯入", "exp:+60"),
    enc("enc_flower_garden_1", "map_flower_sect_garden", "灵花园中百花争艳，每一朵都蕴含着独特药性。", "采集灵花", "item:+herb_sky_jade_fruit", "静坐感悟", "soul:+5"),
    enc("enc_flower_heritage_1", "map_flower_sect_heritage_hall", "传承殿中悬浮着花宗历代宗主的影像，云韵正承受传承考验。", "守护传承", "rel:npc_yun_yun:+10,exp:+100"),

    # dragon_island_war sub-areas
    enc("enc_west_palace_1", "map_west_dragon_palace", "西龙王端坐于龙骨王座，龙目俯视着闯入者。", "挑战西龙王", "exp:+200,reputation:+10", "外交劝降", "reputation:+8,soul:+3"),
    enc("enc_south_battle_1", "map_south_dragon_battlefield", "南龙岛上龙息交错，三方混战的局面难以分清敌友。", "协助东龙岛", "exp:+180,rel:npc_ziyan:+5", "独立作战", "exp:+160"),
    enc("enc_north_throne_1", "map_north_dragon_throne", "北龙王坐在龙座之上，眼中闪烁着疯狂的光芒。", "发动总攻", "exp:+250,reputation:+15", "设阵困敌", "exp:+200,soul:+5"),

    # demon_flame sub-areas
    enc("enc_demon_illusion_1", "map_demon_flame_illusion_realm", "妖火幻境中浮现出最深的恐惧与欲望，每一幕都栩栩如生。", "硬抗幻境", "soul:+10", "寻找破绽", "soul:+8,exp:+50"),
    enc("enc_demon_core_1", "map_demon_flame_core", "净莲妖火本体熊熊燃烧，炽热的温度足以熔化灵魂。", "强行收服", "item:+item_purifying_demon_flame,douqi:+50", "联合压制", "exp:+200"),
    enc("enc_demon_saint_1", "map_demon_flame_saint_remains", "净莲妖圣的残像浮现于火焰之中，目光中带着悲悯。", "倾听教诲", "soul:+15", "请求帮助", "soul:+10,exp:+100"),

    # medicine_ceremony sub-areas
    enc("enc_yao_square_1", "map_yao_realm_ceremony_square", "药典广场上万药汇聚，各方炼药师展示着自己的巅峰之作。", "参加药典", "alchemy:+5,reputation:+10", "观摩学习", "alchemy:+3"),
    enc("enc_yao_garden_1", "map_yao_realm_herb_garden", "药族药园中生长着外界早已绝迹的万年灵药。", "采集灵药", "item:+herb_ancient_green_vine", "守护药园", "reputation:+5"),
    enc("enc_yao_survivor_1", "map_yao_realm_survivor_camp", "药族幸存者藏身于废墟之中，眼中仍残留着灭族之夜的恐惧。", "安抚救治", "reputation:+8,rel:npc_yao_lao:+5", "组织撤离", "exp:+100"),

    # ancient_clan_war sub-areas
    enc("enc_soul_battle_1", "map_soul_realm_battlefield", "魂界战场空间撕裂，远古种族联军与魂族的决战已至白热。", "冲锋陷阵", "exp:+250,reputation:+15", "侧翼包抄", "exp:+200,soul:+5"),
    enc("enc_hun_ritual_1", "map_hun_clan_ritual_site", "魂族祭坛上，帝玉碎片在血光中汇聚，魂天帝的身影若隐若现。", "打断仪式", "exp:+200,reputation:+10", "夺取帝玉", "flag:emperor_jade_crisis=1"),
    enc("enc_alliance_camp_1", "map_ancient_alliance_camp", "远古联盟营地中，各族首领齐聚，为最后的决战做准备。", "协调各族", "reputation:+10,soul:+3", "整编军队", "reputation:+8"),

    # ancient_emperor sub-areas
    enc("enc_emperor_gate_1", "map_emperor_cave_gate", "魂族已经用集齐的八块陀舍古帝玉开启石门，联盟大军必须立刻追入洞府。", "跟随联盟追入洞府", "exp:+200,flag:emperor_cave_opened=1", "观察洞府入口", "soul:+10"),
    enc("enc_emperor_inner_1", "map_emperor_cave_inner", "洞府内殿灵气浓郁到化雾，帝品丹药的香气若隐若现。", "深入探索", "exp:+250,douqi:+30", "小心前进", "exp:+150,soul:+5"),
    enc("enc_emperor_treasure_1", "map_emperor_cave_treasure_room", "古帝丹室中悬浮着帝品雏丹，周围布满了斗帝级别的禁制。", "夺取雏丹", "item:+item_embryonic_emperor_pill", "破解禁制", "soul:+15"),

    # final_war sub-areas
    enc("enc_double_peak_1", "map_double_emperor_peak", "双帝峰顶，萧炎与魂天帝遥遥对峙，天地为之变色。", "全力迎战", "exp:+500,reputation:+30", "布置阵法", "exp:+300,soul:+10"),
    enc("enc_allied_camp_1", "map_allied_forces_camp", "联军大营连绵百里，来自大陆各处的强者枕戈待旦。", "鼓舞士气", "reputation:+15", "部署战术", "soul:+5,reputation:+5"),

    # star_domain
    enc("enc_star_domain_1", "map_star_domain", "星域中繁星点点，三千焱炎火化作一条火龙在星空间游弋。", "收服星焱", "item:+item_three_thousand_flame,douqi:+30", "观察星空", "soul:+8"),

    # chuyun sub-areas
    enc("enc_chuyun_border_1", "map_chuyun_border", "出云边塞的哨塔上，毒宗探子的身影若隐若现。", "潜入边境", "exp:+60", "正面突破", "exp:+80,reputation:+5"),
    enc("enc_poison_hall_1", "map_poison_sect_hall", "毒宗大殿中毒雾弥漫，宗主蝎毕岩端坐于王座，蝎尾轻轻摆动。", "挑战蝎毕岩", "exp:+150,reputation:+10", "谈判结盟", "reputation:+5"),
    enc("enc_poison_cave_1", "map_poison_sect_herb_cave", "毒宗药窟深处生长着外界罕见的剧毒灵草，每走一步都需小心翼翼。", "采集毒材", "item:+herb_desert_mandala,alchemy:+2", "销毁毒窟", "reputation:+8"),

    # xuanhuang sub-areas
    enc("enc_xuanhuang_hall_1", "map_xuanhuang_war_hall", "战备厅的沙盘上标注着密密麻麻的敌军动向。", "制定作战计划", "soul:+5,reputation:+5", "调集援军", "reputation:+8"),
    enc("enc_xuanhuang_wall_1", "map_xuanhuang_defense_wall", "要塞城墙上，守军与魂殿联军展开了激烈的攻防战。", "坚守城墙", "exp:+150,reputation:+10", "率队出击", "exp:+180"),

    # nether_spring sub-areas
    enc("enc_nether_pool_1", "map_nether_spring_pool", "黄泉血池中的精血翻涌起浪，远古九幽蟒的虚影在池中游动。", "浸泡血池", "douqi:+20", "收集精血", "item:+item_huangquan_blood_crystal"),
    enc("enc_nether_throne_1", "map_nether_python_throne", "冥蟒王座被篡位者占据，妖暝的灵魂在暗处注视着一切。", "帮助妖暝复位", "flag:yaoming_restored=1,exp:+200", "旁观局势", "soul:+5"),

    # soul_hall sub-areas
    enc("enc_soul_person_1", "map_soul_hall_person_hall", "魂殿人殿中囚禁着无数灵魂体，哀嚎声不绝于耳。", "释放灵魂", "soul:+10,reputation:+10", "摧毁设施", "exp:+150"),
    enc("enc_soul_well_1", "map_soul_hall_soul_well", "灵魂井深不见底，井口的魂力波动令人灵魂震颤。", "吸取魂力", "soul:+20", "封印井口", "exp:+120,soul:+5"),

    # tianfu alliance hub
    enc("enc_alliance_hub_1", "map_star_pavilion_alliance_hub", "联盟枢纽中各方使者往来穿梭，情报如流水般汇聚。", "整合情报", "soul:+5,reputation:+3", "接取联盟任务", "exp:+80"),

    # ancient_ruins core
    enc("enc_ruins_core_1", "map_ancient_ruins_core", "遗迹核心殿中悬浮着一具散发威压的斗圣骸骨。", "夺取骸骨", "flag:saint_bones_acquired=1,exp:+200", "先探查陷阱", "soul:+8"),
    enc("enc_beast_hub_1", "map_beast_region_trade_hub", "兽域交易区内，化形魔兽与人类商贩讨价还价。", "交易魔兽材料", "item:+core_magic", "打听情报", "soul:+2"),

    # world_gate sub-area
    enc("enc_ascend_1", "map_emperor_ascension_platform", "飞升台上，五位斗帝的身影在光芒中若隐若现。", "踏上升仙台", "flag:story_finished=1,reputation:+100", "目送前辈", "soul:+20,reputation:+30"),
]

# ══════════════════════════════════════════════════════════════
# GATHERING SPOTS (expand for more maps)
# ══════════════════════════════════════════════════════════════

def gather(gid, map_id, item_id, chance, min_q=1, max_q=3, pre="", effect="", limit="", exclusive="", notes=""):
    return {
        "Gather_ID": gid,
        "Map_ID": map_id,
        "Item_ID": item_id,
        "Chance_Percent": chance,
        "Min_Qty": min_q,
        "Max_Qty": max_q,
        "Pre_Condition": pre,
        "Explore_Effect": effect,
        "Unique_Limit": limit,
        "Location_Exclusive": exclusive,
        "Notes": notes,
    }

NEW_GATHERING = [
    # wutan region - basic herbs
    gather("gather_wutan_back_mountain_herb", "map_wutan_back_mountain", "herb_coagulation", 40),
    gather("gather_wutan_back_mountain_fruit", "map_wutan_back_mountain", "herb_qi_fruit", 25),
    gather("gather_wutan_back_spirit", "map_wutan_back_mountain", "herb_spirit_gathering", 15),

    # wutan commercial street - chance to find items
    gather("gather_wutan_street_elixir", "map_wutan_commercial_street", "item_elixir", 10),

    # wutan pharmacy
    gather("gather_wutan_pharmacy_herb", "map_wutan_pharmacy", "herb_coagulation", 50, 1, 5),
    gather("gather_wutan_pharmacy_fruit", "map_wutan_pharmacy", "herb_qi_fruit", 35, 1, 3),

    # jia_ma_road
    gather("gather_jia_ma_road_herb", "map_jia_ma_road", "herb_purple_leaf", 20),
    gather("gather_jia_ma_road_spirit", "map_jia_ma_road", "herb_spirit_gathering", 15),

    # jia_ma_border
    gather("gather_jia_ma_border_herb", "map_jia_ma_border", "herb_bone_wash", 18),
    gather("gather_jia_ma_border_core", "map_jia_ma_border", "core_earth", 12),

    # jia_ma_mountain_pass
    gather("gather_mountain_pass_herb", "map_jia_ma_mountain_pass", "herb_purple_leaf", 22),
    gather("gather_mountain_pass_core", "map_jia_ma_mountain_pass", "core_earth", 15),

    # magic_mountains sub-areas
    gather("gather_magic_herb_valley", "map_magic_herb_valley", "herb_purple_blood_lingzhi", 15, 1, 2),
    gather("gather_magic_herb_valley_lingzhi", "map_magic_herb_valley", "herb_ink_leaf_lotus", 20),
    gather("gather_magic_inner_core", "map_magic_inner", "core_magic", 30),
    gather("gather_magic_inner_herb", "map_magic_inner", "herb_snake_saliva_fruit", 18),
    gather("gather_magic_hidden_cave", "map_magic_hidden_cave", "herb_purple_blood_lingzhi", 8, 1, 1),
    gather("gather_magic_hidden_cave_core", "map_magic_hidden_cave", "core_wind", 12),
    gather("gather_wolfhead_camp", "map_wolfhead_camp", "core_wood", 15),

    # qingshan sub-areas
    gather("gather_qingshan_camp_core", "map_qingshan_mercenary_camp", "core_magic", 20),
    gather("gather_qingshan_medical", "map_qingshan_medical_hall", "herb_coagulation", 40, 1, 3),
    gather("gather_qingshan_market_core", "map_qingshan_market", "core_magic", 25),

    # tager desert sub-areas
    gather("gather_desert_trade_route", "map_desert_trade_route", "herb_desert_mandala", 20),
    gather("gather_snake_oasis", "map_snake_oasis", "herb_ice_flame_grass", 10),
    gather("gather_snake_oasis_water", "map_snake_oasis", "core_water", 15),
    gather("gather_desert_camp", "map_desert_camp", "item_elixir", 8),
    gather("gather_snake_temple", "map_snake_temple_outer", "core_fire", 12),
    gather("gather_desert_salt_lake", "map_desert_salt_lake", "herb_blood_lotus_essence", 8),
    gather("gather_desert_well", "map_desert_ancient_well", "herb_earth_fire_lotus_seed", 3, 1, 1),

    # jia_ma_capital sub-areas
    gather("gather_capital_commercial", "map_capital_commercial", "item_elixir", 12),
    gather("gather_capital_alchemist", "map_capital_alchemist_market", "herb_spirit_gathering", 30),
    gather("gather_capital_alchemist_herb", "map_capital_alchemist_market", "herb_green_flame_grass", 15),
    gather("gather_capital_nalan", "map_capital_nalan_mansion", "item_elixir", 10),
    gather("gather_capital_miteer", "map_capital_miteer_hq", "silver", 40, 5, 20),

    # yunlan sub-areas
    gather("gather_yunlan_back_cliff", "map_yunlan_back_cliff", "herb_green_flame_grass", 12),
    gather("gather_yunlan_back_cliff_core", "map_yunlan_back_cliff", "core_wind", 15),

    # canaan sub-areas
    gather("gather_canaan_library", "map_canaan_library", "item_elixir", 5),
    gather("gather_canaan_mission", "map_canaan_mission_hall", "core_magic", 18),

    # canaan_inner sub-areas
    gather("gather_inner_arena", "map_inner_arena", "core_magic", 20),
    gather("gather_inner_trade", "map_inner_trade_district", "herb_spirit_gathering", 25),

    # skyfire sub-areas
    gather("gather_skyfire_lower", "map_skyfire_lower", "herb_green_flame_grass", 22),
    gather("gather_skyfire_lower_core", "map_skyfire_lower", "core_fire", 18),
    gather("gather_skyfire_seal", "map_skyfire_seal_core", "item_fallen_heart_flame", 1, 1, 1, "flag:fallen_heart_unstable=1", "", "1", ""),

    # black_corner sub-areas
    gather("gather_black_blood_plain", "map_black_blood_plain", "core_magic", 22),
    gather("gather_black_herb_market", "map_black_herb_market", "herb_blood_essence_fruit", 10),
    gather("gather_black_herb_market_hemp", "map_black_herb_market", "herb_black_hemp", 25),

    # zhongzhou sub-areas
    gather("gather_zhongzhou_market", "map_zhongzhou_north_market", "herb_sky_jade_fruit", 12),
    gather("gather_zhongzhou_market_core", "map_zhongzhou_north_market", "core_magic", 20),
    gather("gather_tianbei_han", "map_tianbei_han_clan", "herb_spirit_gathering", 20),
    gather("gather_tianbei_hong", "map_tianbei_hong_clan", "core_fire", 15),

    # ye_city sub-areas
    gather("gather_ye_mansion", "map_ye_mansion", "herb_green_wood_vine", 15),
    gather("gather_ye_alchemy", "map_ye_alchemy_room", "herb_green_flame_grass", 25),

    # dan_region sub-areas
    gather("gather_dan_herb_street", "map_dan_herb_street", "herb_ancient_green_vine", 8),
    gather("gather_dan_herb_street_jade", "map_dan_herb_street", "herb_jade_bone_fruit", 12),
    gather("gather_sacred_dan_market", "map_sacred_dan_market", "item_elixir", 15),
    gather("gather_dan_tower_trial", "map_dan_tower_trial_room", "core_fire", 10),
    gather("gather_dan_beast", "map_dan_beast_enclosure", "herb_dragon_saliva", 10),

    # star_pavilion sub-areas
    gather("gather_star_back_mountain", "map_star_pavilion_back_mountain", "herb_spirit_gathering", 30),
    gather("gather_star_back_herb", "map_star_pavilion_back_mountain", "herb_green_wood_vine", 15),

    # soul_hall sub-area
    gather("gather_soul_prison", "map_soul_hall_prison", "item_soul_baby_fruit", 3, 1, 1),

    # ancient_ruins sub-areas
    gather("gather_ruins_gate", "map_ancient_ruins_gate", "core_earth", 15),
    gather("gather_beast_bone", "map_beast_bone_mountains", "core_magic", 25),
    gather("gather_beast_bone_jade", "map_beast_bone_mountains", "herb_jade_bone_fruit", 8),

    # wilderness sub-areas
    gather("gather_wilderness_outpost", "map_wilderness_outpost", "core_magic", 18),
    gather("gather_wilderness_poison", "map_wilderness_poison_swamp", "herb_ancient_green_vine", 6),

    # dragon island sub-areas
    gather("gather_dragon_harbor", "map_dragon_island_harbor", "core_magic", 20),
    gather("gather_dragon_harbor_core", "map_dragon_island_harbor", "core_wind", 12),

    # ancient_city sub-area
    gather("gather_ancient_city_market", "map_ancient_city_market", "herb_sky_jade_fruit", 15),

    # heaven_tomb sub-area
    gather("gather_heaven_tomb_camp", "map_heaven_tomb_camp", "herb_soul_baby_fruit", 5),

    # new main maps gathering
    gather("gather_chuyun_empire", "map_chuyun_empire", "herb_desert_mandala", 25),
    gather("gather_poison_sect", "map_poison_sect", "herb_blood_lotus_essence", 12),
    gather("gather_mulan_valley", "map_mulan_valley", "herb_purple_leaf", 30),
    gather("gather_xuanhuang_fortress", "map_xuanhuang_fortress", "core_earth", 20),
    gather("gather_northwest_front", "map_northwest_battle_front", "core_magic", 22),
    gather("gather_nether_spring", "map_nether_spring", "item_huangquan_blood_crystal", 5, 1, 1),
    gather("gather_nether_python_tribe", "map_nether_python_tribe", "core_fire", 18),
    gather("gather_magma_world", "map_skyfire_magma_world", "core_fire", 30),
    gather("gather_magma_world_lotus", "map_skyfire_magma_world", "herb_earth_fire_lotus_seed", 8),

    # new sub-areas gathering
    gather("gather_feng_alchemy", "map_feng_alchemy_room", "herb_green_flame_grass", 20),
    gather("gather_star_realm_core", "map_star_realm_core", "herb_spirit_gathering", 35),
    gather("gather_flower_garden", "map_flower_sect_garden", "herb_sky_jade_fruit", 20),
    gather("gather_west_dragon_palace", "map_west_dragon_palace", "core_magic", 25),
    gather("gather_demon_core", "map_demon_flame_core", "item_purifying_demon_flame", 1, 1, 1, "flag:demon_flame_map=1", "", "1", ""),
    gather("gather_yao_garden", "map_yao_realm_herb_garden", "herb_ancient_green_vine", 20),
    gather("gather_yao_garden_jade", "map_yao_realm_herb_garden", "herb_jade_bone_fruit", 15),
    gather("gather_soul_battlefield", "map_soul_realm_battlefield", "core_magic", 30),
    gather("gather_emperor_inner", "map_emperor_cave_inner", "core_fire", 25),
    gather("gather_emperor_treasure", "map_emperor_cave_treasure_room", "item_embryonic_emperor_pill", 1, 1, 1, "flag:emperor_cave_opened=1", "", "1", ""),
    gather("gather_double_peak", "map_double_emperor_peak", "item_emperor_flame", 0.5, 1, 1, "flag:soul_emperor_defeated=1", "", "1", ""),
    gather("gather_star_domain", "map_star_domain", "item_three_thousand_flame", 1, 1, 1, "flag:dan_meeting_won=1", "", "1", ""),
    gather("gather_chuyun_border", "map_chuyun_border", "herb_coagulation", 25),
    gather("gather_poison_cave", "map_poison_sect_herb_cave", "herb_desert_mandala", 30),
    gather("gather_nether_pool", "map_nether_spring_pool", "item_huangquan_blood_crystal", 8, 1, 2),
    gather("gather_alliance_hub", "map_star_pavilion_alliance_hub", "item_elixir", 15),
    gather("gather_ruins_core", "map_ancient_ruins_core", "core_earth", 20),
    gather("gather_beast_hub", "map_beast_region_trade_hub", "core_magic", 30),
]

# ══════════════════════════════════════════════════════════════
# APPLY TO XLSX
# ══════════════════════════════════════════════════════════════

all_new_maps = NEW_MAIN_MAPS + NEW_SUB_MAPS
print(f"\nAdding {len(all_new_maps)} new maps...")
for data in all_new_maps:
    map_id = data[0]
    if map_id in existing_maps:
        print(f"  SKIP (exists): {map_id}")
        continue
    row = {
        "Map_ID": data[0],
        "Name": data[1],
        "Region": data[2],
        "Unlock_Condition": data[3],
        "Recommend_Level": data[4],
        "Explore_Stamina_Cost": data[5],
        "Default_Encounter_Weight": 1,
        "Safe_Zone": "yes" if data[6] == "yes" else "no",
        "Exit_Event": data[7] if data[7] else "",
        "Description": data[8],
        "Notes": "",
    }
    append_row(ws_maps, row)
    existing_maps.add(map_id)
    print(f"  + {map_id}")

print(f"\nAdding {len(NEW_ENCOUNTERS)} new encounters...")
for data in NEW_ENCOUNTERS:
    evt_id = data["Event_ID"]
    if evt_id in existing_encounters:
        print(f"  SKIP (exists): {evt_id}")
        continue
    append_row(ws_encounters, data)
    existing_encounters.add(evt_id)
    print(f"  + {evt_id}")

print(f"\nAdding {len(NEW_GATHERING)} new gathering spots...")
for data in NEW_GATHERING:
    gid = data["Gather_ID"]
    if gid in existing_gathering:
        print(f"  SKIP (exists): {gid}")
        continue
    append_row(ws_gathering, data)
    existing_gathering.add(gid)
    print(f"  + {gid}")

# ── save ─────────────────────────────────────────────────────
wb.save(WORKBOOK_PATH)
print(f"\nSaved to {WORKBOOK_PATH}")
print("Done!")
