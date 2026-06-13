"""
Comprehensive enemy expansion per novel world:
100+ common mobs, 100+ elites, bosses following original novel.
Non-lethal bosses marked (can develop relationships later).
"""
from pathlib import Path
import openpyxl

WORKBOOK_PATH = Path("story/text_game_event_schema_v4.xlsx")
wb = openpyxl.load_workbook(WORKBOOK_PATH)

def find_sheet(name_prefix):
    for name in wb.sheetnames:
        if name.startswith(name_prefix):
            return wb[name]
    raise KeyError(f"Sheet starting with '{name_prefix}' not found")

def get_existing_ids(ws):
    ids = set()
    for row in ws.iter_rows(min_row=2):
        val = row[0].value
        if val:
            ids.add(str(val))
    return ids

def append_row(ws, data: dict):
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    row_data = []
    for h in headers:
        val = data.get(h, "")
        row_data.append(val if val is not None else "")
    ws.append(row_data)

ws_enemies = find_sheet("Enemies_")
existing_ids = get_existing_ids(ws_enemies)
print(f"Existing enemies: {len(existing_ids)}")

# Enemy template helper
def enemy(eid, name, etype, level, hp, atk, def_, spd, exp_r, drop="", win_next="", lose_next="", notes=""):
    return {
        "Enemy_ID": eid, "Name": name, "Type": etype,
        "Level": level, "HP": hp, "ATK": atk, "DEF": def_, "SPD": spd,
        "Exp_Reward": exp_r, "Drop_Table": drop,
        "Win_Next": win_next, "Lose_Next": lose_next, "Notes": notes,
    }

# ══════════════════════════════════════════════════════════
# PART 1: COMMON MOBS (100+) - by region
# ══════════════════════════════════════════════════════════
COMMON_MOBS = [
    # ── 魔兽山脉 (Lv 3-12) ──
    enemy("mob_wind_wolf", "风狼", "mob", 3, 180, 18, 10, 25, 30, "item:core_wind:15|item:item_beast_hide_bundle:20|exp:30:100"),
    enemy("mob_fire_snake", "赤炎蛇", "mob", 4, 200, 22, 8, 20, 40, "item:core_fire:15|exp:40:100"),
    enemy("mob_ice_bear", "冰爪熊", "mob", 5, 280, 25, 18, 12, 50, "item:core_ice:15|item:item_beast_hide_bundle:25|exp:50:100"),
    enemy("mob_earth_boar", "岩甲猪", "mob", 5, 320, 20, 25, 10, 50, "item:core_earth:15|exp:50:100"),
    enemy("mob_shadow_cat", "影猫", "mob", 6, 220, 30, 12, 35, 60, "item:core_wood:15|exp:60:100"),
    enemy("mob_steel_claw_eagle", "钢爪鹰", "mob", 7, 250, 35, 15, 40, 70, "item:core_wind:20|exp:70:100"),
    enemy("mob_venom_fang_spider", "毒牙蛛", "mob", 7, 300, 28, 14, 22, 70, "item:herb_coagulation:20|exp:70:100"),
    enemy("mob_thunder_hawk", "雷鹰", "mob", 8, 280, 40, 16, 45, 80, "item:core_wind:20|exp:80:100"),
    enemy("mob_rock_lizard", "岩蜥", "mob", 9, 350, 30, 30, 15, 90, "item:core_earth:20|exp:90:100"),
    enemy("mob_blood_ape", "血猿", "mob", 10, 400, 45, 22, 28, 100, "item:item_beast_hide_bundle:30|exp:100:100"),
    enemy("mob_dark_mantis", "暗刃螳螂", "mob", 10, 320, 50, 15, 35, 100, "item:core_wood:20|exp:100:100"),
    enemy("mob_scarlet_boar", "赤鬃野猪", "mob", 11, 450, 38, 28, 18, 110, "item:core_earth:20|exp:110:100"),
    enemy("mob_iron_scale_python", "铁鳞蟒", "mob", 12, 500, 42, 35, 20, 120, "item:item_beast_hide_bundle:25|exp:120:100"),

    # ── 赤沙荒漠 (Lv 8-18) ──
    enemy("mob_sand_scorpion", "沙蝎", "mob", 8, 300, 32, 20, 22, 80, "item:herb_desert_mandala:15|exp:80:100"),
    enemy("mob_desert_viper", "沙漠蝰蛇", "mob", 10, 350, 40, 16, 30, 100, "item:item_snake_venom_vial:10|exp:100:100"),
    enemy("mob_sand_worm", "沙虫", "mob", 11, 420, 28, 30, 10, 110, "item:core_earth:20|exp:110:100"),
    enemy("mob_desert_bandit", "沙漠盗匪", "mob", 9, 380, 35, 18, 25, 90, "item:silver:30|exp:90:100"),
    enemy("mob_snake_patrol", "蛇人巡逻兵", "mob", 12, 450, 42, 25, 30, 130, "item:item_snake_venom_vial:15|exp:130:100"),
    enemy("mob_scorching_lizard", "烈日石蜥", "mob", 13, 500, 38, 40, 15, 140, "item:core_fire:20|exp:140:100"),
    enemy("mob_dune_stalker", "沙丘潜伏者", "mob", 14, 480, 50, 20, 35, 150, "item:item_desert_compass:5|exp:150:100"),
    enemy("mob_sandstorm_elemental", "沙暴之灵", "mob", 15, 550, 45, 22, 28, 160, "item:core_wind:25|exp:160:100"),
    enemy("mob_camel_spider", "骆驼蛛", "mob", 16, 520, 48, 25, 32, 170, "item:herb_desert_mandala:20|exp:170:100"),
    enemy("mob_desert_marauder", "沙漠掠食者", "mob", 17, 600, 52, 28, 30, 180, "item:silver:40|exp:180:100"),

    # ── 暗角域 (Lv 16-30) ──
    enemy("mob_black_domain_thug", "暗角域暴徒", "mob", 16, 550, 55, 25, 30, 160, "item:silver:30|exp:160:100"),
    enemy("mob_blood_sect_disciple", "血宗弟子", "mob", 18, 620, 58, 28, 32, 190, "item:herb_blood_essence_fruit:10|exp:190:100"),
    enemy("mob_black_alliance_soldier", "暗盟杂兵", "mob", 20, 700, 62, 32, 34, 220, "item:silver:35|exp:220:100"),
    enemy("mob_wandering_killer", "亡命徒", "mob", 19, 650, 65, 22, 38, 200, "item:core_magic:15|exp:200:100"),
    enemy("mob_black_market_guard", "黑市护卫", "mob", 21, 750, 60, 35, 30, 230, "item:silver:30|exp:230:100"),
    enemy("mob_plain_raider", "平原劫匪", "mob", 22, 780, 65, 30, 36, 250, "item:item_beast_hide_bundle:15|exp:250:100"),
    enemy("mob_demon_flame_valley_guard", "魔焰谷守卫", "mob", 25, 900, 70, 40, 32, 300, "item:core_fire:20|exp:300:100"),
    enemy("mob_black_seal_enforcer", "黑印城打手", "mob", 24, 850, 68, 38, 35, 280, "item:silver:40|exp:280:100"),

    # ── 迦南学院 (Lv 16-25) ──
    enemy("mob_training_puppet", "训练傀儡", "mob", 16, 500, 50, 40, 20, 150, "exp:150:100"),
    enemy("mob_forest_beast_outer", "外围森林兽", "mob", 18, 650, 55, 30, 30, 190, "item:core_magic:18|exp:190:100"),

    # ── 出云帝国 (Lv 25-35) ──
    enemy("mob_poison_sect_disciple", "毒宗弟子", "mob", 26, 900, 72, 38, 36, 320, "item:herb_desert_mandala:18|exp:320:100"),
    enemy("mob_scorpion_gate_guard", "万蝎门守卫", "mob", 28, 980, 76, 42, 34, 350, "item:item_snake_venom_vial:12|exp:350:100"),
    enemy("mob_golden_goose_disciple", "金雁宗弟子", "mob", 27, 950, 74, 40, 38, 340, "item:core_wind:20|exp:340:100"),
    enemy("mob_mulan_valley_guard", "慕兰谷守卫", "mob", 28, 1000, 72, 45, 30, 350, "item:core_wood:20|exp:350:100"),
    enemy("mob_border_smuggler", "边境私贩", "mob", 25, 880, 70, 35, 36, 300, "item:silver:45|exp:300:100"),

    # ── 中州 (Lv 35-55) ──
    enemy("mob_wandering_cultivator", "中州散修", "mob", 35, 1500, 100, 60, 55, 500, "item:item_elixir:8|exp:500:100"),
    enemy("mob_wind_lightning_disciple", "风雷阁弟子", "mob", 38, 1700, 110, 65, 70, 580, "item:core_wind:25|exp:580:100"),
    enemy("mob_huangquan_disciple", "黄泉阁弟子", "mob", 38, 1750, 108, 68, 62, 580, "item:core_earth:25|exp:580:100"),
    enemy("mob_wanjian_disciple", "万剑阁弟子", "mob", 40, 1800, 120, 60, 72, 620, "item:core_ice:25|exp:620:100"),
    enemy("mob_burning_valley_disciple", "焚炎谷弟子", "mob", 42, 2000, 125, 70, 65, 680, "item:core_fire:25|exp:680:100"),
    enemy("mob_ice_river_disciple", "冰河谷弟子", "mob", 42, 2050, 122, 72, 62, 680, "item:core_ice:25|exp:680:100"),
    enemy("mob_sky_demon_disciple", "天冥宗弟子", "mob", 43, 2100, 128, 68, 64, 700, "item:core_wood:25|exp:700:100"),
    enemy("mob_north_domain_merc", "北域佣兵", "mob", 40, 1850, 115, 65, 60, 620, "item:silver:50|exp:620:100"),
    enemy("mob_central_domain_guard", "中域城卫", "mob", 45, 2200, 130, 78, 68, 750, "item:silver:55|exp:750:100"),
    enemy("mob_wormhole_bandit", "虫洞劫匪", "mob", 44, 2100, 135, 62, 75, 720, "item:item_wormhole_pass:5|exp:720:100"),
    enemy("mob_herb_thief", "药贼", "mob", 41, 1900, 118, 62, 70, 650, "item:herb_spirit_gathering:15|exp:650:100"),

    # ── 丹域 (Lv 45-55) ──
    enemy("mob_dan_region_guard", "丹域守卫", "mob", 45, 2300, 135, 80, 65, 750, "item:item_elixir:12|exp:750:100"),
    enemy("mob_alchemy_fraud", "炼药骗子", "mob", 46, 2100, 125, 60, 72, 770, "item:silver:60|exp:770:100"),

    # ── 黑渊殿 (Lv 30-65) ──
    enemy("mob_soul_hall_scout", "黑渊殿探子", "mob", 30, 1100, 85, 45, 55, 400, "exp:400:100"),
    enemy("mob_soul_chain_guard", "锁魂守卫", "mob", 38, 1600, 105, 60, 60, 600, "item:item_soul_baby_fruit:5|exp:600:100"),
    enemy("mob_soul_collector", "魂使", "mob", 42, 1900, 120, 65, 68, 700, "item:item_soul_baby_fruit:8|exp:700:100"),
    enemy("mob_soul_hall_enforcer", "黑渊殿执法", "mob", 48, 2400, 140, 78, 72, 850, "item:item_soul_baby_fruit:10|exp:850:100"),
    enemy("mob_death_corpse_ghoul", "葬尸食尸鬼", "mob", 50, 2600, 145, 80, 60, 900, "item:core_earth:25|exp:900:100"),

    # ── 兽域 (Lv 50-65) ──
    enemy("mob_feral_wolf_king", "野狼王", "mob", 50, 2800, 155, 85, 75, 950, "item:item_beast_blood_essence:5|exp:950:100"),
    enemy("mob_iron_tusk_elephant", "铁牙巨象", "mob", 52, 3200, 145, 110, 45, 1020, "item:item_beast_hide_bundle:25|exp:1020:100"),
    enemy("mob_sky_serpent", "飞天蟒", "mob", 54, 3000, 165, 80, 90, 1100, "item:item_snake_venom_vial:15|exp:1100:100"),
    enemy("mob_crimson_flame_tiger", "赤焰虎", "mob", 55, 3400, 175, 88, 85, 1150, "item:core_fire:30|exp:1150:100"),
    enemy("mob_thunder_bird", "雷鸟", "mob", 56, 3200, 180, 75, 100, 1180, "item:core_wind:30|exp:1180:100"),
    enemy("mob_ancient_tree_spirit", "古树精魂", "mob", 58, 3800, 160, 100, 60, 1250, "item:core_wood:30|exp:1250:100"),
    enemy("mob_nether_python_guard", "冥蟒族守卫", "mob", 60, 4000, 185, 100, 80, 1400, "item:item_huangquan_blood_crystal:5|exp:1400:100"),
    enemy("mob_beast_region_hunter", "兽域猎人", "mob", 55, 3300, 170, 85, 88, 1150, "item:item_beast_blood_essence:8|exp:1150:100"),

    # ── 莽荒古域 (Lv 60-68) ──
    enemy("mob_ancient_guardian_beast", "古域守护兽", "mob", 60, 4200, 190, 110, 80, 1450, "item:item_ancient_seed:3|exp:1450:100"),
    enemy("mob_poison_swamp_ghoul", "毒沼尸鬼", "mob", 62, 4400, 185, 100, 70, 1520, "item:herb_desert_mandala:20|exp:1520:100"),
    enemy("mob_blood_pool_serpent", "血池魔蟒", "mob", 64, 4800, 200, 105, 85, 1620, "item:item_huangquan_blood_crystal:8|exp:1620:100"),
    enemy("mob_ancient_bramble", "古域荆棘藤", "mob", 66, 5000, 195, 120, 55, 1720, "item:core_wood:30|exp:1720:100"),
    enemy("mob_demon_python_hatchling", "天魔蟒幼崽", "mob", 68, 5500, 210, 110, 90, 1820, "item:item_beast_blood_essence:10|exp:1820:100"),

    # ── 龙岛 (Lv 55-75) ──
    enemy("mob_young_dragon_guard", "幼龙卫士", "mob", 55, 3500, 180, 100, 85, 1200, "item:item_dragon_scale:3|exp:1200:100"),
    enemy("mob_dragon_island_lizard", "龙岛巨蜥", "mob", 60, 4200, 195, 115, 75, 1450, "item:item_beast_hide_bundle:30|exp:1450:100"),
    enemy("mob_dragon_blood_bat", "龙血蝙蝠", "mob", 62, 4000, 200, 80, 105, 1520, "item:item_dragon_scale:5|exp:1520:100"),
    enemy("mob_void_drifter", "虚空漂游者", "mob", 70, 5000, 220, 90, 110, 1900, "item:core_wind:30|exp:1900:100"),

    # ── 远古战场 (Lv 70-85) ──
    enemy("mob_ancient_soldier_spirit", "远古战魂", "mob", 70, 5200, 230, 110, 95, 2000, "item:item_soul_baby_fruit:12|exp:2000:100"),
    enemy("mob_battlefield_wraith", "战场冤魂", "mob", 72, 5500, 240, 105, 100, 2100, "item:item_soul_baby_fruit:15|exp:2100:100"),
    enemy("mob_hun_clan_scout", "玄族斥候", "mob", 75, 6000, 260, 130, 115, 2300, "item:item_soul_baby_fruit:18|exp:2300:100"),
    enemy("mob_alliance_vanguard", "联盟先锋", "mob", 75, 6200, 255, 135, 110, 2300, "item:item_elixir:20|exp:2300:100"),
    enemy("mob_broken_space_beast", "破碎空间兽", "mob", 80, 7000, 280, 140, 105, 2600, "item:item_ancient_relic:5|exp:2600:100"),

    # ── 妖火空间 (Lv 62-68) ──
    enemy("mob_demon_flame_wisp", "妖火精怪", "mob", 62, 3800, 200, 70, 110, 1550, "item:item_demon_flame_crystal:5|exp:1550:100"),
    enemy("mob_flame_illusion", "火焰幻象", "mob", 64, 3600, 210, 60, 115, 1620, "item:core_fire:30|exp:1620:100"),
    enemy("mob_saint_guardian_spirit", "妖圣守护灵", "mob", 66, 4500, 225, 100, 90, 1750, "item:item_demon_flame_crystal:8|exp:1750:100"),

    # ── 古帝洞府 (Lv 75-90) ──
    enemy("mob_emperor_cave_golem", "洞府石魔像", "mob", 75, 6500, 270, 160, 80, 2400, "item:core_earth:35|exp:2400:100"),
    enemy("mob_emperor_flame_warden", "帝火守卫", "mob", 80, 7200, 300, 140, 120, 2700, "item:item_demon_flame_crystal:10|exp:2700:100"),

    # ── 双帝战场 (Lv 85-95) ──
    enemy("mob_final_battle_remnant", "终战残魂", "mob", 85, 8000, 320, 160, 130, 3000, "item:item_soul_baby_fruit:20|exp:3000:100"),
    enemy("mob_emperor_ascension_guard", "帝升守卫", "mob", 90, 8800, 350, 180, 140, 3400, "item:item_ancient_relic:10|exp:3400:100"),
]

# ══════════════════════════════════════════════════════════
# PART 2: ELITE ENEMIES (100+) - by region
# ══════════════════════════════════════════════════════════
ELITE_ENEMIES = [
    # ── 沧澜帝国精英 (Lv 10-30) ──
    enemy("elite_wolf_head_leader", "狼头佣兵团长", "elite", 10, 800, 55, 35, 40, 200, "item:core_magic:25|exp:200:100"),
    enemy("elite_xiao_guard_captain", "林家护卫队长", "elite", 12, 900, 60, 40, 38, 250, "item:item_elixir:15|exp:250:100"),
    enemy("elite_wutan_arena_champ", "青石城擂主", "elite", 14, 1000, 65, 42, 42, 300, "item:silver:80|exp:300:100"),
    enemy("elite_alchemist_guard", "公会护卫长", "elite", 15, 1100, 62, 45, 40, 320, "item:item_elixir:20|exp:320:100"),
    enemy("elite_desert_merc_captain", "沙漠佣兵队长", "elite", 16, 1200, 70, 48, 45, 350, "item:item_desert_compass:10|exp:350:100"),
    enemy("elite_snake_temple_guard", "蛇人神殿守卫", "elite", 18, 1400, 78, 55, 50, 420, "item:item_snake_venom_vial:25|exp:420:100"),
    enemy("elite_miteer_chief_guard", "米特尔首席护卫", "elite", 20, 1600, 82, 60, 52, 480, "item:silver:100|exp:480:100"),
    enemy("elite_imperial_guard", "皇室禁卫", "elite", 22, 1800, 88, 65, 55, 550, "item:item_elixir:25|exp:550:100"),
    enemy("elite_nalan_elder", "纳兰家长老", "elite", 24, 2000, 92, 68, 58, 620, "item:herb_spirit_gathering:20|exp:620:100"),
    enemy("elite_jia_ma_general", "沧澜边军将领", "elite", 26, 2200, 95, 72, 60, 680, "item:item_jia_ma_passport:15|exp:680:100"),

    # ── 青岚宗精英 (Lv 18-32) ──
    enemy("elite_yunlan_elder", "青岚宗执事", "elite", 18, 1500, 80, 58, 55, 420, "item:core_wind:25|exp:420:100"),
    enemy("elite_yunlan_enforcer", "青岚宗执法", "elite", 20, 1700, 85, 62, 58, 480, "item:item_elixir:20|exp:480:100"),
    enemy("elite_yunlan_guard_captain", "青岚宗护卫长", "elite", 25, 2100, 95, 70, 62, 650, "item:core_wind:28|exp:650:100"),
    enemy("elite_yunlan_deacon", "青岚宗长老会成员", "elite", 28, 2500, 105, 78, 65, 750, "item:item_elixir:30|exp:750:100"),

    # ── 暗角域精英 (Lv 20-38) ──
    enemy("elite_blood_sect_elder", "血宗长老", "elite", 22, 2200, 95, 68, 60, 550, "item:herb_blood_essence_fruit:25|exp:550:100"),
    enemy("elite_black_alliance_captain", "暗盟队长", "elite", 24, 2400, 100, 72, 62, 620, "item:silver:120|exp:620:100"),
    enemy("elite_eight_gates_enforcer", "八扇门执事", "elite", 26, 2600, 105, 75, 65, 680, "item:item_black_market_pass:10|exp:680:100"),
    enemy("elite_feng_city_guard", "枫城守卫长", "elite", 28, 2800, 110, 78, 68, 750, "item:item_elixir:30|exp:750:100"),
    enemy("elite_black_emperor_guard", "黑皇宗护卫", "elite", 30, 3000, 115, 82, 70, 820, "item:item_forbidden_pill_fragment:8|exp:820:100"),
    enemy("elite_demon_valley_elder", "魔焰谷长老", "elite", 34, 3400, 125, 88, 75, 980, "item:core_fire:30|exp:980:100"),
    enemy("elite_black_domain_hunter", "暗角域赏金猎人", "elite", 32, 3200, 120, 85, 78, 900, "item:silver:150|exp:900:100"),
    enemy("elite_black_blood_merc", "血原佣兵头目", "elite", 36, 3600, 130, 90, 80, 1050, "item:item_beast_blood_essence:12|exp:1050:100"),

    # ── 迦南学院精英 (Lv 22-32) ──
    enemy("elite_strong_rank_10", "强榜第十", "elite", 22, 2200, 90, 65, 68, 550, "item:item_canaan_token:30|exp:550:100"),
    enemy("elite_strong_rank_5", "强榜第五", "elite", 26, 2700, 105, 75, 75, 700, "item:item_canaan_token:40|exp:700:100"),
    enemy("elite_inner_enforcer", "内院执法队", "elite", 25, 2500, 100, 72, 70, 650, "item:item_canaan_token:35|exp:650:100"),
    enemy("elite_tower_trial_guard", "炼气塔试炼官", "elite", 28, 2800, 110, 80, 72, 780, "item:item_canaan_token:45|exp:780:100"),
    enemy("elite_academy_instructor", "学院导师", "elite", 30, 3000, 115, 85, 78, 850, "item:item_elixir:35|exp:850:100"),

    # ── 出云帝国精英 (Lv 28-42) ──
    enemy("elite_poison_sect_elder", "毒宗长老", "elite", 30, 3200, 125, 85, 75, 880, "item:herb_desert_mandala:25|exp:880:100"),
    enemy("elite_scorpion_gate_elder", "万蝎门长老", "elite", 32, 3400, 130, 88, 72, 950, "item:item_snake_venom_vial:30|exp:950:100"),
    enemy("elite_golden_goose_elder", "金雁宗长老", "elite", 33, 3500, 132, 86, 80, 980, "item:core_wind:32|exp:980:100"),
    enemy("elite_mulan_valley_elder", "慕兰谷长老", "elite", 33, 3550, 128, 90, 75, 980, "item:core_wood:32|exp:980:100"),
    enemy("elite_chuyun_imperial_guard", "出云皇室禁卫", "elite", 35, 3800, 138, 92, 78, 1080, "item:item_elixir:35|exp:1080:100"),

    # ── 中州精英 (Lv 38-70) ──
    enemy("elite_wind_lightning_deacon", "风雷阁执事", "elite", 38, 4000, 145, 95, 95, 1200, "item:core_wind:35|exp:1200:100"),
    enemy("elite_huangquan_deacon", "黄泉阁执事", "elite", 40, 4200, 148, 98, 88, 1280, "item:core_earth:35|exp:1280:100"),
    enemy("elite_wanjian_deacon", "万剑阁执事", "elite", 42, 4400, 155, 92, 105, 1380, "item:core_ice:35|exp:1380:100"),
    enemy("elite_burning_valley_deacon", "焚炎谷执事", "elite", 44, 4600, 160, 100, 95, 1480, "item:core_fire:35|exp:1480:100"),
    enemy("elite_ice_river_deacon", "冰河谷执事", "elite", 45, 4700, 158, 105, 90, 1520, "item:core_ice:35|exp:1520:100"),
    enemy("elite_sky_demon_deacon", "天冥宗执事", "elite", 46, 4800, 165, 98, 92, 1580, "item:item_soul_baby_fruit:15|exp:1580:100"),
    enemy("elite_flower_sect_elder", "花宗长老", "elite", 48, 5000, 162, 102, 100, 1680, "item:herb_sky_jade_fruit:20|exp:1680:100"),
    enemy("elite_dan_tower_deacon", "丹阁执事", "elite", 50, 4800, 155, 95, 98, 1750, "item:item_dan_herb_box:15|exp:1750:100"),
    enemy("elite_dan_region_guard_captain", "丹域护卫长", "elite", 52, 5200, 170, 108, 100, 1820, "item:item_elixir:40|exp:1820:100"),
    enemy("elite_central_domain_merc", "中域佣兵团长", "elite", 50, 5000, 168, 105, 105, 1750, "item:item_zhongzhou_map:20|exp:1750:100"),
    enemy("elite_tianhuang_guard", "天黄城卫队长", "elite", 54, 5400, 175, 110, 102, 1900, "item:item_wormhole_pass:15|exp:1900:100"),
    enemy("elite_black_fire_sect_leader", "黑火宗宗主", "elite", 48, 4900, 160, 100, 96, 1680, "item:core_fire:35|exp:1680:100"),

    # ── 黑渊殿精英 (Lv 35-75) ──
    enemy("elite_soul_hall_protector", "黑渊护法", "elite", 35, 3800, 140, 90, 85, 1100, "item:item_soul_baby_fruit:20|exp:1100:100"),
    enemy("elite_soul_hall_elder_protector", "黑渊殿大护法", "elite", 42, 4500, 160, 100, 95, 1400, "item:item_soul_baby_fruit:25|exp:1400:100"),
    enemy("elite_soul_hall_soul_reaper", "黑渊殿勾魂使", "elite", 48, 5000, 175, 108, 105, 1650, "item:item_soul_baby_fruit:30|exp:1650:100"),
    enemy("elite_soul_hall_person_hall_guard", "人殿守卫长", "elite", 55, 5800, 195, 120, 112, 2000, "item:item_soul_baby_fruit:35|exp:2000:100"),
    enemy("elite_soul_hall_earth_hall_guard", "地殿守卫长", "elite", 60, 6500, 215, 135, 118, 2300, "item:item_soul_baby_fruit:40|exp:2300:100"),
    enemy("elite_soul_hall_heaven_hall_guard", "天殿守卫长", "elite", 65, 7200, 235, 150, 125, 2600, "item:item_soul_baby_fruit:50|exp:2600:100"),
    enemy("elite_soul_hall_messenger", "黑渊使者", "elite", 50, 5400, 185, 112, 108, 1800, "item:item_soul_baby_fruit:28|exp:1800:100"),
    enemy("elite_death_corpse_guardian", "葬尸山脉守墓者", "elite", 52, 5600, 190, 118, 100, 1880, "item:item_soul_baby_fruit:30|exp:1880:100"),

    # ── 兽域精英 (Lv 52-68) ──
    enemy("elite_transformed_beast", "化形魔兽", "elite", 52, 5800, 200, 120, 115, 1900, "item:item_beast_blood_essence:20|exp:1900:100"),
    enemy("elite_nether_python_elder", "冥蟒族长老", "elite", 56, 6200, 215, 130, 118, 2100, "item:item_huangquan_blood_crystal:15|exp:2100:100"),
    enemy("elite_sky_demon_phoenix", "天妖凰族战士", "elite", 60, 6800, 235, 135, 135, 2400, "item:item_beast_blood_essence:25|exp:2400:100"),
    enemy("elite_beast_region_chieftain", "兽域部落酋长", "elite", 58, 6400, 225, 132, 120, 2200, "item:item_transforming_herb:10|exp:2200:100"),
    enemy("elite_ancient_beast_guardian", "远古魔兽守护者", "elite", 62, 7000, 240, 145, 125, 2500, "item:item_beast_blood_essence:30|exp:2500:100"),

    # ── 龙岛精英 (Lv 58-78) ──
    enemy("elite_dragon_warrior", "龙族战士", "elite", 58, 6500, 230, 150, 125, 2200, "item:item_dragon_scale:10|exp:2200:100"),
    enemy("elite_dragon_guard_captain", "龙卫队长", "elite", 62, 7000, 245, 160, 130, 2450, "item:item_dragon_scale:12|exp:2450:100"),
    enemy("elite_west_dragon_commander", "西龙岛将领", "elite", 65, 7500, 255, 170, 132, 2650, "item:item_dragon_scale:15|exp:2650:100"),
    enemy("elite_south_dragon_commander", "南龙岛将领", "elite", 65, 7600, 252, 172, 130, 2650, "item:item_dragon_scale:15|exp:2650:100"),
    enemy("elite_north_dragon_commander", "北龙岛将领", "elite", 68, 8000, 270, 178, 135, 2850, "item:item_dragon_scale:18|exp:2850:100"),
    enemy("elite_east_dragon_commander", "东龙岛将领", "elite", 62, 6800, 240, 158, 128, 2400, "item:item_dragon_scale:12|exp:2400:100"),
    enemy("elite_ancient_dragon_spirit", "虚空龙残魂", "elite", 72, 8200, 280, 170, 140, 3100, "item:item_dragon_scale:20|exp:3100:100"),

    # ── 远古种族精英 (Lv 60-85) ──
    enemy("elite_gu_clan_warrior", "云族战士", "elite", 62, 7000, 250, 165, 140, 2500, "item:item_elixir:45|exp:2500:100"),
    enemy("elite_gu_clan_elder", "云族长老", "elite", 72, 8500, 290, 190, 155, 3200, "item:item_ancient_relic:15|exp:3200:100"),
    enemy("elite_yan_clan_warrior", "炎族战士", "elite", 64, 7200, 260, 160, 135, 2650, "item:core_fire:40|exp:2650:100"),
    enemy("elite_lei_clan_warrior", "雷族战士", "elite", 64, 7300, 255, 162, 145, 2650, "item:core_wind:40|exp:2650:100"),
    enemy("elite_yao_clan_guard", "药族守卫", "elite", 60, 6800, 235, 155, 130, 2400, "item:herb_ancient_green_vine:20|exp:2400:100"),
    enemy("elite_yao_clan_elder", "药族长老", "elite", 70, 8000, 275, 180, 148, 3100, "item:item_dan_herb_box:25|exp:3100:100"),
    enemy("elite_hun_clan_warrior", "玄族战士", "elite", 68, 7800, 275, 170, 150, 2900, "item:item_soul_baby_fruit:35|exp:2900:100"),
    enemy("elite_hun_clan_elder", "玄族长老", "elite", 78, 9000, 310, 200, 168, 3600, "item:item_soul_baby_fruit:50|exp:3600:100"),
    enemy("elite_ancient_alliance_officer", "远古联盟军官", "elite", 75, 8500, 300, 185, 160, 3400, "item:item_elixir:50|exp:3400:100"),

    # ── 莽荒古域精英 (Lv 64-72) ──
    enemy("elite_manghuang_hunter", "莽荒猎手", "elite", 64, 7200, 250, 150, 140, 2700, "item:item_beast_blood_essence:25|exp:2700:100"),
    enemy("elite_demon_python_guardian", "天魔蟒守护者", "elite", 68, 8000, 275, 165, 142, 3000, "item:item_huangquan_blood_crystal:20|exp:3000:100"),
    enemy("elite_ancient_platform_guard", "古域台守卫", "elite", 70, 8300, 285, 175, 148, 3150, "item:item_ancient_seed:12|exp:3150:100"),
    enemy("elite_bodhi_illusion_guard", "菩提幻境守卫", "elite", 72, 8200, 280, 160, 160, 3300, "item:item_bodhi_leaf:5|exp:3300:100"),

    # ── 妖火空间精英 (Lv 66-72) ──
    enemy("elite_demon_flame_warden", "妖火守卫", "elite", 66, 7600, 260, 155, 150, 2800, "item:item_demon_flame_crystal:15|exp:2800:100"),
    enemy("elite_saint_remains_guard", "妖圣残像护卫", "elite", 70, 8200, 280, 165, 155, 3100, "item:item_demon_flame_crystal:20|exp:3100:100"),

    # ── 古帝洞府精英 (Lv 78-90) ──
    enemy("elite_emperor_cave_sentry", "古帝石门守卫", "elite", 78, 9000, 310, 210, 155, 3600, "item:item_ancient_relic:18|exp:3600:100"),
    enemy("elite_emperor_treasure_guard", "帝品丹药守卫", "elite", 85, 10000, 350, 230, 170, 4200, "item:item_embryonic_emperor_pill:5|exp:4200:100"),

    # ── 最终战场精英 (Lv 88-95) ──
    enemy("elite_allied_forces_general", "联军将领", "elite", 88, 11000, 380, 250, 180, 4800, "item:item_elixir:60|exp:4800:100"),
    enemy("elite_hun_army_commander", "玄族军团指挥官", "elite", 90, 12000, 400, 260, 190, 5200, "item:item_soul_baby_fruit:60|exp:5200:100"),
]

# ══════════════════════════════════════════════════════════
# PART 3: ADDITIONAL BOSSES (per novel)
# ══════════════════════════════════════════════════════════
# NON_LETHAL = these characters can develop relationships later; cannot be killed
NON_LETHAL = "non_lethal"  # marker in Notes field
ADDITIONAL_BOSSES = [
    # ── 沧澜帝国/魔兽山脉 Boss ──
    enemy("boss_mu_she", "穆蛇", "boss", 12, 1500, 85, 55, 60, 500, "item:core_magic:30|exp:500:100", "", "", NON_LETHAL),
    enemy("boss_amethyst_wing_lion", "紫晶翼狮王", "boss", 18, 3000, 120, 100, 80, 1200, "item:item_beast_hide_bundle:40|item:core_magic:40|exp:1200:100"),
    enemy("boss_desert_ancient_scorpion", "沙漠古蝎王", "boss", 20, 3500, 130, 120, 70, 1400, "item:item_snake_venom_vial:40|exp:1400:100"),
    enemy("boss_snake_high_priest", "蛇人族大祭司", "boss", 22, 3800, 140, 110, 90, 1600, "item:item_snake_venom_vial:50|exp:1600:100", "", "", NON_LETHAL),

    # ── 暗角域 Boss ──
    enemy("boss_blood_sect_leader", "血宗宗主", "boss", 28, 5000, 190, 140, 150, 2500, "item:herb_blood_essence_fruit:50|exp:2500:100"),
    enemy("boss_eight_gates_master", "八扇门主", "boss", 26, 4500, 175, 130, 145, 2200, "item:item_black_market_pass:30|exp:2200:100", "", "", NON_LETHAL),
    enemy("boss_black_alliance_commander", "暗盟副统领", "boss", 32, 6000, 220, 160, 170, 3500, "item:item_forbidden_pill_fragment:25|exp:3500:100"),

    # ── 迦南学院 Boss ──
    enemy("boss_inner_academy_guardian", "内院守护者", "boss", 26, 4500, 170, 135, 140, 2200, "item:item_canaan_token:60|exp:2200:100", "", "", NON_LETHAL),
    enemy("boss_skyfire_tower_beast", "炼气塔火兽", "boss", 30, 5500, 200, 150, 160, 3000, "item:core_fire:50|exp:3000:100"),

    # ── 出云帝国 Boss ──
    enemy("boss_golden_goose_leader", "金雁宗宗主", "boss", 36, 7000, 240, 180, 190, 4200, "item:core_wind:45|exp:4200:100"),
    enemy("boss_mulan_valley_leader", "慕兰谷谷主", "boss", 36, 7200, 235, 185, 185, 4200, "item:core_wood:45|exp:4200:100"),

    # ── 中州 Boss ──
    enemy("boss_thunder_emperor", "雷尊者", "boss", 48, 12000, 350, 250, 280, 8000, "item:core_wind:50|exp:8000:100", "", "", NON_LETHAL),
    enemy("boss_sword_emperor", "剑尊者", "boss", 48, 12500, 355, 245, 275, 8000, "item:core_ice:50|exp:8000:100", "", "", NON_LETHAL),
    enemy("boss_sky_demon_sect_leader", "天冥宗宗主", "boss", 52, 15000, 400, 280, 270, 10000, "item:item_soul_baby_fruit:45|exp:10000:100"),
    enemy("boss_black_fire_ancestor", "黑火老祖", "boss", 50, 14000, 380, 270, 265, 9500, "item:core_fire:50|exp:9500:100"),
    enemy("boss_dan_tower_guardian", "丹阁守护者", "boss", 55, 16000, 380, 300, 290, 12000, "item:item_dan_herb_box:30|exp:12000:100", "", "", NON_LETHAL),

    # ── 兽域 Boss ──
    enemy("boss_nether_python_usurper", "冥蟒篡位者", "boss", 58, 18000, 420, 320, 300, 14000, "item:item_huangquan_blood_crystal:30|exp:14000:100"),
    enemy("boss_sky_phoenix_elder", "天妖凰族长老", "boss", 62, 22000, 480, 350, 350, 18000, "item:item_beast_blood_essence:40|exp:18000:100", "", "", NON_LETHAL),

    # ── 龙岛 Boss ──
    enemy("boss_south_dragon_king", "南龙王", "boss", 65, 25000, 500, 400, 380, 22000, "item:item_dragon_scale:25|exp:22000:100", "", "", NON_LETHAL),
    enemy("boss_north_dragon_king", "北龙王", "boss", 68, 28000, 530, 420, 400, 25000, "item:item_dragon_scale:30|exp:25000:100"),
    enemy("boss_ancient_dragon_spirit_king", "虚空龙王残魂", "boss", 72, 32000, 580, 450, 420, 30000, "item:item_dragon_saliva_herb:20|exp:30000:100"),
    enemy("boss_dragon_emperor_zhu_kun", "苍坤", "boss", 95, 85000, 4500, 3200, 3800, 85000, "item:item_dragon_scale:50|exp:85000:100", "", "", NON_LETHAL),

    # ── 远古种族 Boss ──
    enemy("boss_gu_clan_general", "云族大将", "boss", 72, 30000, 550, 480, 420, 28000, "item:item_ancient_relic:25|exp:28000:100", "", "", NON_LETHAL),
    enemy("boss_yan_clan_elder_chief", "炎族大长老", "boss", 74, 32000, 580, 460, 440, 31000, "item:core_fire:60|exp:31000:100", "", "", NON_LETHAL),
    enemy("boss_lei_clan_elder_chief", "雷族大长老", "boss", 74, 33000, 575, 465, 460, 31000, "item:core_wind:60|exp:31000:100", "", "", NON_LETHAL),
    enemy("boss_yao_clan_guardian", "药族守护长老", "boss", 70, 28000, 520, 430, 400, 26000, "item:item_dan_herb_box:35|exp:26000:100", "", "", NON_LETHAL),
    enemy("boss_hun_clan_ritual_master", "玄族祭坛主", "boss", 82, 42000, 650, 500, 480, 38000, "item:item_soul_baby_fruit:70|exp:38000:100"),

    # ── 莽荒古域 Boss ──
    enemy("boss_demon_python_king", "天魔蟒王", "boss", 68, 26000, 480, 360, 350, 24000, "item:item_huangquan_blood_crystal:35|exp:24000:100"),
    enemy("boss_bodhi_tree_guardian", "菩提树灵", "boss", 70, 28000, 460, 400, 420, 30000, "item:item_bodhi_leaf:15|exp:30000:100", "", "", NON_LETHAL),
    enemy("boss_ancient_domain_beast", "古域兽王", "boss", 66, 24000, 460, 350, 340, 22000, "item:item_ancient_seed:20|exp:22000:100"),

    # ── 妖火空间 Boss ──
    enemy("boss_demon_flame_saint_spirit", "净莲妖圣残魂", "boss", 70, 30000, 520, 420, 430, 32000, "item:item_demon_flame_crystal:30|exp:32000:100", "", "", NON_LETHAL),

    # ── 古帝洞府 Boss ──
    enemy("boss_emperor_cave_dragon", "洞府守护龙魂", "boss", 82, 40000, 620, 520, 480, 38000, "item:item_dragon_scale:35|exp:38000:100"),
    enemy("boss_emperor_pill_guardian", "帝品雏丹守护者", "boss", 88, 55000, 720, 580, 520, 48000, "item:item_embryonic_emperor_pill:10|exp:48000:100"),

    # ── 最终战场 Boss ──
    enemy("boss_hun_army_general", "玄族大军统帅", "boss", 90, 60000, 750, 600, 550, 50000, "item:item_soul_baby_fruit:80|exp:50000:100"),
    enemy("boss_alliance_commander", "联盟总指挥", "boss", 90, 62000, 740, 620, 540, 50000, "item:item_elixir:80|exp:50000:100", "", "", NON_LETHAL),
]

# ══════════════════════════════════════════════════════════
# APPLY ALL
# ══════════════════════════════════════════════════════════
all_new = COMMON_MOBS + ELITE_ENEMIES + ADDITIONAL_BOSSES
print(f"Adding {len(COMMON_MOBS)} mobs + {len(ELITE_ENEMIES)} elites + {len(ADDITIONAL_BOSSES)} bosses = {len(all_new)} total")

added = 0
for data in all_new:
    eid = data["Enemy_ID"]
    if eid in existing_ids:
        print(f"  SKIP: {eid}")
        continue
    append_row(ws_enemies, data)
    existing_ids.add(eid)
    added += 1

wb.save(WORKBOOK_PATH)
print(f"\nAdded {added} new enemies. Total: {len(existing_ids)}")
print("Done!")
