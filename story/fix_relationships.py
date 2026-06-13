"""
Adjust relationship initial values per user requirements:
- Hostile: keep as is
- Friendly: default to 0 (need to cultivate)
- Xiao Zhan, Xiao Ding, Xiao Li, Xiao Xun'er, Yao Lao: default 100 (immune)
"""
from pathlib import Path
import openpyxl

WORKBOOK_PATH = Path("story/text_game_event_schema_v4.xlsx")

wb = openpyxl.load_workbook(WORKBOOK_PATH)

def find_sheet(name_prefix):
    for name in wb.sheetnames:
        if name.startswith(name_prefix):
            return wb[name]
    raise KeyError(f"Sheet starting with '{name_prefix}' not found")

def append_row(ws, data: dict):
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    row_data = []
    for h in headers:
        val = data.get(h, "")
        row_data.append(val if val is not None else "")
    ws.append(row_data)

ws_rels = find_sheet("Relationships_")

# Find the header row and column indices
headers = [cell.value for cell in next(ws_rels.iter_rows(min_row=1, max_row=1))]
print(f"Headers: {headers}")
# Find indices
id_col = headers.index("Relation_ID")
initial_col = headers.index("Initial_Value")
min_col = headers.index("Min_Value")
max_col = headers.index("Max_Value")
stage_col = headers.index("Stage_Rule")
on_reach_col = headers.index("On_Reach_Effect")
pre_cond_col = headers.index("Pre_Condition")
src_col = headers.index("Source_ID")
tgt_col = headers.index("Target_ID")
type_col = headers.index("Relation_Type")
bidir_col = headers.index("Is_Bidirectional")
vis_col = headers.index("Visible_To_Player")

# Changes to make:
# 1. Set 5 immune characters to initial=100, max=100 (can't be changed)
# 2. Set all other friendly relationships to initial=0
# 3. Add new relationships for xiao_ding and xiao_li
IMMUNE_IDS = {
    "rel_player_xiao_zhan", "rel_player_xun_er", "rel_player_xuanlu_elder",
    "rel_player_xiao_ding", "rel_player_xiao_li",
}
FRIENDLY_TO_ZERO = {
    "rel_player_yun_yun", "rel_player_xiaoyixian",
    "rel_player_ziyan", "rel_player_qinglin",
    "rel_player_su_qian", "rel_player_feng_zunzhe",
}

# Scan all rows and update values
changes_made = 0
for row_idx, row in enumerate(ws_rels.iter_rows(min_row=2), start=2):
    rel_id = row[id_col].value
    if not rel_id:
        continue

    if rel_id in IMMUNE_IDS:
        old_val = row[initial_col].value
        ws_rels.cell(row=row_idx, column=initial_col+1).value = 100
        ws_rels.cell(row=row_idx, column=max_col+1).value = 100
        ws_rels.cell(row=row_idx, column=min_col+1).value = 0
        print(f"  IMMUNE: {rel_id} initial {old_val} -> 100, max -> 100")
        changes_made += 1
    elif rel_id in FRIENDLY_TO_ZERO:
        old_val = row[initial_col].value
        ws_rels.cell(row=row_idx, column=initial_col+1).value = 0
        print(f"  ZERO: {rel_id} initial {old_val} -> 0")
        changes_made += 1

# Add new relationships for Xiao Ding and Xiao Li
existing_ids = set()
for row in ws_rels.iter_rows(min_row=2):
    val = row[id_col].value
    if val:
        existing_ids.add(str(val))

new_rels = [
    {
        "Relation_ID": "rel_player_xiao_ding",
        "Source_ID": "player",
        "Target_ID": "npc_xiao_ding",
        "Relation_Type": "family",
        "Initial_Value": 100,
        "Min_Value": 0,
        "Max_Value": 100,
        "Stage_Rule": "",
        "Is_Bidirectional": "false",
        "Visible_To_Player": "true",
        "Pre_Condition": "",
        "On_Reach_Effect": "",
        "Notes": "萧鼎-二哥，免疫事件影响",
    },
    {
        "Relation_ID": "rel_player_xiao_li",
        "Source_ID": "player",
        "Target_ID": "npc_xiao_li",
        "Relation_Type": "family",
        "Initial_Value": 100,
        "Min_Value": 0,
        "Max_Value": 100,
        "Stage_Rule": "",
        "Is_Bidirectional": "false",
        "Visible_To_Player": "true",
        "Pre_Condition": "",
        "On_Reach_Effect": "",
        "Notes": "萧厉-三哥，免疫事件影响",
    },
]

for rel_data in new_rels:
    if rel_data["Relation_ID"] in existing_ids:
        print(f"  SKIP (exists): {rel_data['Relation_ID']}")
        continue
    append_row(ws_rels, rel_data)
    existing_ids.add(rel_data["Relation_ID"])
    print(f"  + ADDED: {rel_data['Relation_ID']}")

wb.save(WORKBOOK_PATH)
print(f"\nSaved. {changes_made} changes made, {len(new_rels)} new relationships added.")
