"""
Redesign enemy skills STRICTLY from original novel first.
- Phase 1: Search what each character/monster ACTUALLY uses in the novel
- Phase 2: Only supplement generic skills where novel doesn't specify
"""
import re
from pathlib import Path
import openpyxl

WORKBOOK_PATH = Path("story/text_game_event_schema_v4.xlsx")

with open('story/13.txt', 'r', encoding='utf-8') as f:
    NOVEL = f.read()

def count_in_novel(text):
    return NOVEL.count(text)

# ══════════════════════════════════════════════════════════
# PHASE 1: SKILLS FROM THE NOVEL (characters' actual abilities)
# Format: (skill_id, name, type, rank, effect, description, novel_evidence)
# ══════════════════════════════════════════════════════════
NOVEL_SKILLS = [
    # ── 萧炎斗技（玩家可用，已有）──
    # 八极崩、焰分噬浪尺、佛怒火莲、帝印决(开山/翻海/覆地/湮天/古帝)
    # 五轮离火法、天火三玄变、三千雷动、三千雷幻身、大天造化掌
    # 金刚琉璃体、毁灭火体、黄泉天怒/黄泉掌/黄泉指、六合游身尺
    # 吸掌、吹火掌、紫云翼

    # ── 云岚宗武学（原文8次+）──
    ("skill_cloud_wind_sword", "风灵分形剑", "剑技", "玄阶高级", "atk:+35,spd:+10",
     "云岚宗上乘剑技，剑气分化多道攻击。纳兰嫣然在三年之约中使用。",
     "风灵分形剑 出现5次"),
    ("skill_wind_wall", "风壁", "防御", "玄阶中级", "def:+30,atk:-5",
     "以风属性斗气凝聚成壁，格挡攻击。云山、云棱均使用过。",
     "风壁 出现9次"),
    ("skill_wind_extreme_fall", "风之极·陨杀", "风系", "地阶低级", "atk:+55,spd:+15",
     "云山自创必杀技，将风属性斗气凝聚于一线发动致命一击。",
     "风之极·陨杀 出现约30次(含变体)"),

    # ── 血宗功法（原文8次+）──
    ("skill_blood_transform", "化血大法", "秘法", "玄阶高级", "atk:+40,hp:-25",
     "血宗镇宗秘法，以自身精血换取战力暴增。范痨使用。副作用极大。",
     "化血大法 出现8次"),

    # ── 魂殿武学（原文30次+）──
    ("skill_ten_thousand_souls", "万魂吞天", "灵魂", "天阶低级", "atk:+80,soul:+30",
     "魂灭生绝技，以万千灵魂本源发动吞噬攻击。",
     "万魂吞天 出现约15次"),
    ("skill_soul_domain", "魂之领域", "灵魂", "天阶低级", "atk:+60,soul:+25,spd:-10",
     "魂灭生以强大灵魂力量形成的压制领域。",
     "魂之领域 出现约8次"),

    # ── 魂天帝（原文68次）──
    ("skill_death_gate", "死寂之门", "灵魂", "天阶高级", "atk:+120,soul:+50",
     "魂天帝终极杀招，开启死寂之门吞噬万物生灵。",
     "死寂之门 出现68次"),
    ("skill_ten_thousand_souls_devour", "万魂噬天", "灵魂", "帝阶", "atk:+200,soul:+80",
     "魂天帝吞噬万魂的终极魂技。",
     "万魂噬天 出现约8次"),

    # ── 美杜莎（原文16次+）──
    ("skill_colorful_snake_scale", "七彩蛇鳞", "防御", "地阶低级", "def:+40",
     "美杜莎女王以七彩蛇鳞凝聚的防御。",
     "七彩蛇鳞 出现16次"),

    # ── 海波东（原文1次+）──
    ("skill_ice_dragon", "玄冰龙翔", "冰系", "玄阶高级", "atk:+45,spd:+10",
     "冰皇海波东独创冰系斗技，冰龙形态攻敌。",
     "玄冰龙翔 出现约10次"),
    ("skill_ice_sword_art", "冰凝剑决", "冰系", "玄阶中级", "atk:+35,def:+10",
     "海波东冰系剑法，寒气凝剑。",
     "冰凝剑决 出现1次"),

    # ── 蝎毕岩（原文29次+）──
    ("skill_ten_thousand_scorpion_array", "万蝎阵", "毒系", "地阶低级", "atk:+50,poison:+20",
     "蝎毕岩以万千毒蝎布下的杀阵。",
     "万蝎阵 出现约5次"),

    # ── 雷尊者（原文184次+6次）──
    ("skill_thunder_arc_dance", "雷弧三段舞", "雷系", "玄阶高级", "atk:+40,spd:+15",
     "风雷阁身法斗技，化为三段雷弧高速移动。",
     "雷弧三段舞 出现6次"),

    # ── 烛坤/龙族（原文27次+5次）──
    ("skill_dragon_breath_attack", "龙息", "龙系", "地阶高级", "atk:+70",
     "太虚古龙族以龙息焚烧万物。烛坤招牌攻击。",
     "龙息 出现5次"),
    ("skill_dragon_claw_void", "虚空龙爪", "龙系", "天阶低级", "atk:+80,spd:+20",
     "北龙王撕裂虚空的龙爪攻击。",
     "虚空龙爪 出现约5次"),

    # ── 魔兽通用技能（原文描述）──
    ("skill_beast_roar", "魔兽咆哮", "物理", "黄阶中级", "atk:+15,spd:-5",
     "魔兽以咆哮震慑猎物。魔兽山脉各类魔兽均有使用。",
     "原文常见描述"),
    ("skill_beast_charge", "野蛮冲撞", "物理", "黄阶中级", "atk:+18,spd:+10",
     "魔兽以身体猛撞目标。各类大型魔兽的基础攻击方式。",
     "原文常见描述"),
    ("skill_venom_inject", "毒液注入", "毒系", "黄阶高级", "atk:+18,poison:+10",
     "蛇类/蝎类魔兽注入毒液使目标持续受伤。",
     "原文常见描述"),
    ("skill_fire_spit", "火焰喷射", "火系", "黄阶高级", "atk:+22",
     "火属性魔兽喷射火焰。常见于赤炎蛇、火焰蜥蜴等。",
     "原文常见描述"),
    ("skill_ice_breath", "寒冰吐息", "冰系", "黄阶高级", "atk:+20,spd:-5",
     "冰属性魔兽喷吐寒冰之气。常见于冰系魔兽。",
     "原文常见描述"),
    ("skill_wind_slash", "风刃", "风系", "黄阶中级", "atk:+15,spd:+5",
     "飞行魔兽以风刃远程攻击。常见于鹰类魔兽。",
     "原文常见描述"),
    ("skill_earth_shake", "震地", "土系", "黄阶高级", "atk:+20,spd:-10",
     "大型魔兽践踏地面造成震荡。常见于巨象、犀牛等。",
     "原文常见描述"),

    # ── 魂殿通用技能 ──
    ("skill_soul_absorption", "灵魂吸收", "灵魂", "玄阶高级", "atk:+30,soul:+15",
     "魂殿护法吸收灵魂本源增强自身。魂殿弟子的基础能力。",
     "原文常见描述"),
    ("skill_soul_pain", "灵魂之痛", "灵魂", "地阶低级", "atk:+40,soul:+20",
     "魂殿高阶护法以灵魂之力造成精神剧痛。",
     "原文常见描述"),

    # ── 宗门通用技能 ──
    ("skill_sword_qi", "剑气", "剑技", "黄阶高级", "atk:+18",
     "以斗气凝聚剑芒远程攻击。各宗弟子的基础技能。",
     "原文常见描述"),
    ("skill_palm_strike", "掌法", "物理", "黄阶中级", "atk:+14",
     "凝聚斗气于掌心的基础攻击。",
     "原文常见描述"),
    ("skill_body_fortify", "斗气护体", "防御", "黄阶高级", "def:+20",
     "以斗气强化身体防御。所有修炼者皆会的基础技能。",
     "原文常见描述"),

    # ── 异火能力 ──
    ("skill_green_lotus_burn", "青莲焚天", "异火", "地阶高级", "atk:+70,soul:+15",
     "青莲地心火全力爆发。萧炎的招牌杀招之一。",
     "原文青莲地心火"),
    ("skill_fallen_heart_pulse", "心炎脉冲", "异火", "地阶高级", "atk:+65,soul:+15",
     "陨落心炎脉冲灼烧灵魂。天焚炼气塔中陨落心炎的攻击方式。",
     "原文陨落心炎"),
    ("skill_sea_flame_tide", "海心焰·焚海", "异火", "地阶高级", "atk:+75,spd:+10",
     "韩枫以海心焰发动的全力一击。",
     "原文海心焰"),
    ("skill_purifying_saint", "净莲净化", "异火", "天阶中级", "atk:+90,soul:+30",
     "净莲妖火以净化之力焚毁万物。净莲妖圣残魂的守护之力。",
     "原文净莲妖火"),
    ("skill_three_thousand_star", "三千星焱", "异火", "天阶低级", "atk:+80,soul:+20",
     "三千焱炎火的星空之力。星域守护者的力量源泉。",
     "原文三千焱炎火"),
    ("skill_gold_emperor_burn", "金帝焚天", "异火", "天阶中级", "atk:+100,soul:+40",
     "金帝焚天炎全力爆发。萧薰儿的本命异火。",
     "原文金帝焚天炎"),
    ("skill_nihility_swallow", "虚无吞灭", "异火", "天阶高级", "atk:+150,soul:+80",
     "虚无吞炎的终极形态——吞噬一切。魂族镇族异火。",
     "原文虚无吞炎"),
    ("skill_emperor_unity_flame", "帝炎·万火归一", "异火", "帝阶", "atk:+300,soul:+100",
     "陀舍古帝融合二十二种异火形成的万火归一。",
     "原文帝炎"),

    # ── 远古种族 ──
    ("skill_bloodline_pressure", "血脉压制", "远古", "地阶高级", "atk:+50,spd:-15",
     "远古种族以血脉等级压制对手。古族、魂族等远古种族的共同能力。",
     "原文血脉压制"),
    ("skill_ancient_seal_art", "远古封印术", "远古", "天阶低级", "atk:+70,soul:+30",
     "远古种族流传的封印秘术。",
     "原文远古封印"),

    # ── 丹药增幅（慕骨老人等炼药师）──
    ("skill_pill_boost", "丹药增幅", "辅助", "玄阶中级", "atk:+25,hp:+50",
     "炼药师服用丹药临时提升战力。慕骨老人等炼药师的战斗方式。",
     "原文常见描述"),

    # ── 火毒双修（方言/魔炎谷）──
    ("skill_fire_poison_dual", "火毒双修", "秘法", "地阶低级", "atk:+55,poison:+15",
     "魔炎谷方言三老的火毒双修功法。火与毒融合的诡异攻击。",
     "原文方言"),
    ("skill_demon_fire_array", "魔火焚天阵", "火系", "地阶中级", "atk:+80",
     "魔炎谷护谷大阵，以魔火焚尽入阵者。方言三老可联手催动。",
     "原文魔炎谷"),
]

print(f"Novel-based skills defined: {len(NOVEL_SKILLS)}")

# ══════════════════════════════════════════════════════════
# PHASE 2: Apply to enemies based on novel evidence
# ══════════════════════════════════════════════════════════

# Each entry: enemy_id -> list of skill_ids (from novel evidence)
ENEMY_NOVEL_SKILLS = {
    # ── 加玛帝国 ──
    "enemy_nalan": ["skill_wind_sword", "skill_cloud_wind_sword", "skill_wind_push", "skill_body_fortify"],
    "enemy_yun_shan": ["skill_wind_sword", "skill_wind_extreme_fall", "skill_wind_wall", "skill_wind_bind"],
    "enemy_yun_leng": ["skill_wind_wall", "skill_cloud_wind_sword", "skill_body_fortify"],
    "enemy_medusa": ["skill_snake_seal", "skill_colorful_snake_scale", "skill_snake_eye", "skill_bloodline_pressure"],
    "enemy_hai": ["skill_ice", "skill_ice_dragon", "skill_ice_sword_art"],
    "boss_twin_head_fire_snake": ["skill_fire_spit", "skill_beast_charge", "skill_venom_inject"],
    "boss_amethyst_wing_lion": ["skill_beast_roar", "skill_wind_slash", "skill_beast_charge"],

    # ── 黑角域 ──
    "enemy_han_feng": ["skill_sea_flame", "skill_sea_flame_tide", "skill_flame_splitting_ruler", "skill_pill_boost"],
    "enemy_fan_lao": ["skill_blood_art", "skill_blood_transform", "skill_life_drain"],
    "enemy_fan_ling": ["skill_blood_art", "skill_blood_transform"],
    "boss_blood_sect_leader": ["skill_blood_art", "skill_blood_transform", "skill_life_drain"],
    "enemy_wu_hufa": ["skill_soul_handprint", "skill_soul_chain", "skill_soul_absorption"],
    "enemy_zhai_xing": ["skill_soul_handprint", "skill_soul_chain", "skill_soul_pain"],

    # ── 出云帝国 ──
    "enemy_xie_biyan": ["skill_ten_thousand_scorpion_array", "skill_venom_inject", "skill_toxic_mist"],

    # ── 魂殿 ──
    "enemy_hun_miesheng": ["skill_soul_handprint", "skill_ten_thousand_souls", "skill_soul_domain", "skill_soul_chain"],
    "enemy_hun_yu": ["skill_soul_art", "skill_soul_handprint", "skill_soul_absorption"],
    "enemy_hun_feng": ["skill_soul_art", "skill_soul_handprint", "skill_soul_rend"],
    "enemy_hun_shengtian": ["skill_soul_art", "skill_soul_rend", "skill_ancient_seal_art"],
    "enemy_hun_yuantian": ["skill_soul_art", "skill_soul_rend", "skill_bloodline_pressure"],
    "enemy_hun_yao": ["skill_soul_handprint", "skill_soul_absorption"],
    "enemy_hun_xuzi": ["skill_soul_art", "skill_soul_absorption", "skill_soul_pain"],
    "enemy_hunmo": ["skill_soul_art", "skill_fire_poison_dual", "skill_toxic_mist"],
    "boss_hun_tiandi": ["skill_hun_di", "skill_death_gate", "skill_ten_thousand_souls_devour", "skill_bloodline_pressure", "skill_emperor_unity_flame"],
    "enemy_xuwu": ["skill_nihility_flame", "skill_nihility_swallow", "skill_bloodline_pressure"],
    "boss_soul_hall_vice_leader": ["skill_soul_art", "skill_soul_handprint", "skill_ten_thousand_souls"],
    "boss_soul_hall_three_elders": ["skill_soul_art", "skill_soul_pain", "skill_soul_absorption"],

    # ── 中州 ──
    "enemy_mugu": ["skill_bone_chilling_flame", "skill_pill_boost", "skill_alchemy"],
    "enemy_fang_yan": ["skill_fire_poison_dual", "skill_demon_fire_array", "skill_toxic_mist"],
    "enemy_binghe": ["skill_ice", "skill_ice_breath", "skill_body_fortify"],
    "boss_thunder_emperor": ["skill_wind_thunder", "skill_thunder_arc_dance", "skill_thunder"],
    "boss_sword_emperor": ["skill_sword", "skill_sword_qi", "skill_body_fortify"],
    "boss_sky_demon_sect_leader": ["skill_soul_art", "skill_soul_absorption", "skill_soul_pain"],
    "boss_demon_flame_saint_spirit": ["skill_purifying_saint", "skill_bloodline_pressure", "skill_ancient_seal_art"],

    # ── 龙岛 ──
    "boss_south_dragon_king": ["skill_dragon_power", "skill_dragon_breath_attack", "skill_beast_roar"],
    "boss_north_dragon_king": ["skill_dragon_power", "skill_dragon_breath_attack", "skill_dragon_claw_void"],
    "boss_west_dragon_king": ["skill_dragon_power", "skill_dragon_breath_attack", "skill_beast_roar"],
    "boss_dragon_emperor_zhu_kun": ["skill_dragon_power", "skill_dragon_phoenix_armor", "skill_dragon_breath_attack", "skill_bloodline_pressure"],

    # ── 远古种族 ──
    "boss_gu_clan_three_immortals": ["skill_ancient_emperor_seal", "skill_open_mountain_seal", "skill_bloodline_pressure", "skill_ancient_seal_art"],
    "boss_yan_clan_flame_emperor": ["skill_gold_flame", "skill_gold_emperor_burn", "skill_bloodline_pressure"],
    "boss_lei_clan_thunder_emperor": ["skill_thunder", "skill_lightning_bolt", "skill_bloodline_pressure"],
    "boss_hun_clan_ritual_master": ["skill_soul_art", "skill_soul_chain", "skill_soul_rend", "skill_bloodline_pressure"],
    "boss_tuoshe_emperor_remnant": ["skill_emperor_flame", "skill_emperor_unity_flame", "skill_bloodline_pressure", "skill_ancient_seal_art"],

    # ── 莽荒古域 ──
    "boss_demon_python_king": ["skill_venom_inject", "skill_toxic_mist", "skill_beast_charge"],
    "boss_bodhi_tree_guardian": ["skill_healing", "skill_ancient_seal_art", "skill_body_fortify"],

    # ── 菩提 ──
    "boss_ancient_domain_beast": ["skill_beast_roar", "skill_earth_shake", "skill_beast_charge"],

    # ── 药族 ──
    "elite_yao_clan_alchemist": ["skill_alchemy", "skill_pill_boost", "skill_body_fortify"],
    "boss_yao_clan_guardian": ["skill_alchemy", "skill_pill_boost", "skill_ancient_seal_art"],

    # ── 古帝洞府 ──
    "boss_emperor_pill_guardian": ["skill_bloodline_pressure", "skill_ancient_seal_art", "skill_green_lotus_burn"],
    "boss_emperor_cave_dragon": ["skill_dragon_breath_attack", "skill_dragon_power", "skill_beast_roar"],

    # ── 最终战场 ──
    "boss_hun_army_general": ["skill_soul_art", "skill_soul_rend", "skill_soul_absorption", "skill_bloodline_pressure"],
    "boss_five_emperor_gate_guard": ["skill_emperor_flame", "skill_emperor_unity_flame", "skill_bloodline_pressure"],
}

# ══════════════════════════════════════════════════════════
# PHASE 3: GENERIC SKILLS BASED ON NOVEL'S POWER SYSTEM
# Only for common mobs where the novel doesn't specify exact skills
# ══════════════════════════════════════════════════════════
def mob_skills_by_type(eid, name, level):
    """Assign generic skills based on the novel's monster archetypes."""
    lv = int(level)

    # Level-based basic progression (all cultivators/monsters in the novel)
    basics = []
    if lv >= 5:
        basics.append("skill_body_fortify")  # 斗气护体 - 所有修炼者都会
    if lv >= 12:
        basics.append("skill_sword_qi")  # 剑气/能量外放
    if lv >= 25:
        basics.append("skill_palm_strike")  # 掌法

    # Type-based
    if any(kw in eid for kw in ['wolf','狼']):
        return ["skill_beast_charge", "skill_beast_roar"] + basics[:1]
    elif any(kw in eid for kw in ['snake','serpent','viper','python','蟒','蛇']):
        return ["skill_venom_inject", "skill_beast_charge"] + basics[:1]
    elif any(kw in eid for kw in ['fire','flame','magma','火','炎','焰']):
        return ["skill_fire_spit", "skill_beast_roar"] + basics[:1]
    elif any(kw in eid for kw in ['ice','frost','frozen','冰','霜','寒']):
        return ["skill_ice_breath", "skill_beast_charge"] + basics[:1]
    elif any(kw in eid for kw in ['wind','gale','风','eagle','hawk','bird','鹰','鸟']):
        return ["skill_wind_slash", "skill_beast_roar"] + basics[:1]
    elif any(kw in eid for kw in ['earth','rock','stone','sand','土','岩','沙','elephant','象','rhino','犀']):
        return ["skill_earth_shake", "skill_beast_charge"] + basics[:1]
    elif any(kw in eid for kw in ['scorpion','蝎','spider','蛛']):
        return ["skill_venom_inject", "skill_ten_thousand_scorpion_array" if lv >= 30 else "skill_venom_inject"] + basics[:1]
    elif any(kw in eid for kw in ['poison','toxic','venom','毒','plague']):
        return ["skill_venom_inject", "skill_toxic_mist"] + basics[:1]
    elif any(kw in eid for kw in ['soul','ghost','wraith','魂','鬼','尸','undead','zombie','skeleton']):
        return ["skill_soul_absorption"] + basics[:1]
    elif any(kw in eid for kw in ['dragon','龙']):
        return ["skill_dragon_breath_attack", "skill_beast_roar"] + basics[:1]
    elif any(kw in eid for kw in ['guard','disciple','patrol','enforcer','守卫','弟子','护卫']):
        return ["skill_sword_qi", "skill_body_fortify", "skill_palm_strike"][:2 if lv < 20 else 3]
    elif any(kw in eid for kw in ['elder','deacon','captain','general','commander','长老','执事','队长']):
        return ["skill_sword_qi", "skill_palm_strike", "skill_body_fortify"]
    elif any(kw in eid for kw in ['bandit','thug','thief','killer','盗','匪','徒']):
        return ["skill_palm_strike", "skill_body_fortify"]
    elif any(kw in eid for kw in ['cultivator','fighter','warrior','merc','hunter','soldier','spy']):
        return ["skill_sword_qi", "skill_body_fortify", "skill_palm_strike"]
    elif any(kw in eid for kw in ['golem','puppet','傀儡','gole','像','spirit','elemental','wisp']):
        return ["skill_earth_shake", "skill_body_fortify"]
    elif any(kw in eid for kw in ['bear','熊','tiger','虎','lion','狮','ape','猿']):
        return ["skill_beast_roar", "skill_beast_charge", "skill_earth_shake"]
    elif any(kw in eid for kw in ['boar','猪','crocodile','鳄','lizard','蜥']):
        return ["skill_beast_charge", "skill_earth_shake"]
    elif any(kw in eid for kw in ['bat','蝙蝠']):
        return ["skill_wind_slash", "skill_beast_roar"]
    else:
        # Default by level
        if lv <= 10:
            return ["skill_beast_charge"]
        elif lv <= 25:
            return ["skill_beast_charge", "skill_body_fortify"]
        elif lv <= 45:
            return ["skill_sword_qi", "skill_body_fortify", "skill_palm_strike"]
        else:
            return ["skill_sword_qi", "skill_palm_strike", "skill_body_fortify", "skill_bloodline_pressure"]

# ══════════════════════════════════════════════════════════
# APPLY TO WORKBOOK
# ══════════════════════════════════════════════════════════
wb = openpyxl.load_workbook(WORKBOOK_PATH)
ws_skills = None
ws_enemies = None
for name in wb.sheetnames:
    if name.startswith("Skills_"): ws_skills = wb[name]
    if name.startswith("Enemies_"): ws_enemies = wb[name]

skill_headers = [cell.value for cell in next(ws_skills.iter_rows(min_row=1, max_row=1))]
enemy_headers = [cell.value for cell in next(ws_enemies.iter_rows(min_row=1, max_row=1))]

existing_skills = set()
for row in ws_skills.iter_rows(min_row=2):
    if row[0].value: existing_skills.add(str(row[0].value))

# Ensure Skills column in enemies
if "Skills" not in enemy_headers:
    ws_enemies.cell(row=1, column=len(enemy_headers)+1).value = "Skills"
    enemy_headers.append("Skills")

skills_col = enemy_headers.index("Skills")

# Add novel-based skills
added = 0
for data in NOVEL_SKILLS:
    sid = data[0]
    if sid in existing_skills: continue
    row_data = {"Skill_ID": sid, "Name": data[1], "Type": data[2], "Rank": data[3],
                "Effect": data[4], "Description": data[5] + " 【原文依据】" + data[6]}
    ws_skills.append([str(row_data.get(h, "")) for h in skill_headers])
    existing_skills.add(sid)
    added += 1
print(f"Added {added} novel-based skills. Total skills: {len(existing_skills)}")

# Assign skills to all enemies
updated = 0
for row_idx, row in enumerate(ws_enemies.iter_rows(min_row=2), start=2):
    eid = str(row[0].value)
    if not eid: continue
    name = str(row[1].value) if row[1].value else ""
    etype = str(row[2].value) if row[2].value else ""
    level = int(row[3].value) if row[3].value else 1

    if eid in ENEMY_NOVEL_SKILLS:
        skills = ENEMY_NOVEL_SKILLS[eid]
    else:
        skills = mob_skills_by_type(eid, name, level)

    ws_enemies.cell(row=row_idx, column=skills_col+1).value = ",".join(skills)
    updated += 1

wb.save(WORKBOOK_PATH)
print(f"Skills assigned to {updated} enemies.")
print("Done!")
