"""
Add missing locations from the novel to the game workbook.
Run once to add maps, encounters.
"""
import sys, copy
from pathlib import Path
import openpyxl

WORKBOOK_PATH = Path("story/text_game_event_schema_v4.xlsx")
BACKUP_PATH = Path("story/text_game_event_schema_v4_backup.xlsx")

# Backup
if not BACKUP_PATH.exists():
    import shutil
    shutil.copy2(WORKBOOK_PATH, BACKUP_PATH)
    print(f"Backed up to {BACKUP_PATH}")

wb = openpyxl.load_workbook(WORKBOOK_PATH)


def find_sheet(name_prefix):
    for name in wb.sheetnames:
        if name.startswith(name_prefix):
            return wb[name]
    raise KeyError(f"Sheet starting with '{name_prefix}' not found")


def get_existing_ids(ws):
    ids = set()
    for row in ws.iter_rows(min_row=2):
        val = row[0].value
        if val:
            ids.add(str(val))
    return ids


def append_row(ws, data: dict):
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    row_data = []
    for h in headers:
        val = data.get(h, "")
        row_data.append(val if val is not None else "")
    ws.append(row_data)


ws_maps = find_sheet("Maps_")
ws_encounters = find_sheet("Map_Encounters_")

existing_maps = get_existing_ids(ws_maps)
existing_encounters = get_existing_ids(ws_encounters)
print(f"Existing maps: {len(existing_maps)}, encounters: {len(existing_encounters)}")

# ========== NEW MAPS ==========
NEW_MAPS = [
    # ── 加玛帝国扩展 ──
    ("map_mo_city", "漠城", "加玛帝国", "", 6, 5, "yes", "",
     "塔戈尔沙漠边缘的最后城池，冰皇海波东隐居于此。城内佣兵与探险者云集，沙漠情报与稀有材料时有流通。",
     "沙漠剧情关键节点"),
    ("map_stone_mo_city", "石漠城", "加玛帝国", "", 7, 5, "yes", "",
     "深入沙漠前的补给重镇，蛇人族活动频繁。萧炎在此遇见青鳞，命运的齿轮开始转动。",
     "青鳞相遇剧情"),
    ("map_black_rock_city", "黑岩城", "加玛帝国", "", 4, 5, "yes", "",
     "加玛帝国东部大城，设有炼药师公会分会。城内帮派林立，地下黑市暗中流通禁售药材。",
     "炼药师考核剧情"),
    ("map_salt_city", "盐城", "加玛帝国", "", 5, 5, "yes", "",
     "以盐矿闻名的贸易城市，通往魔兽山脉东麓的门户。盐帮掌控城内大部分生意。",
     "加玛帝国城市"),
    ("map_ghost_pass", "镇鬼关", "加玛帝国", "", 6, 6, "no", "",
     "拱卫加玛帝国北境的雄关，驻军日夜警戒着边境魔兽与敌国斥候。",
     "边境关隘"),

    # ── 出云帝国扩展 ──
    ("map_scorpion_gate", "万蝎门", "出云帝国", "", 34, 8, "no", "",
     "出云帝国内与毒宗分庭抗礼的蝎毒宗门，蝎毕岩掌控的万蝎大阵令人闻风丧胆。",
     "出云剧情扩展"),

    # ── 黑角域扩展 ──
    ("map_black_seal_city", "黑印城", "黑角域", "", 20, 6, "yes", "",
     "黑角域核心大城，八扇门掌控的地下拍卖会名震黑角域。珍品功法、药材与禁物在此暗流涌动。",
     "黑角域拍卖城"),
    ("map_demon_flame_valley", "魔炎谷", "黑角域", "", 36, 10, "no", "",
     "黑角域三大势力之一，方言三老坐镇的魔道重地。谷中毒火终年不熄，与韩枫联手对抗迦南学院。",
     "黑角域清算战"),
    ("map_tianya_city", "天涯城", "黑角域边境", "", 26, 6, "yes", "",
     "方圆千里唯一拥有通往中州空间虫洞的城市，罗家世代守护虫洞。无数强者在此歇脚，等待虫洞开启。",
     "黑角域→中州虫洞枢纽"),

    # ── 中州扩展 ──
    ("map_tianmu_mountains", "天目山脉", "中州", "", 46, 8, "no", "",
     "灵气充沛的古老山脉，天山血潭隐于云雾深处。能量潮汐周期性冲刷山脉，是淬炼肉身的绝佳之地。",
     "血潭修炼剧情"),
    ("map_heaven_mountain_blood_pool", "天山血潭", "中州", "", 48, 10, "no", "",
     "天目山脉之巅的天然血池，地脉精华汇聚而成。能够淬炼筋骨、加速斗气凝练，中州青年才俊趋之若鹜。",
     "血潭核心"),
    ("map_sky_demon_sect", "天冥宗", "中州", "", 42, 8, "no", "",
     "中州老牌宗门，冥系功法诡异莫测。山门常年笼罩在幽暗雾气中，外人难窥其真面目。",
     "中州宗门"),
    ("map_death_corpse_mountains", "葬尸山脉", "中州", "", 55, 10, "no", "",
     "魂殿在中州的巢穴之一，尸气弥漫的死寂山岭。据说山脉深处隐藏着魂殿培养护法的秘地。",
     "魂殿据点"),

    # ── 魂殿扩展 ──
    ("map_heavenly_gang_hall", "天罡殿", "魂殿", "", 62, 10, "no", "",
     "魂殿总部核心大殿，殿主魂灭生掌控的灵魂中枢。殿内囚禁着无数被收集的强大灵魂体。",
     "魂殿总部"),

    # ── 妖火相关 ──
    ("map_demon_flame_plain", "妖火平原", "中州", "", 64, 8, "no", "",
     "净莲妖火空间现世前妖火气息外泄形成的焦土平原。奇异的火属性能量扭曲了整片区域。",
     "妖火前哨"),

    # ── 古帝洞府扩展 ──
    ("map_strange_flame_square", "异火广场", "古帝洞府", "", 78, 10, "no", "",
     "陀舍古帝以异火本源塑造的传承广场，二十二种异火的虚影在此永恒燃烧。帝品雏丹的气息弥漫其中。",
     "古帝传承核心"),
]


# ========== ENCOUNTERS ==========
def enc(evt_id, map_id, text, opt1_text="", opt1_effect="",
        opt2_text="", opt2_effect="", opt3_text="", opt3_effect="", weight=1):
    return {
        "Event_ID": evt_id, "Map_ID": map_id, "Weight": weight,
        "Text": text, "Pre_Condition": "",
        "Opt1_Text": opt1_text, "Opt1_Condition": "", "Opt1_Next": "", "Opt1_Effect": opt1_effect,
        "Opt2_Text": opt2_text, "Opt2_Condition": "", "Opt2_Next": "", "Opt2_Effect": opt2_effect,
        "Opt3_Text": opt3_text, "Opt3_Condition": "", "Opt3_Next": "", "Opt3_Effect": opt3_effect,
        "Notes": "",
    }


NEW_ENCOUNTERS = [
    # ── 漠城 ──
    enc("enc_mo_1", "map_mo_city",
        "漠城街角，一名披着斗篷的老者静坐于茶摊旁，斗篷下隐隐透出冰寒气息。他的目光掠过你，微微一顿。",
        "上前搭话", "soul:+3,rel:npc_hai_bodong:+5",
        "暗中观察", "soul:+2",
        "购买沙漠补给", "item:+item_elixir"),
    enc("enc_mo_2", "map_mo_city",
        "一间不起眼的古图店铺门可罗雀，店内墙上挂着几张泛黄的残破地图。其中一张残图散发出奇异的温热。",
        "仔细查看残图", "soul:+5",
        "询问店主来历", "soul:+2,reputation:+1"),
    enc("enc_mo_3", "map_mo_city",
        "沙漠商队在漠城集市招募护卫，领队高喊需要熟悉沙漠地形的斗者。报酬丰厚，但路程凶险。",
        "接受护卫委托", "exp:+40,silver:+15",
        "打听沙漠情报", "soul:+2",
        "采购沙漠药材", "item:+herb_desert_mandala"),

    # ── 石漠城 ──
    enc("enc_stone_mo_1", "map_stone_mo_city",
        "石漠城佣兵大厅内，一名怯生生的少女正在被几个佣兵调笑。少女的眼眸深处，隐约有碧绿色的蛇瞳一闪而过。",
        "出手解围", "rel:npc_qing_lin:+10,reputation:+5",
        "静观其变", "soul:+2"),
    enc("enc_stone_mo_2", "map_stone_mo_city",
        "城外传来蛇人族巡逻队的嘶鸣声，城墙上守卫拉响了警铃。石漠城进入了战备状态。",
        "协助守城", "exp:+50,reputation:+8",
        "护送平民撤离", "reputation:+5"),
    enc("enc_stone_mo_3", "map_stone_mo_city",
        "城中沙尘飞扬，一支疲惫的商队刚从沙漠深处归来。他们声称在沙暴中看到了一座埋藏的古殿。",
        "询问古殿位置", "soul:+3",
        "购买商队货物", "item:+herb_blood_lotus_essence"),

    # ── 黑岩城 ──
    enc("enc_black_rock_1", "map_black_rock_city",
        "黑岩城炼药师公会分会门前排起了长队，今天是月度考核日。守门的执事目光挑剔地扫过排队者。",
        "报名参加考核", "alchemy:+2,reputation:+3",
        "观摩考核过程", "alchemy:+1",
        "购买公会药材", "item:+herb_spirit_gathering"),
    enc("enc_black_rock_2", "map_black_rock_city",
        "黑岩城地下黑市入口藏在一间废弃的铁匠铺后，有两个壮汉守着暗门。门缝中透出药材与丹药的异香。",
        "潜入黑市", "exp:+30,soul:+2",
        "与黑市商人交易", "item:+herb_green_flame_grass",
        "通报公会", "reputation:+3"),
    enc("enc_black_rock_3", "map_black_rock_city",
        "城东帮派火并，硝烟中夹杂着斗气爆裂声。一个受伤的帮派成员跌跌撞撞向你求救。",
        "出手平息纷争", "exp:+60,reputation:+5",
        "趁机搜刮物资", "silver:+20",
        "坐山观虎斗", "soul:+2"),

    # ── 盐城 ──
    enc("enc_salt_1", "map_salt_city",
        "盐城码头边，盐帮的苦力正在搬运成袋的盐矿。监工挥鞭催促，空气中弥漫着咸涩的气味。",
        "找监工打听情报", "soul:+2",
        "帮忙搬运换取报酬", "silver:+10,exp:+20"),
    enc("enc_salt_2", "map_salt_city",
        "盐帮总舵大门紧闭，据传帮主正在与神秘来客密谈。墙外有几个鬼祟的身影正在偷听。",
        "加入偷听", "soul:+3",
        "从正门拜访", "reputation:+2"),

    # ── 镇鬼关 ──
    enc("enc_ghost_pass_1", "map_ghost_pass",
        "镇鬼关城墙上箭痕累累，守关将领正在集结士兵。斥候回报：北方有大股魔兽异动。",
        "协助侦查", "exp:+40,soul:+3",
        "加固城防", "exp:+30,reputation:+5"),
    enc("enc_ghost_pass_2", "map_ghost_pass",
        "关隘下的黑市中，走私者正以低价出售边境缴获的魔兽材料。价格诱人，但来路不正。",
        "购买战利品", "item:+core_magic",
        "举报走私", "reputation:+5",
        "无视", ""),

    # ── 万蝎门 ──
    enc("enc_scorpion_1", "map_scorpion_gate",
        "万蝎门山谷入口爬满了拳头大小的毒蝎，蝎尾在月光下泛着幽绿光芒。空气中弥漫着腐臭的毒雾。",
        "闯入万蝎阵", "exp:+120,alchemy:+3",
        "采集蝎毒", "item:+herb_desert_mandala",
        "绕道而行", "exp:+30"),
    enc("enc_scorpion_2", "map_scorpion_gate",
        "蝎毕岩端坐于万蝎王座之上，巨大的蝎尾虚影在其身后缓缓摆动。他的目光如毒刺般锁定你。",
        "正面迎战", "exp:+200,reputation:+15",
        "谈判回旋", "reputation:+5,soul:+3"),

    # ── 黑印城 ──
    enc("enc_black_seal_1", "map_black_seal_city",
        "八扇门总部的铁门紧闭，门上八个铜环在风中叮当作响。门缝中透出拍卖厅的喧哗声。",
        "设法进入拍卖会", "silver:-20,item:+herb_blood_essence_fruit",
        "在门外偷听", "soul:+2"),
    enc("enc_black_seal_2", "map_black_seal_city",
        "黑印城地下拍卖会人声鼎沸，一件件珍品被推上展台。压轴拍品是一卷地阶斗技残卷。",
        "参与竞拍", "silver:-50,exp:+50",
        "放弃竞拍观察买家", "soul:+3"),
    enc("enc_black_seal_3", "map_black_seal_city",
        "拍卖会结束后的暗巷中，两个竞拍失败者正在密谋劫杀刚刚拍得宝物的买家。",
        "暗中保护买家", "reputation:+8,exp:+40",
        "黑吃黑", "silver:+30",
        "当做没看见", ""),

    # ── 魔炎谷 ──
    enc("enc_demon_valley_1", "map_demon_flame_valley",
        "魔炎谷入口毒火肆虐，方言三老布下的护谷大阵将闯入者困于火海。火焰中夹杂着凄厉的哀嚎。",
        "以异火开路", "exp:+150,flag:demon_flame_valley_defeated=1",
        "寻找阵法破绽", "soul:+5"),
    enc("enc_demon_valley_2", "map_demon_flame_valley",
        "谷中密室中存放着魔炎谷多年积累的功法卷轴与炼药秘方，韩枫在此留下了海心焰的使用笔记。",
        "搜刮功法", "exp:+80,alchemy:+3",
        "只取炼药笔记", "alchemy:+5"),
    enc("enc_demon_valley_3", "map_demon_flame_valley",
        "方言三老中的最后一位负隅顽抗，他燃烧灵魂催动了谷中隐藏的禁术——魔火焚天阵。",
        "正面击杀", "exp:+200,reputation:+10",
        "摧毁阵眼", "soul:+8,exp:+100"),

    # ── 天涯城 ──
    enc("enc_tianya_1", "map_tianya_city",
        "天涯城空间虫洞广场上人群熙攘，罗家护卫维持着秩序。虫洞出口处，一位刚从虫洞中踏出的老者面色苍白地扶着墙喘息。",
        "打听虫洞通行规则", "soul:+2,reputation:+1",
        "观察虫洞状态", "soul:+3"),
    enc("enc_tianya_2", "map_tianya_city",
        "罗家大小姐在街市上横冲直撞，一个卖药老人被她撞翻在地。围观者敢怒不敢言——罗家掌控着天涯城的虫洞。",
        "扶起老人", "reputation:+5,rel:npc_yao_lao:+1",
        "与罗家大小姐理论", "reputation:+3,exp:+40"),
    enc("enc_tianya_3", "map_tianya_city",
        "罗家族长宣布虫洞出现不稳定波动，暂时关闭维修。大批等待前往中州的强者被滞留城中。",
        "帮忙维修虫洞", "soul:+8,exp:+80",
        "在城中等待", "exp:+20,soul:+2",
        "打听其他路线", "soul:+3"),
    enc("enc_tianya_4", "map_tianya_city",
        "虫洞重新开启，罗家护卫逐一查验通行资格。前方有人因没有足够的过路费而被拦下。",
        "支付过路费", "silver:-30,exp:+50",
        "以实力争取免费通行", "exp:+80,reputation:+5"),

    # ── 天目山脉 ──
    enc("enc_tianmu_1", "map_tianmu_mountains",
        "天目山脉能量潮汐如期而至，天地间的斗气如浪涛般冲刷着山脉。无数修炼者逆着能量流向上攀登。",
        "借此机会淬炼肉身", "douqi:+15,exp:+100",
        "寻找血潭入口", "soul:+5"),
    enc("enc_tianmu_2", "map_tianmu_mountains",
        "山腰处传来打斗声，两名年轻强者正在争夺一处能量最浓郁的修炼洞穴。",
        "调解争端", "reputation:+5",
        "抢夺洞穴", "exp:+80,douqi:+5"),

    # ── 天山血潭 ──
    enc("enc_blood_pool_1", "map_heaven_mountain_blood_pool",
        "天山血潭呈深邃的血红色，地脉精华在潭水中翻涌。潭边铭刻着古老的淬体符文。",
        "跳入血潭淬炼", "douqi:+25,exp:+150",
        "先研究符文", "soul:+8,douqi:+10"),
    enc("enc_blood_pool_2", "map_heaven_mountain_blood_pool",
        "血潭深处浮现出一具远古强者遗留的骸骨，骸骨手中的纳戒仍在散发微弱光芒。",
        "取走纳戒", "exp:+100,item:+item_elixir",
        "拜祭骸骨", "soul:+5,reputation:+3"),

    # ── 天冥宗 ──
    enc("enc_sky_demon_1", "map_sky_demon_sect",
        "天冥宗山门笼罩在幽暗雾气中，守门弟子以冥气凝聚成鬼面，质问来意。",
        "递交拜帖", "reputation:+3",
        "展示实力强闯", "exp:+100,reputation:+5"),
    enc("enc_sky_demon_2", "map_sky_demon_sect",
        "天冥宗大殿内，宗主正与魂殿护法密谈。你隐约听到'灵魂本源'和'天府联盟'的字眼。",
        "暗中偷听", "soul:+8",
        "正面现身质问", "exp:+120,reputation:+10"),

    # ── 葬尸山脉 ──
    enc("enc_corpse_mt_1", "map_death_corpse_mountains",
        "葬尸山脉尸气弥漫，山道两旁堆满了不知名的骸骨。远处传来魂殿锁链拖曳灵魂的凄厉声响。",
        "循声追踪", "exp:+100,soul:+5",
        "小心绕行", "exp:+40,soul:+2"),
    enc("enc_corpse_mt_2", "map_death_corpse_mountains",
        "一处隐蔽的山洞中囚禁着数十个被铁链锁住的灵魂体。他们的记忆正在被魂殿秘法抽取。",
        "释放被囚灵魂", "soul:+10,reputation:+10",
        "记录魂殿手法", "soul:+5,alchemy:+3"),

    # ── 天罡殿 ──
    enc("enc_heavenly_gang_1", "map_heavenly_gang_hall",
        "天罡殿正厅深处，魂殿殿主魂灭生端坐于灵魂王座。无数灵魂本源在他身边缭绕，发出痛苦的哀鸣。",
        "正面挑战魂灭生", "exp:+300,reputation:+20",
        "释放灵魂本源", "soul:+15,reputation:+15"),
    enc("enc_heavenly_gang_2", "map_heavenly_gang_hall",
        "天罡殿偏殿中陈列着魂殿收集的灵魂档案——记录了数百年间被夺取灵魂的强者名单。",
        "翻阅档案", "soul:+8",
        "销毁档案", "reputation:+10"),

    # ── 妖火平原 ──
    enc("enc_demon_plain_1", "map_demon_flame_plain",
        "妖火平原上空的云层被染成了诡异的粉红色，净莲妖火的气息从空间中渗出，在平原上制造出无数火焰幻象。",
        "感悟妖火气息", "soul:+10",
        "收集火属性能量", "douqi:+10,exp:+80"),
    enc("enc_demon_plain_2", "map_demon_flame_plain",
        "各方势力在妖火平原上扎营对峙，都在等待妖火空间开启的最佳时机。空气中弥漫着剑拔弩张的紧张感。",
        "建立据点", "reputation:+5,exp:+50",
        "刺探各方情报", "soul:+5"),

    # ── 异火广场 ──
    enc("enc_flame_square_1", "map_strange_flame_square",
        "异火广场上二十二种异火虚影徐徐燃烧，每一种都散发着独特的能量波动。陀舍古帝的意志在其中低语。",
        "感悟古帝意志", "soul:+20,douqi:+30",
        "尝试收取异火虚影", "exp:+150"),
    enc("enc_flame_square_2", "map_strange_flame_square",
        "广场中央，帝品雏丹悬浮于光柱之中。斗帝级别的禁制层层叠叠地将其保护在核心。",
        "尝试夺取雏丹", "exp:+300,soul:+10",
        "研究禁制结构", "soul:+15,alchemy:+5"),
]

# ========== APPLY ==========
print(f"\nAdding {len(NEW_MAPS)} new maps...")
for data in NEW_MAPS:
    map_id = data[0]
    if map_id in existing_maps:
        print(f"  SKIP (exists): {map_id}")
        continue
    row = {
        "Map_ID": data[0],
        "Name": data[1],
        "Region": data[2],
        "Unlock_Condition": data[3],
        "Recommend_Level": data[4],
        "Explore_Stamina_Cost": data[5],
        "Default_Encounter_Weight": 1,
        "Safe_Zone": data[6],
        "Exit_Event": data[7],
        "Description": data[8],
        "Notes": data[9],
    }
    append_row(ws_maps, row)
    existing_maps.add(map_id)
    print(f"  + {map_id}: {data[1]}")

print(f"\nAdding {len(NEW_ENCOUNTERS)} new encounters...")
for data in NEW_ENCOUNTERS:
    evt_id = data["Event_ID"]
    if evt_id in existing_encounters:
        print(f"  SKIP (exists): {evt_id}")
        continue
    append_row(ws_encounters, data)
    existing_encounters.add(evt_id)
    print(f"  + {evt_id}")

wb.save(WORKBOOK_PATH)
print(f"\nSaved to {WORKBOOK_PATH}")
print("Done!")
