"""
Fix item acquisition channels:
- Remove legendary/rare gathering spots (should be boss drop / encounter only)
- Add drop tables to appropriate bosses
- Keep only common/uncommon items in gathering
"""
from pathlib import Path
import openpyxl

WORKBOOK_PATH = Path("story/text_game_event_schema_v4.xlsx")
wb = openpyxl.load_workbook(WORKBOOK_PATH)

def find_sheet(name_prefix):
    for name in wb.sheetnames:
        if name.startswith(name_prefix):
            return wb[name]
    raise KeyError(f"Sheet starting with '{name_prefix}' not found")

def get_existing_ids(ws):
    ids = set()
    for row in ws.iter_rows(min_row=2):
        val = row[0].value
        if val:
            ids.add(str(val))
    return ids

# ── 1. REMOVE HIGH-VALUE GATHERING SPOTS ──
ws_gathering = find_sheet("Map_Gathering")
gather_to_remove = {
    # Legendary items - boss drops only
    "gather_dragon_scale",      # 龙鳞碎片 → 北龙王/南龙王/西龙王 drop
    "gather_dragon_saliva",     # 龙涎灵草 → 龙岛boss drop
    "gather_bodhi_leaf",        # 菩提叶 → 菩提古树 encounter
    "gather_demon_crystal",     # 妖火结晶 → 净世白莲火 encounter
    "gather_ancient_relic",     # 远古遗物 → 远古遗迹 boss
    # Rare items - encounter/NPC only
    "gather_forbidden_pill",    # 禁药配方残页 → 暗角域 encounter
    "gather_transforming_herb", # 化形草 → 兽域 encounter
    "gather_ancient_seed",      # 古域灵种 → 莽荒古域 encounter
    "gather_beast_blood",       # 魔兽精血瓶 → 兽域 encounter/boss
    # Quest items - special acquisition only
    "gather_wormhole_pass_tianya", # 虫洞通行令 → 天涯城 encounter
    "gather_wormhole_pass_th",     # 虫洞通行令 → 天黄城 encounter
    "gather_black_market_pass",    # 黑市令牌 → 黑印城 encounter
}

removed = 0
rows_to_delete = []
for row_idx, row in enumerate(ws_gathering.iter_rows(min_row=2), start=2):
    gid = row[0].value
    if str(gid) in gather_to_remove:
        rows_to_delete.append(row_idx)
        print(f"  REMOVE gather: {gid}")

# Delete from bottom to keep indices valid
for row_idx in sorted(rows_to_delete, reverse=True):
    ws_gathering.delete_rows(row_idx)
    removed += 1
print(f"Removed {removed} high-value gathering spots")

# ── 2. ADD BOSS DROP TABLES ──
ws_enemies = find_sheet("Enemies_")
headers = [cell.value for cell in next(ws_enemies.iter_rows(min_row=1, max_row=1))]
drop_col = headers.index("Drop_Table")

# Boss drop assignments
boss_drops = {
    "enemy_medusa": "item:item_snake_venom_vial:80|item:item_desert_compass:50|exp:1500:100",
    "enemy_han_feng": "item:item_sea_heart_flame:30|item:item_forbidden_pill_fragment:50|exp:5000:100",
    "enemy_yun_shan": "item:item_elixir:30|exp:4000:100",
    "enemy_wu_hufa": "item:item_soul_baby_fruit:20|exp:4500:100",
    "enemy_fan_lao": "item:item_blood_essence_fruit:40|exp:3000:100",
    "enemy_di_mo_laogui": "item:item_demon_puppet:30|exp:6500:100",
    "enemy_zhai_xing": "item:item_soul_baby_fruit:40|exp:15000:100",
    "enemy_mugu": "item:item_elixir:80|item:item_alchemy_handbook:40|exp:12000:100",
    "enemy_xuwu": "item:item_nihility_flame:100|item:item_heavenly_flame:100|exp:80000:100",
    "enemy_mo_tianxing": "item:item_elixir:40|exp:5000:100",
    "enemy_hun_yu": "item:item_soul_baby_fruit:50|exp:25000:100",
    "enemy_hun_feng": "item:item_soul_baby_fruit:60|exp:30000:100",
    "enemy_xie_biyan": "item:item_snake_venom_vial:60|exp:10000:100",
    "enemy_hunmo": "item:item_soul_baby_fruit:50|exp:10000:100",
    "enemy_hun_yuantian": "item:item_soul_baby_fruit:70|exp:10000:100",
    "enemy_hun_miesheng": "item:item_soul_baby_fruit:80|exp:10000:100",
    "enemy_binghe": "item:item_elixir:50|exp:10000:100",
    "enemy_four_demon_saints": "item:item_forbidden_pill_fragment:60|exp:10000:100",
    "enemy_hun_shengtian": "item:item_soul_baby_fruit:90|exp:10000:100",
    "enemy_hun_yao": "item:item_soul_baby_fruit:40|exp:10000:100",
    # New boss drops for regional items
    "enemy_fang_yan": "item:item_forbidden_pill_fragment:35|exp:10000:100",
}

print(f"\nUpdating {len(boss_drops)} boss drop tables...")
for row_idx, row in enumerate(ws_enemies.iter_rows(min_row=2), start=2):
    eid = str(row[0].value)
    if eid in boss_drops:
        old_drop = row[drop_col].value or ""
        new_drop = boss_drops[eid]
        ws_enemies.cell(row=row_idx, column=drop_col+1).value = new_drop
        print(f"  UPDATE {eid}: {old_drop[:30]}... -> {new_drop[:60]}...")

# ── 3. KEEP ONLY COMMON GATHERING (herbs, cores, basic mats) ──
# The remaining gathering spots are for:
# - Common herbs (凝血草, 紫叶兰草, etc.) in appropriate maps
# - Basic monster cores in appropriate maps
# - Regional common specialties (特产药材包, 通关文牒, 魔兽皮料包, etc.)
# These are fine to keep.

wb.save(WORKBOOK_PATH)
print(f"\nSaved. Gathering cleaned, boss drops enriched.")
print("Done!")
