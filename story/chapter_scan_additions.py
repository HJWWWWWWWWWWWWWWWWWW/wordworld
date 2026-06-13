"""
Final chapter-scan additions: NPCs, maps, encounters.
Adds all missing content found in the chapter-by-chapter scan.
"""
from pathlib import Path
import openpyxl

WORKBOOK_PATH = Path("story/text_game_event_schema_v4.xlsx")

wb = openpyxl.load_workbook(WORKBOOK_PATH)


def find_sheet(name_prefix):
    for name in wb.sheetnames:
        if name.startswith(name_prefix):
            return wb[name]
    raise KeyError(f"Sheet starting with '{name_prefix}' not found")


def get_existing_ids(ws):
    ids = set()
    for row in ws.iter_rows(min_row=2):
        val = row[0].value
        if val:
            ids.add(str(val))
    return ids


def append_row(ws, data: dict):
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    row_data = []
    for h in headers:
        val = data.get(h, "")
        row_data.append(val if val is not None else "")
    ws.append(row_data)


ws_maps = find_sheet("Maps_")
ws_encounters = find_sheet("Map_Encounters_")
ws_npcs = find_sheet("NPCs_")
existing_maps = get_existing_ids(ws_maps)
existing_encounters = get_existing_ids(ws_encounters)

# ══════════════════════════════════════════════════════════
# 1. MISSING NPCs
# ══════════════════════════════════════════════════════════
NEW_NPCS = [
    ("npc_liu_ling", "柳翎", "沧澜帝国炼药师公会，法犸弟子，丹王传人"),
    ("npc_gu_ni", "谷尼", "青石城二品炼药师，米特尔拍卖场鉴定师"),
    ("npc_mu_zhan", "木战", "迦南学院学员，木家子弟"),
    ("npc_yao_ming", "妖暝", "九幽地冥蟒族前族长，被囚于九幽地宫"),
    ("npc_yao_tian", "药天", "药族天才炼药师"),
    ("npc_yao_ling", "药灵", "药族炼药师"),
    ("npc_yao_huo", "药火", "药族"),
    ("npc_hun_sha", "魂煞", "玄族强者"),
    ("npc_hun_jing", "魂镜", "玄族"),
    ("npc_hun_tu", "魂屠", "玄族"),
    ("npc_gu_qian", "古谦", "云族长老"),
    ("npc_gu_xu", "古虚", "云族强者"),
    ("npc_gu_dao", "古道", "云族三仙之一"),
    ("npc_yan_huo", "炎火", "炎族强者"),
    ("npc_yan_yao", "炎耀", "炎族"),
    ("npc_lei_yun", "雷云", "雷族强者"),
    ("npc_lei_dian", "雷电", "雷族强者"),
    ("npc_jiutian_zun", "九天尊", "黑渊殿九大天尊统称"),
]

# Check existing NPCs
existing_npc_ids = set()
for row in ws_npcs.iter_rows(min_row=2):
    val = row[0].value
    if val:
        existing_npc_ids.add(str(val))

print(f"Adding {len(NEW_NPCS)} new NPCs...")
for npc_id, npc_name, npc_notes in NEW_NPCS:
    if npc_id in existing_npc_ids:
        print(f"  SKIP (exists): {npc_id}")
        continue
    row = {"Character_ID": npc_id, "Name": npc_name, "Notes": npc_notes}
    append_row(ws_npcs, row)
    print(f"  + {npc_id}: {npc_name}")

# ══════════════════════════════════════════════════════════
# 2. MISSING MAPS from chapter scan
# ══════════════════════════════════════════════════════════
NEW_MAPS = [
    # ── 沧澜帝国→暗角域途中 ──
    ("map_daling_city", "大岭城", "沧澜帝国", "", 8, 5, "yes", "",
     "拱卫沧澜帝国东境的边陲重镇，是离开沧澜前往暗角域的必经之路。往来商旅在此补给，守军戒备森严。",
     "边境枢纽"),
    ("map_black_domain_plain", "黑域大平原", "暗角域", "", 18, 12, "no", "",
     "连接沧澜帝国方向与暗角域核心地带的辽阔平原。法则混乱，亡命徒横行，没有任何规则可言。",
     "暗角域入口"),

    # ── 暗角域组织 ──
    ("map_blood_sect", "血宗", "暗角域", "", 24, 8, "no", "",
     "暗角域老牌势力，血宗秘法以血炼之术著称。范痨坐镇的阴森山门，血池翻涌令人作呕。",
     "暗角域势力"),
    ("map_eight_gates", "八扇门", "黑印城", "", 20, 5, "yes", "",
     "黑印城地下秩序的掌控者，八扇铜门背后是暗角域最大的地下拍卖网络。",
     "黑印城势力"),
    ("map_pan_gate", "磐门", "迦南内院", "", 22, 4, "yes", "",
     "林烬在内院创建的新生势力，以团结新生学员为核心。磐门之名寓意坚如磐石、不可动摇。",
     "内院组织"),
    ("map_xiao_gate", "林门", "迦南内院", "", 25, 4, "yes", "",
     "林烬离开后磐门发展壮大的延续，成为内院最强新生势力之一。林门弟子遍布内院。",
     "内院组织进阶"),

    # ── 中州地点 ──
    ("map_scorching_mountains", "炙火山脉", "中州", "", 42, 10, "no", "",
     "焚炎谷坐落于炙火山脉核心，火山口终年喷涌。天地火属性能量在此浓郁到极致。",
     "焚炎谷所在山脉"),
    ("map_qi_feng_mountain", "栖凤山", "中州", "", 40, 8, "no", "",
     "风雷阁北阁领地内的灵山，凤清儿修炼之地。山巅罡风如刀，是风属性修炼者圣地。",
     "风雷阁领地"),
    ("map_tianhuang_city", "天黄城", "中域", "", 52, 5, "yes", "",
     "中域繁华大城，通往丹域的重要中转站。虫洞广场常年拥挤，各方势力在此交换情报。",
     "中域中转"),
    ("map_black_fire_sect", "黑火宗", "中州", "", 45, 8, "no", "",
     "修炼黑火功法的小宗门，因得罪冰河谷而覆灭。焦黑的废墟中仍残留着不甘的怨念。",
     "中州小宗门"),
    ("map_wan_yao_mountains", "万药山脉", "丹域", "", 54, 8, "no", "",
     "丹界附近的原始山脉，盛产万年灵药。丹阁在此设有药兽保护区，进入需丹阁许可。",
     "丹域药山"),

    # ── 莽荒古域 ──
    ("map_manghuang_town", "莽荒镇", "莽荒古域", "", 62, 4, "yes", "",
     "莽荒古域外最后的人类聚居点。探索古域的强者在此做最后的补给与情报交换。",
     "古域前哨"),
    ("map_heaven_demon_blood_pool", "天魔血池", "莽荒古域", "", 66, 10, "no", "",
     "莽荒古域深处的远古血池，天魔蟒一族守护的禁地。血池之力可淬炼肉身至极端。",
     "古域禁地"),
    ("map_ancient_domain_platform", "古域台", "莽荒古域", "", 65, 8, "no", "",
     "莽荒古域中央的远古石台，菩提古树出世之地。石台上铭刻着上古符文，威压惊人。",
     "古域核心"),

    # ── 丹阁 ──
    ("map_small_dan_tower", "小丹阁", "丹阁", "", 65, 6, "yes", "",
     "丹阁内部最隐秘的圣地，只有丹阁最高层知晓其存在。大长老在此闭关，守护炼药师一脉的至高传承。",
     "丹阁核心圣地"),
    ("map_black_emperor_pavilion", "黑皇阁", "黑皇城", "", 28, 5, "yes", "",
     "黑皇宗接待各方强者的奢华会馆，大厅宽广如广场。暗角域最顶级的交易在此暗中进行。",
     "黑皇宗会馆"),

    # ── 葬天山脉 ──
    ("map_burial_sky_mountains", "葬天山脉", "中州", "", 88, 15, "no", "",
     "远古大战形成的破碎山脉，无数强者葬身于此。玄冥帝在此布置最终战场，天地为之变色。",
     "最终战场所在地"),
]

# ══════════════════════════════════════════════════════════
# 3. ENCOUNTERS for new maps
# ══════════════════════════════════════════════════════════
def enc(evt_id, map_id, text, opt1_text="", opt1_effect="",
        opt2_text="", opt2_effect="", opt3_text="", opt3_effect="", weight=1):
    return {
        "Event_ID": evt_id, "Map_ID": map_id, "Weight": weight,
        "Text": text, "Pre_Condition": "",
        "Opt1_Text": opt1_text, "Opt1_Condition": "", "Opt1_Next": "", "Opt1_Effect": opt1_effect,
        "Opt2_Text": opt2_text, "Opt2_Condition": "", "Opt2_Next": "", "Opt2_Effect": opt2_effect,
        "Opt3_Text": opt3_text, "Opt3_Condition": "", "Opt3_Next": "", "Opt3_Effect": opt3_effect,
        "Notes": "",
    }

NEW_ENCOUNTERS = [
    # ── 大岭城 ──
    enc("enc_daling_1", "map_daling_city",
        "大岭城城门处排着长队，守城士兵逐一盘查过往旅人。告示板上贴着边境外的最新魔兽动向。",
        "打听边境情报", "soul:+2", "帮助守军巡逻", "exp:+30,reputation:+3"),
    enc("enc_daling_2", "map_daling_city",
        "城中商队正在招募护卫，即将穿越边境前往暗角域方向。领队强调路途凶险，亡命徒与魔兽并存。",
        "接受护卫委托", "exp:+50,silver:+20", "独自上路", "exp:+30"),

    # ── 黑域大平原 ──
    enc("enc_black_plain_1", "map_black_domain_plain",
        "平原上狂风卷起沙尘，远处传来厮杀声。一队商旅正被流寇围攻，马车上的货物散落一地。",
        "出手救援", "reputation:+8,exp:+60", "绕道避开", "exp:+10"),
    enc("enc_black_plain_2", "map_black_domain_plain",
        "平原中央有一座无名石碑，碑文被风沙侵蚀得看不清了。碑下压着一柄锈迹斑斑的长剑。",
        "拔出长剑", "exp:+50,soul:+3", "研究碑文", "soul:+5"),

    # ── 血宗 ──
    enc("enc_blood_sect_1", "map_blood_sect",
        "血宗山门弥漫着浓重的血腥味，门前的血池中翻滚着不知名的红色液体。范痨的声音从深处传来。",
        "闯入血宗", "exp:+100,reputation:+8", "在门外伏击血宗弟子", "exp:+60"),
    enc("enc_blood_sect_2", "map_blood_sect",
        "血宗密室内收藏着血炼秘法的完整卷轴。但强行修炼会侵蚀心智——这是范痨都不敢触碰的禁忌篇章。",
        "抄录秘法", "exp:+80,soul:+5", "销毁秘法", "reputation:+10"),

    # ── 八扇门 ──
    enc("enc_eight_gates_1", "map_eight_gates",
        "八扇铜门前站着两排黑衣护卫，门环上刻着复杂的禁制纹路。门内隐约传出拍卖声。",
        "申请进入拍卖", "silver:-15,exp:+30", "在门外观察进出者", "soul:+2"),

    # ── 磐门 ──
    enc("enc_pan_gate_1", "map_pan_gate",
        "磐门总部内新生们正在切磋灵技，墙上挂着磐门的规章——团结、互助、不欺压同门。",
        "指点新生修炼", "reputation:+5,exp:+30", "申请加入磐门", "reputation:+3"),
    enc("enc_pan_gate_2", "map_pan_gate",
        "磐门仓库管理员正在清点火能收入。他悄悄告诉你，最近有人在暗中打压磐门的交易渠道。",
        "追查打压者", "exp:+80,soul:+3", "加强仓库守卫", "reputation:+5"),

    # ── 林门 ──
    enc("enc_xiao_gate_1", "map_xiao_gate",
        "林门大厅中悬挂着林烬的画像，弟子们在此议事。一名刚从暗角域回来的弟子带来了冷煜的消息。",
        "听取情报", "soul:+3", "派遣弟子打探更多", "reputation:+3"),

    # ── 炙火山脉 ──
    enc("enc_scorching_1", "map_scorching_mountains",
        "炙火山脉火山口喷涌出的岩浆汇成一条赤红河流，空气中火属性能量浓郁到化雾。",
        "在火山口修炼", "douqi:+15,exp:+100", "采集火山矿石", "item:+core_fire"),
    enc("enc_scorching_2", "map_scorching_mountains",
        "山腰的天然火窟中，唐火儿正引导地火淬炼筋骨。她的控火手法精妙绝伦。",
        "请教控火之术", "alchemy:+3,rel:npc_tang_huoer:+5", "旁观学习", "alchemy:+1"),

    # ── 栖凤山 ──
    enc("enc_qifeng_1", "map_qi_feng_mountain",
        "栖凤山巅罡风凛冽，凤清儿乘着天妖凰翼盘旋于空。她的目光如刀锋般扫过你的位置。",
        "迎风修炼身法", "spd:+10,exp:+80", "与凤清儿对话", "reputation:+3"),
    enc("enc_qifeng_2", "map_qi_feng_mountain",
        "山腰洞穴中发现了远古天妖凰留下的一根褪羽，羽毛上天然铭刻着风属性灵力运转路线。",
        "感悟褪羽", "douqi:+10,soul:+5", "收起褪羽", "item:+item_elixir"),

    # ── 天黄城 ──
    enc("enc_tianhuang_1", "map_tianhuang_city",
        "天黄城虫洞广场比天涯城更为繁忙——这里是中域交通枢纽。十余座虫洞同时运转，灵力光芒照亮半边天空。",
        "打听各方虫洞路线", "soul:+3", "前往丹域方向虫洞", "exp:+40"),
    enc("enc_tianhuang_2", "map_tianhuang_city",
        "城中最繁华的交易街上，一个云族装扮的摊主正在出售远古遗迹中挖出的残破玉简。",
        "购买玉简", "silver:-20,soul:+5", "鉴定真伪", "soul:+3"),

    # ── 黑火宗 ──
    enc("enc_black_fire_1", "map_black_fire_sect",
        "黑火宗的废墟中焦臭味未散，倒塌的殿墙上残留着冰河谷的寒冰掌印。几具被冻裂的尸体仍未入土。",
        "搜索幸存者", "reputation:+5,soul:+3", "搜寻遗物", "exp:+60"),
    enc("enc_black_fire_2", "map_black_fire_sect",
        "废墟地窖中藏着一卷黑火宗最后的功法——《黑火冥炎功》。虽是旁门左道，但控火之法确有独到之处。",
        "研习功法", "alchemy:+3,exp:+50", "带走出售", "silver:+30"),

    # ── 万药山脉 ──
    enc("enc_wan_yao_1", "map_wan_yao_mountains",
        "万药山脉中一株万年青灵藤攀附在悬崖上，藤蔓间挂着数颗青翠欲滴的灵果。守护魔兽在崖壁上留下深深爪痕。",
        "采摘灵果", "item:+herb_ancient_green_vine,exp:+80", "先引开魔兽", "exp:+50,soul:+3"),
    enc("enc_wan_yao_2", "map_wan_yao_mountains",
        "丹阁在此设有药兽园入口，一名丹阁执事正在审核入山采药者的资质。",
        "出示炼药师资格", "reputation:+3,alchemy:+2", "试图偷渡入山", "exp:+60"),

    # ── 莽荒镇 ──
    enc("enc_manghuang_town_1", "map_manghuang_town",
        "莽荒镇唯一的客栈里挤满了等待进入古域的各方强者。角落里一位老者自称上次古域开启时曾闯入深处。",
        "请老者喝酒套话", "silver:-5,soul:+5", "在镇中采购补给", "item:+item_elixir"),
    enc("enc_manghuang_town_2", "map_manghuang_town",
        "镇口的神木下，几个刚从古域退出的伤者正在接受治疗。他们的伤口上蔓延着诡异的黑色纹路——古域中的诅咒。",
        "询问古域内部情况", "soul:+3", "协助治疗伤员", "reputation:+5"),

    # ── 天魔血池 ──
    enc("enc_demon_blood_1", "map_heaven_demon_blood_pool",
        "天魔血池中池水呈紫黑色，远古天魔蟒的精血在其中翻涌。池边散落着被血毒侵蚀而死的闯入者骨骸。",
        "以源火护体进入血池", "douqi:+30,exp:+200", "在池边收集血晶", "item:+item_huangquan_blood_crystal"),
    enc("enc_demon_blood_2", "map_heaven_demon_blood_pool",
        "血池深处浮现出一具天魔蟒始祖的骸骨，头骨中嵌着一颗紫黑色的蛇珠——品质远超普通魔核。",
        "取出蛇珠", "exp:+150,item:+core_magic", "祭拜骸骨后离开", "soul:+8"),

    # ── 古域台 ──
    enc("enc_ancient_platform_1", "map_ancient_domain_platform",
        "古域台石板上铭刻的上古符文在月光下泛起微光。菩提古树的虚影在台上若隐若现。",
        "参悟符文", "soul:+12,douqi:+10", "静待古树虚影显化", "soul:+8"),
    enc("enc_ancient_platform_2", "map_ancient_domain_platform",
        "各方势力在古域台周围对峙，争夺距离菩提古树最近的位置。空气中弥漫着剑拔弩张的氛围。",
        "以实力争取位置", "exp:+180,reputation:+10", "与其他势力协商", "reputation:+8"),

    # ── 小丹阁 ──
    enc("enc_small_dan_1", "map_small_dan_tower",
        "小丹阁中，大长老正在考核你的炼药造诣。一尊古朴的药鼎中，九色丹雷正在凝聚。",
        "完成炼制", "alchemy:+8,soul:+5", "请教大长老", "alchemy:+5,rel_player_xuanlu_elder:+0"),
    enc("enc_small_dan_2", "map_small_dan_tower",
        "小丹阁的藏经阁中收藏着失传的远古丹方。守阁人递给你一卷《帝丹谱》的残卷。",
        "研读丹方", "alchemy:+10,soul:+8", "抄录带走", "alchemy:+6"),

    # ── 黑皇阁 ──
    enc("enc_black_pavilion_1", "map_black_emperor_pavilion",
        "黑皇阁大厅中灯火辉煌，暗角域各方势力首领围坐圆桌。莫天行正展示一件压轴的远古异宝。",
        "参与鉴赏", "soul:+3", "暗中记录各方关系", "soul:+5"),
    enc("enc_black_pavilion_2", "map_black_emperor_pavilion",
        "宴会间隙，黑皇宗的密使悄然接近你，提出一项秘密交易——以情报换取拍卖会上的优先竞拍权。",
        "接受交易", "soul:+3,silver:-20", "婉拒", "reputation:+3"),

    # ── 葬天山脉 ──
    enc("enc_burial_sky_1", "map_burial_sky_mountains",
        "葬天山脉连绵的群峰被玄冥帝以无上魂力扭曲成一座天然大阵。黑渊殿的黑旗插满每一座山头。",
        "率联军突入阵中", "exp:+300,reputation:+20", "寻找阵眼破解", "soul:+15,exp:+200"),
    enc("enc_burial_sky_2", "map_burial_sky_mountains",
        "山脉最高峰上，玄冥帝与虚无吞炎的身影并肩而立。两人的威压让整片天空都在颤抖。",
        "发起总攻", "exp:+500,reputation:+30", "布置联军合围", "exp:+350,soul:+10"),
]

# ══════════════════════════════════════════════════════════
# APPLY
# ══════════════════════════════════════════════════════════
print(f"\nAdding {len(NEW_MAPS)} new maps...")
for data in NEW_MAPS:
    map_id = data[0]
    if map_id in existing_maps:
        print(f"  SKIP (exists): {map_id}")
        continue
    row = {
        "Map_ID": data[0], "Name": data[1], "Region": data[2],
        "Unlock_Condition": data[3], "Recommend_Level": data[4],
        "Explore_Stamina_Cost": data[5], "Default_Encounter_Weight": 1,
        "Safe_Zone": data[6], "Exit_Event": data[7],
        "Description": data[8], "Notes": data[9],
    }
    append_row(ws_maps, row)
    existing_maps.add(map_id)
    print(f"  + {map_id}: {data[1]}")

print(f"\nAdding {len(NEW_ENCOUNTERS)} new encounters...")
for data in NEW_ENCOUNTERS:
    evt_id = data["Event_ID"]
    if evt_id in existing_encounters:
        print(f"  SKIP (exists): {evt_id}")
        continue
    append_row(ws_encounters, data)
    existing_encounters.add(evt_id)
    print(f"  + {evt_id}")

wb.save(WORKBOOK_PATH)
print(f"\nSaved to {WORKBOOK_PATH}")
print("Done!")
