"""
Supplementary enemy additions - deeper novel coverage.
Adds missing enemy types, regional variants, faction hierarchies.
"""
from pathlib import Path
import openpyxl

WORKBOOK_PATH = Path("story/text_game_event_schema_v4.xlsx")
wb = openpyxl.load_workbook(WORKBOOK_PATH)

def find_sheet(name_prefix):
    for name in wb.sheetnames:
        if name.startswith(name_prefix):
            return wb[name]
    raise KeyError

ws = find_sheet("Enemies_")
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
existing = set()
for row in ws.iter_rows(min_row=2):
    if row[0].value: existing.add(str(row[0].value))

def add(eid, name, etype, lv, hp, atk, df, spd, exp_r, drop="", notes=""):
    if eid in existing: return False
    row_data = {
        "Enemy_ID": str(eid), "Name": str(name), "Type": str(etype),
        "Level": str(lv), "HP": str(hp), "ATK": str(atk), "DEF": str(df),
        "SPD": str(spd), "Drop_Table": str(drop), "Exp_Reward": str(exp_r),
        "Win_Next": "", "Lose_Next": "", "Notes": str(notes) if notes else "",
    }
    vals = [row_data.get(h, "") for h in headers]
    ws.append(vals)
    existing.add(eid)
    return True

count = 0
def A(*a):
    global count
    if add(*a): count += 1

# ══════════════════════════════════════════════════════════
# MORE REGIONAL MOBS
# ══════════════════════════════════════════════════════════
REGIONAL_MOBS = [
    # ── 加玛帝国野外 ──
    ("mob_forest_boar", "森林野猪", "mob", 2, 120, 12, 8, 10, 20, "item:herb_coagulation:10|exp:20:100"),
    ("mob_grassland_wolf", "草原狼", "mob", 3, 150, 15, 10, 18, 30, "item:core_magic:10|exp:30:100"),
    ("mob_giant_mantis", "巨型螳螂", "mob", 6, 260, 32, 16, 28, 65, "item:core_wood:18|exp:65:100"),
    ("mob_venomous_centipede", "毒蜈蚣", "mob", 7, 280, 28, 18, 22, 70, "item:herb_coagulation:15|exp:70:100"),
    ("mob_cave_bat", "洞穴蝙蝠", "mob", 4, 160, 20, 8, 30, 40, "exp:40:100"),
    ("mob_mountain_bandit", "山贼", "mob", 5, 220, 24, 14, 20, 50, "item:silver:20|exp:50:100"),
    ("mob_river_crocodile", "河鳄", "mob", 8, 320, 34, 28, 15, 85, "item:item_beast_hide_bundle:15|exp:85:100"),
    ("mob_poison_frog", "毒箭蛙", "mob", 9, 280, 36, 15, 32, 95, "item:herb_desert_mandala:10|exp:95:100"),
    ("mob_rock_golem", "岩石傀儡", "mob", 11, 480, 38, 40, 10, 120, "item:core_earth:22|exp:120:100"),
    # 魔兽山脉内围
    ("mob_winged_serpent", "翼蛇", "mob", 10, 380, 45, 22, 35, 105, "item:core_wind:20|exp:105:100"),
    ("mob_iron_horn_rhino", "铁角犀牛", "mob", 12, 550, 40, 40, 15, 130, "item:item_beast_hide_bundle:20|exp:130:100"),
    ("mob_flame_salamander", "火焰蜥蜴", "mob", 11, 420, 48, 18, 30, 115, "item:core_fire:20|exp:115:100"),
    # 沙漠
    ("mob_quicksand_trap", "流沙陷阱兽", "mob", 10, 400, 30, 35, 10, 100, "item:core_earth:20|exp:100:100"),
    ("mob_sand_fox", "沙狐", "mob", 8, 280, 38, 15, 40, 85, "item:herb_desert_mandala:12|exp:85:100"),
    ("mob_oasis_bandit", "绿洲盗匪", "mob", 11, 420, 45, 22, 30, 115, "item:silver:35|exp:115:100"),
    ("mob_desert_eagle", "沙漠巨鹰", "mob", 13, 450, 50, 22, 45, 145, "item:core_wind:22|exp:145:100"),
    # 黑角域
    ("mob_alley_thief", "暗巷盗贼", "mob", 17, 580, 58, 24, 35, 175, "item:silver:30|exp:175:100"),
    ("mob_arena_slave", "竞技场奴隶斗士", "mob", 20, 700, 65, 30, 30, 220, "item:item_elixir:8|exp:220:100"),
    ("mob_poison_market_thug", "黑市毒贩", "mob", 22, 720, 68, 28, 38, 240, "item:herb_desert_mandala:15|exp:240:100"),
    ("mob_black_inn_assassin", "客栈刺客", "mob", 23, 680, 75, 25, 45, 260, "exp:260:100"),
    # 出云帝国
    ("mob_poison_master", "毒师", "mob", 26, 920, 75, 40, 35, 330, "item:herb_desert_mandala:20|exp:330:100"),
    ("mob_scorpion_tamer", "驭蝎者", "mob", 28, 950, 78, 42, 36, 350, "item:item_snake_venom_vial:15|exp:350:100"),
    ("mob_plague_doctor", "瘟疫医师", "mob", 29, 880, 70, 38, 42, 360, "item:herb_blood_lotus_essence:12|exp:360:100"),
    # 中州
    ("mob_arena_challenger", "竞技场挑战者", "mob", 36, 1600, 105, 62, 58, 520, "item:silver:60|exp:520:100"),
    ("mob_treasure_hunter", "寻宝猎人", "mob", 38, 1700, 108, 60, 65, 560, "item:item_elixir:12|exp:560:100"),
    ("mob_space_trader", "空间商人护卫", "mob", 42, 2000, 120, 72, 68, 680, "item:item_wormhole_pass:8|exp:680:100"),
    ("mob_tianmu_energy_beast", "天目能量兽", "mob", 46, 2400, 140, 80, 72, 780, "item:core_fire:28|exp:780:100"),
    ("mob_blood_pool_cultivator", "血潭修炼者", "mob", 48, 2500, 148, 78, 75, 820, "item:item_elixir:18|exp:820:100"),
    # 魂殿
    ("mob_soul_hall_warden", "魂牢狱卒", "mob", 35, 1300, 95, 55, 58, 480, "item:item_soul_baby_fruit:8|exp:480:100"),
    ("mob_soul_extractor", "抽魂师", "mob", 40, 1800, 115, 62, 65, 640, "item:item_soul_baby_fruit:12|exp:640:100"),
    ("mob_soul_puppet", "魂傀儡", "mob", 45, 2200, 125, 75, 55, 760, "item:core_earth:25|exp:760:100"),
    ("mob_ritual_sacrificer", "祭坛献祭者", "mob", 52, 2700, 155, 90, 78, 950, "item:item_soul_baby_fruit:18|exp:950:100"),
    # 兽域
    ("mob_three_headed_snake", "三头蛇", "mob", 53, 3200, 160, 92, 82, 1050, "item:item_snake_venom_vial:20|exp:1050:100"),
    ("mob_blood_mane_lion", "血鬃狮", "mob", 54, 3400, 168, 90, 88, 1100, "item:item_beast_blood_essence:10|exp:1100:100"),
    ("mob_fire_feather_vulture", "火羽秃鹫", "mob", 56, 3300, 175, 82, 100, 1180, "item:core_fire:32|exp:1180:100"),
    ("mob_ice_scale_serpent", "冰鳞蟒", "mob", 55, 3500, 165, 100, 80, 1120, "item:core_ice:32|exp:1120:100"),
    # 莽荒古域
    ("mob_ancient_moss_beast", "古苔巨兽", "mob", 62, 4600, 195, 115, 65, 1500, "item:core_wood:32|exp:1500:100"),
    ("mob_poison_spore_ghoul", "毒孢尸", "mob", 63, 4200, 185, 90, 75, 1550, "item:herb_desert_mandala:22|exp:1550:100"),
    ("mob_ancient_beast_skeleton", "远古兽骸", "mob", 66, 5200, 200, 130, 70, 1700, "item:item_ancient_relic:8|exp:1700:100"),
    # 龙岛
    ("mob_void_serpent", "虚空蛇", "mob", 58, 3800, 200, 95, 100, 1300, "item:core_wind:32|exp:1300:100"),
    ("mob_dragon_island_wolf", "龙岛狼", "mob", 56, 3600, 190, 105, 90, 1220, "item:item_beast_hide_bundle:28|exp:1220:100"),
    # 远古战场
    ("mob_battlefield_ghost", "战场鬼火", "mob", 72, 5600, 235, 105, 105, 2150, "item:item_soul_baby_fruit:15|exp:2150:100"),
    ("mob_space_rift_beast", "空间裂缝兽", "mob", 78, 6800, 270, 135, 115, 2550, "item:item_ancient_relic:10|exp:2550:100"),
    ("mob_emperor_jade_guardian", "帝玉守护者", "mob", 80, 7200, 285, 150, 120, 2700, "item:item_ancient_relic:12|exp:2700:100"),
]

for d in REGIONAL_MOBS:
    A(d[0], d[1], d[2], d[3], d[4], d[5], d[6], d[7], d[8], d[9])

# ══════════════════════════════════════════════════════════
# MORE ELITE ENEMIES
# ══════════════════════════════════════════════════════════
ELITES = [
    # ── 加玛帝国 ──
    ("elite_wutan_guard_chief", "乌坦城守卫长", "elite", 12, 950, 62, 42, 40, 260, "item:silver:60|exp:260:100"),
    ("elite_alchemist_examiner", "炼药考核官", "elite", 16, 1200, 72, 50, 48, 380, "item:item_elixir:22|exp:380:100"),
    ("elite_salt_gang_leader", "盐帮头目", "elite", 18, 1500, 80, 55, 50, 440, "item:silver:120|exp:440:100"),
    ("elite_black_rock_gang_boss", "黑岩城帮主", "elite", 20, 1800, 88, 60, 55, 500, "item:silver:150|exp:500:100"),
    ("elite_desert_caravan_master", "沙漠驼队首领", "elite", 20, 1700, 85, 58, 58, 500, "item:item_desert_compass:12|exp:500:100"),
    ("elite_sand_pirate_chief", "沙盗头领", "elite", 22, 2100, 92, 62, 65, 580, "item:silver:180|exp:580:100"),
    ("elite_snake_temple_priest", "蛇人神殿祭司", "elite", 24, 2300, 95, 68, 62, 650, "item:item_snake_venom_vial:30|exp:650:100"),
    ("elite_imperial_court_mage", "皇室宫廷法师", "elite", 26, 2400, 100, 72, 68, 720, "item:item_elixir:28|exp:720:100"),
    ("elite_jia_ma_arena_champion", "加玛竞技场冠军", "elite", 28, 2700, 108, 75, 75, 780, "item:silver:200|exp:780:100"),
    # ── 云岚宗 ──
    ("elite_yunlan_ritual_master", "云岚宗祭师", "elite", 22, 2000, 90, 60, 62, 550, "item:core_wind:28|exp:550:100"),
    ("elite_yunlan_sword_master", "云岚宗剑师", "elite", 26, 2500, 105, 68, 72, 700, "item:core_ice:30|exp:700:100"),
    # ── 黑角域 ──
    ("elite_black_seal_auctioneer", "黑印城首席拍卖师", "elite", 26, 2300, 95, 72, 70, 680, "item:item_black_market_pass:12|exp:680:100"),
    ("elite_blood_refiner", "血炼术士", "elite", 28, 2700, 112, 78, 68, 760, "item:herb_blood_essence_fruit:30|exp:760:100"),
    ("elite_feng_city_spy", "枫城密探", "elite", 30, 2800, 115, 72, 82, 820, "item:item_forbidden_pill_fragment:10|exp:820:100"),
    ("elite_black_emperor_advisor", "黑皇宗幕僚", "elite", 32, 3000, 120, 82, 78, 900, "item:item_elixir:32|exp:900:100"),
    ("elite_plain_mercenary_chief", "平原佣兵头领", "elite", 24, 2400, 100, 68, 68, 620, "item:item_beast_hide_bundle:22|exp:620:100"),
    ("elite_eight_gates_auctioneer", "八扇门首席鉴定师", "elite", 28, 2500, 100, 70, 75, 740, "item:item_black_market_pass:15|exp:740:100"),
    # ── 迦南学院 ──
    ("elite_library_guardian", "藏书阁守阁人", "elite", 30, 2800, 105, 85, 72, 820, "item:item_canaan_token:40|exp:820:100"),
    ("elite_mission_hall_officer", "任务大厅执事", "elite", 27, 2600, 102, 76, 74, 730, "item:item_canaan_token:35|exp:730:100"),
    ("elite_inner_duel_referee", "内院裁判", "elite", 32, 3200, 118, 85, 80, 950, "item:item_canaan_token:50|exp:950:100"),
    # ── 出云帝国 ──
    ("elite_poison_craft_master", "制毒大师", "elite", 34, 3600, 135, 90, 80, 1020, "item:herb_desert_mandala:30|exp:1020:100"),
    ("elite_scorpion_tamer_master", "驭蝎大师", "elite", 35, 3700, 138, 92, 78, 1050, "item:item_snake_venom_vial:35|exp:1050:100"),
    ("elite_golden_goose_rider", "金雁骑士", "elite", 32, 3400, 130, 85, 88, 960, "item:core_wind:35|exp:960:100"),
    ("elite_mulan_valley_herbalist", "慕兰谷药师", "elite", 34, 3300, 125, 88, 82, 980, "item:herb_ancient_green_vine:18|exp:980:100"),
    # ── 魂殿高层 ──
    ("elite_soul_hall_two_star", "魂殿二星斗圣", "elite", 75, 25000, 380, 280, 220, 9000, "item:item_soul_baby_fruit:55|exp:9000:100"),
    ("elite_soul_hall_three_star", "魂殿三星斗圣", "elite", 78, 28000, 400, 300, 235, 10000, "item:item_soul_baby_fruit:60|exp:10000:100"),
    ("elite_soul_hall_five_star", "魂殿五星斗圣", "elite", 82, 34000, 450, 340, 260, 13000, "item:item_soul_baby_fruit:70|exp:13000:100"),
    # ── 中州更多精英 ──
    ("elite_thunder_pavilion_north", "风雷北阁执事", "elite", 44, 4600, 158, 98, 100, 1480, "item:core_wind:38|exp:1480:100"),
    ("elite_thunder_pavilion_east", "风雷东阁执事", "elite", 43, 4500, 155, 96, 98, 1450, "item:core_wind:38|exp:1450:100"),
    ("elite_space_trade_guard", "空间交易会护卫", "elite", 52, 5400, 182, 112, 105, 1850, "item:item_zhongzhou_map:22|exp:1850:100"),
    ("elite_dan_tower_examiner", "丹塔考核官", "elite", 54, 5200, 175, 110, 108, 1880, "item:item_dan_herb_box:18|exp:1880:100"),
    ("elite_dan_region_alchemist", "丹域炼药大师", "elite", 55, 5000, 170, 105, 112, 1900, "item:item_alchemy_handbook:20|exp:1900:100"),
    ("elite_fire_valley_smith", "焚炎谷铸器师", "elite", 46, 4800, 165, 108, 98, 1580, "item:core_fire:38|exp:1580:100"),
    ("elite_ice_valley_freezer", "冰河谷冰封师", "elite", 48, 5000, 162, 112, 95, 1650, "item:core_ice:38|exp:1650:100"),
    ("elite_beast_region_tamer", "兽域驯兽师", "elite", 54, 5800, 200, 122, 112, 1950, "item:item_transforming_herb:12|exp:1950:100"),
    ("elite_ancient_ruin_explorer", "远古遗迹探险者", "elite", 56, 6200, 210, 125, 118, 2050, "item:item_ancient_relic:15|exp:2050:100"),
    ("elite_bodhi_guardian_spirit", "菩提守护灵", "elite", 68, 7500, 260, 155, 150, 2850, "item:item_bodhi_leaf:8|exp:2850:100"),
    ("elite_nine_serene_guard", "九幽黄泉守卫", "elite", 60, 6800, 230, 145, 130, 2350, "item:item_huangquan_blood_crystal:18|exp:2350:100"),
    # ── 远古种族更多 ──
    ("elite_gu_clan_ritual_elder", "古族祭坛长老", "elite", 74, 8500, 295, 195, 160, 3300, "item:item_ancient_relic:18|exp:3300:100"),
    ("elite_yan_clan_flame_master", "炎族控火师", "elite", 68, 7600, 275, 165, 148, 2900, "item:core_fire:48|exp:2900:100"),
    ("elite_lei_clan_thunder_master", "雷族驭雷师", "elite", 68, 7700, 270, 168, 155, 2900, "item:core_wind:48|exp:2900:100"),
    ("elite_yao_clan_alchemist", "药族大药师", "elite", 66, 7200, 250, 160, 142, 2750, "item:item_dan_herb_box:28|exp:2750:100"),
    ("elite_shi_clan_warrior", "石族战士", "elite", 64, 7800, 260, 180, 130, 2700, "item:core_earth:45|exp:2700:100"),
    ("elite_ling_clan_scout", "灵族斥候", "elite", 62, 6800, 240, 140, 155, 2500, "item:item_soul_baby_fruit:32|exp:2500:100"),
    # ── 古帝洞府 ──
    ("elite_emperor_formation_guard", "古帝阵灵", "elite", 82, 9500, 320, 220, 160, 3800, "item:item_ancient_relic:20|exp:3800:100"),
    # ── 最终战场 ──
    ("elite_alliance_champion", "联盟斗圣先锋", "elite", 92, 12500, 420, 270, 200, 5500, "item:item_elixir:70|exp:5500:100"),
]

for d in ELITES:
    A(d[0], d[1], d[2], d[3], d[4], d[5], d[6], d[7], d[8], d[9])

# ══════════════════════════════════════════════════════════
# MORE BOSSES
# ══════════════════════════════════════════════════════════
BOSSES = [
    # ── 加玛帝国 ──
    ("boss_twin_head_fire_snake", "双头火灵蛇", "boss", 15, 3500, 120, 100, 85, 1000, "item:core_fire:40|exp:1000:100"),
    ("boss_sand_thief_king", "沙盗之王", "boss", 18, 3200, 135, 95, 100, 1200, "item:silver:500|item:item_desert_compass:25|exp:1200:100"),
    ("boss_guard_beast_of_temple", "神殿守护兽", "boss", 20, 4500, 145, 130, 90, 1500, "item:item_snake_venom_vial:45|exp:1500:100"),
    # ── 魔兽山脉 ──
    ("boss_fire_ape_king", "火猿王", "boss", 22, 5500, 180, 140, 120, 2000, "item:item_beast_blood_essence:30|exp:2000:100"),
    # ── 黑角域 ──
    ("boss_black_corner_arena_king", "黑角域竞技场之王", "boss", 30, 6500, 240, 160, 180, 3200, "item:silver:800|exp:3200:100"),
    ("boss_black_seal_underground_lord", "黑印城地下之主", "boss", 32, 7000, 260, 175, 185, 3500, "item:item_black_market_pass:40|exp:3500:100", "non_lethal"),
    # ── 迦南学院 ──
    ("boss_skyfire_beast_king", "天焚炼气塔兽王", "boss", 32, 8000, 250, 190, 160, 4000, "item:item_fallen_heart_flame:5|exp:4000:100"),
    # ── 中州 ──
    ("boss_space_trade_master", "空间交易会会长", "boss", 58, 20000, 420, 320, 310, 15000, "item:item_zhongzhou_map:40|exp:15000:100", "non_lethal"),
    ("boss_wan_yao_mountain_lord", "万药山脉之主", "boss", 56, 18000, 400, 300, 290, 13500, "item:herb_ancient_green_vine:40|exp:13500:100"),
    ("boss_manghuang_mayor", "莽荒镇长", "boss", 60, 22000, 430, 340, 310, 16000, "item:item_ancient_seed:25|exp:16000:100", "non_lethal"),
    ("boss_small_dan_tower_elder", "小丹塔大长老", "boss", 72, 35000, 500, 480, 430, 30000, "item:item_alchemy_handbook:50|exp:30000:100", "non_lethal"),
    ("boss_burial_sky_guardian", "葬天山脉守护者", "boss", 86, 50000, 680, 550, 500, 45000, "item:item_ancient_relic:35|exp:45000:100"),
    # ── 魂殿 ──
    ("boss_soul_hall_vice_leader", "魂殿副殿主", "boss", 78, 38000, 580, 450, 430, 32000, "item:item_soul_baby_fruit:75|exp:32000:100"),
    ("boss_soul_hall_three_elders", "魂殿三天尊", "boss", 76, 35000, 550, 430, 420, 30000, "item:item_soul_baby_fruit:70|exp:30000:100"),
    # ── 远古种族 ──
    ("boss_gu_clan_three_immortals", "古族三仙", "boss", 85, 48000, 620, 520, 500, 42000, "item:item_ancient_relic:40|exp:42000:100", "non_lethal"),
    ("boss_yan_clan_flame_emperor", "炎族炎帝", "boss", 86, 50000, 650, 530, 510, 44000, "item:item_emperor_flame:5|exp:44000:100", "non_lethal"),
    ("boss_lei_clan_thunder_emperor", "雷族雷帝", "boss", 86, 51000, 640, 535, 520, 44000, "item:core_wind:80|exp:44000:100", "non_lethal"),
    # ── 龙岛 ──
    ("boss_west_dragon_king", "西龙王", "boss", 66, 26000, 510, 410, 390, 23000, "item:item_dragon_scale:28|exp:23000:100", "non_lethal"),
    # ── 古帝洞府 ──
    ("boss_tuoshe_emperor_remnant", "陀舍古帝残影", "boss", 92, 70000, 800, 600, 580, 65000, "item:item_emperor_flame:15|exp:65000:100", "non_lethal"),
    # ── 最终 ──
    ("boss_five_emperor_gate_guard", "五帝破空之门守卫", "boss", 95, 80000, 900, 700, 650, 80000, "item:item_emperor_flame:20|exp:80000:100"),
]

for d in BOSSES:
    if len(d) > 10:
        A(d[0], d[1], d[2], d[3], d[4], d[5], d[6], d[7], d[8], d[9], d[10])
    else:
        A(d[0], d[1], d[2], d[3], d[4], d[5], d[6], d[7], d[8], d[9])

print(f"Added: {count} new enemies. Total: {len(existing)}")
wb.save(WORKBOOK_PATH)
